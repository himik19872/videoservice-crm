"""
Конвертеры Excel-файлов в унифицированный CSV-формат.

Поддерживаемые форматы:
1. ТСЖ (база клиентов) — 7 колонок
2. ЕРЦ СПб (форма № 30.01.01) — 15 колонок
3. ЕРЦ ЛО (форма № 30.01.01в) — 17 колонок
4. ЕРЦ Агалатово (форма № 03.01.04б) — 19 колонок
5. Красное Село (оборотно-сальдовая) — 12 колонок
6. Стр.6-3 (ТО домофонов) — 8 колонок

Унифицированный CSV (разделитель — запятая, кавычки — двойные):
  type,personal_account,full_name,city,district,street,house,building,apartment,entrance,management_company,period,balance_start,charged,paid,balance_end,raw_address

  type = "client" | "payment"
"""

import csv
import io
import re
from datetime import datetime
from pathlib import Path
import openpyxl


# ══════════════════════════════════════════════════════════════════
# Общие утилиты
# ══════════════════════════════════════════════════════════════════

SPB_SUBURBS = {
    'петергоф': 'Петергоф', 'петродворец': 'Петергоф',
    'ломоносов': 'Ломоносов', 'колпино': 'Колпино',
    'пушкин': 'Пушкин', 'царское село': 'Пушкин',
    'павловск': 'Павловск', 'кронштадт': 'Кронштадт',
    'сестрорецк': 'Сестрорецк', 'зеленогорск': 'Зеленогорск',
    'красное село': 'Красное Село', 'гатчина': 'Гатчина',
    'стрельна': 'Стрельна', 'шушары': 'Шушары',
    'парголово': 'Парголово', 'левашово': 'Левашово',
    'репино': 'Репино', 'комарово': 'Комарово',
    'солнечное': 'Солнечное', 'белоостров': 'Белоостров',
    'серово': 'Серово', 'усть-ижора': 'Усть-Ижора',
    'понтонный': 'Понтонный', 'металлострой': 'Металлострой',
    'сапёрный': 'Сапёрный', 'петро-славянка': 'Петро-Славянка',
    'рощино': 'Рощино', 'молодёжное': 'Молодёжное',
    'сосново': 'Сосново', 'всеволожск': 'Всеволожск',
    'тосно': 'Тосно', 'аннино': 'Аннино',
    'агалатово': 'Агалатово', 'коммунар': 'Коммунар',
    'сертолово': 'Сертолово', 'мурино': 'Мурино',
}

UNIFIED_FIELDS = [
    'type', 'personal_account', 'full_name',
    'city', 'district', 'street', 'house', 'building', 'apartment', 'entrance',
    'management_company',
    'period', 'balance_start', 'charged', 'paid', 'balance_end',
    'raw_address',
]


def _clean_str(val):
    """Очистка строкового значения."""
    if val is None:
        return ''
    s = str(val).strip()
    if s in ('None', '-', '—', ''):
        return ''
    return s


def _parse_house_number(h):
    """Исправляет кривые номера домов: д..5 → 5, д. 37А → 37А."""
    h = _clean_str(h)
    h = re.sub(r'д\.\.+', '', h).strip()
    h = re.sub(r'^д\.?\s*', '', h).strip()
    return h


def parse_address(raw):
    """
    Парсит адресную строку в поля: city, district, street, house, building, apartment.
    Возвращает dict.
    """
    result = {
        'city': '', 'district': '', 'street': '',
        'house': '', 'building': '', 'apartment': '',
    }

    if not raw or not raw.strip():
        return result

    text = raw.strip()

    # 1. Выделяем скобки
    bracket_match = re.search(r'\(([^)]+)\)', text)
    bracket_text = ''
    if bracket_match:
        bracket_text = bracket_match.group(1).strip().lower()
        text = text.replace(bracket_match.group(0), '').strip()
        text = re.sub(r',\s*,', ',', text)
        text = re.sub(r'\s{2,}', ' ', text)

    # 2. Разбиваем по запятым
    parts = [p.strip() for p in text.split(',') if p.strip()]

    # 3. Определяем город
    city_found = ''
    if parts:
        first = parts[0].lower()
        if any(c in first for c in ['санкт-петербург', 'спб']):
            city_found = 'Санкт-Петербург'
            parts = parts[1:]
        elif 'москва' in first:
            city_found = 'Москва'
            parts = parts[1:]

    # Скобки → пригород/район
    if bracket_text:
        for key, val in SPB_SUBURBS.items():
            if key in bracket_text:
                city_found = val
                result['district'] = val
                break
        if not result['district']:
            result['district'] = bracket_text.title()

    # Части → пригород
    for i, p in enumerate(parts):
        for key, val in SPB_SUBURBS.items():
            if key in p.lower():
                if not city_found or city_found == 'Санкт-Петербург':
                    city_found = val
                if not result['district']:
                    result['district'] = val
                parts[i] = re.sub(r'(?i)\b' + re.escape(key) + r'\b\s*', '', p).strip()
                if not parts[i]:
                    parts.pop(i)
                break
        if city_found and city_found != 'Санкт-Петербург':
            break

    result['city'] = city_found or 'Санкт-Петербург'

    # 4. Ищем улицу (по типу улицы)
    street_types = ['ул.', ' ул', 'улица', 'пр-кт', 'пр.', 'пр ', 'проспект',
                    'пер.', 'пер ', 'переулок', 'наб.', 'наб ', 'набережная',
                    'шоссе', 'ш.', 'бульвар', 'б-р', 'аллея', 'проезд',
                    'пл.', 'площадь', 'линия', 'дорога', 'тупик', 'тракт']

    street_idx = -1
    for i, p in enumerate(parts):
        p_lower = p.lower()
        for st in street_types:
            if st in p_lower or p_lower.startswith(st.replace('.', '')):
                street_idx = i
                break
        if street_idx >= 0:
            break

    if street_idx >= 0:
        result['street'] = parts[street_idx]
    elif parts:
        # Специальный случай: "Агалатово д 15" → город=Агалатово, дом=15
        # Ищем "д ХХ" в первом remaining-part
        for i, p in enumerate(parts):
            house_match = re.search(r'д\.?\s*(\d+[а-яa-z]*)', p.lower())
            if house_match:
                result['house'] = _parse_house_number('д ' + house_match.group(1))
                # Всё до "д" — возможно улица или ничего
                before = p[:house_match.start()].strip().rstrip(', ').strip()
                if before and not any(c in before.lower() for c in SPB_SUBURBS):
                    result['street'] = before
                street_idx = i
                break
        if street_idx < 0:
            for i, p in enumerate(parts):
                if not re.search(r'д\.\s*\d|дом\s*\d|корп|кв\.?\s*\d|^\d+$', p.lower()):
                    result['street'] = p
                    street_idx = i
                    break

    # 5. Дом, корпус, квартира
    remaining = parts[max(street_idx + 1, 0):]

    for p in remaining:
        p_clean = p.strip().rstrip(',').strip()

        # Литера
        lit_match = re.search(r'лит\.?\s*([а-яa-z])', p_clean, re.IGNORECASE)
        if lit_match:
            result['building'] = f'литера {lit_match.group(1).upper()}'
            p_clean = re.sub(r'лит\.?\s*[а-яa-z]', '', p_clean, flags=re.IGNORECASE).strip()

        # Корпус/строение
        corp_match = re.search(r'(?:корп|к|стр)\.?\s*(\d+[а-яa-z]?)', p_clean, re.IGNORECASE)
        if corp_match:
            corp_val = corp_match.group(1)
            if result['building']:
                result['building'] += f' корп. {corp_val}'
            else:
                result['building'] = corp_val
            p_clean = re.sub(r'(?:корп|к|стр)\.?\s*\d+[а-яa-z]?', '', p_clean, flags=re.IGNORECASE).strip()

        # Дом
        house_match = re.search(r'д\.?\.?\s*(\d+[а-яa-z]*)', p_clean, re.IGNORECASE)
        if house_match and not result['house']:
            result['house'] = _parse_house_number(house_match.group(0))
            p_clean = re.sub(r'д\.?\.?\s*\d+[а-яa-z]*', '', p_clean, flags=re.IGNORECASE).strip()

        # Дом без префикса
        if not result['house']:
            num_match = re.search(r'^(\d+[а-яa-z]?)$', p_clean.strip().rstrip(','))
            if num_match:
                result['house'] = num_match.group(1)
                continue

        # Квартира: последнее число (если не дом и не корпус)
        apt_match = re.search(r'(\d+)\s*$', p_clean.strip().rstrip(','))
        if apt_match and result['house'] and not result['apartment']:
            apt_val = apt_match.group(1)
            # Исключаем если это не квартира (например "ТСЖ")
            remaining_before = p_clean[:apt_match.start()].strip().rstrip(',').strip()
            if not remaining_before or remaining_before.isdigit() or not any(
                w in remaining_before.lower() for w in ['тсж', 'ук ', 'жск', 'жк ', 'ооо', 'тсн']
            ):
                result['apartment'] = apt_val

    return result


def make_unified_row(row_type, **kwargs):
    """Создаёт строку унифицированного формата."""
    row = {f: '' for f in UNIFIED_FIELDS}
    row['type'] = row_type
    row.update(kwargs)
    return row


def write_csv(rows, filepath):
    """Записывает строки в CSV-файл."""
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=UNIFIED_FIELDS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


# ══════════════════════════════════════════════════════════════════
# Конвертер №1: ТСЖ (база клиентов)
# ══════════════════════════════════════════════════════════════════

def convert_tszh(file_bytes):
    """
    Конвертирует файл ТСЖ (7 колонок: №п/п, л/с, ФИО, Адрес, №парадной, ТСЖ).
    → CSV с type=client.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_data = list(ws.iter_rows(min_row=2, values_only=True))

    result = []
    for row in rows_data:
        if not row:
            continue

        personal_account = _clean_str(row[1]) if len(row) > 1 else ''
        full_name = _clean_str(row[2]) if len(row) > 2 else ''
        raw_address = _clean_str(row[3]) if len(row) > 3 else ''
        entrance = _clean_str(row[4]) if len(row) > 4 else ''
        management_company = _clean_str(row[5]) if len(row) > 5 else ''

        if not personal_account and not full_name and not raw_address:
            continue

        parsed = parse_address(raw_address)

        # entrance может быть пустым в колонке, но квартирой в адресе
        if not entrance and parsed['apartment']:
            entrance = parsed['apartment']  # fallback

        result.append(make_unified_row('client',
            personal_account=personal_account,
            full_name=full_name or 'Не определено',
            city=parsed['city'],
            district=parsed['district'],
            street=parsed['street'],
            house=parsed['house'],
            building=parsed['building'],
            apartment=parsed['apartment'],
            entrance=entrance,
            management_company=management_company,
            raw_address=raw_address,
        ))

    return result


# ══════════════════════════════════════════════════════════════════
# Конвертер №2: ЕРЦ СПб (форма № 30.01.01)
# ══════════════════════════════════════════════════════════════════

def convert_erc_spb(file_bytes, period_date=None):
    """
    ЕРЦ СПб: 15 колонок, заголовки в строках 6-9, данные с 10-й строки.
    Колонки: 0=№п/п, 1=л/с, 2=ФИО, 5=Адрес, 7=жильцов, 8=сальдо на начало,
             9=начислено, 10=начислено факт, 11=оплачено, 12=%оплаты, 13=сальдо дебет, 14=сальдо кредит
    """
    if period_date is None:
        period_date = datetime.now().date().replace(day=1)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Ищем первую строку с данными: колонка 0 — число (№ п/п), колонка 1 — л/с
    rows_data = []
    for row in all_rows:
        if not row:
            continue
        # Первая колонка должна быть числом (порядковый номер)
        try:
            int(str(row[0]).strip())
            acc_num = _clean_str(row[1])
            if acc_num and acc_num not in ('счета', 'лицевого', 'п/п'):
                rows_data.append(row)
        except (ValueError, TypeError):
            continue

    result = []

    for row in rows_data:
        if not row:
            continue

        acc_num = _clean_str(row[1]) if len(row) > 1 else ''
        if not acc_num or acc_num in ('счета', 'лицевого', 'п/п'):
            continue

        full_name = _clean_str(row[2]) if len(row) > 2 else ''
        raw_address = _clean_str(row[5]) if len(row) > 5 else ''

        def _f(val):
            """Конвертирует в float, возвращает 0 при None/пусто/ошибке."""
            if val is None or val == '':
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        saldo_start = _f(row[8]) if len(row) > 8 else 0.0
        accrued = _f(row[9]) if len(row) > 9 else 0.0
        paid = _f(row[11]) if len(row) > 11 else 0.0
        saldo_end_debet = _f(row[13]) if len(row) > 13 else 0.0
        saldo_end_credit = _f(row[14]) if len(row) > 14 else 0.0
        saldo_end = saldo_end_debet - saldo_end_credit

        parsed = parse_address(raw_address)

        result.append(make_unified_row('payment',
            personal_account=acc_num,
            full_name=full_name,
            city=parsed['city'],
            district=parsed['district'],
            street=parsed['street'],
            house=parsed['house'],
            building=parsed['building'],
            apartment=parsed['apartment'],
            management_company='',
            period=str(period_date),
            balance_start=str(saldo_start),
            charged=str(accrued),
            paid=str(paid),
            balance_end=str(saldo_end),
            raw_address=raw_address,
        ))

    return result


# ══════════════════════════════════════════════════════════════════
# Конвертер №3: ЕРЦ ЛО (форма № 30.01.01в)
# ══════════════════════════════════════════════════════════════════

def convert_erc_lo(file_bytes, period_date=None):
    """
    ЕРЦ ЛО: 17 колонок. Отличие — колонки сдвинуты:
    0=№п/п, 1=л/с, 2=ФИО, 5=Адрес, 6=сальдо нач дебет, 7=сальдо нач кредит,
    8=начислено, 9=льгота, 10=перерасчет, 11=всего начислено, 12=оплата,
    14=сальдо кон дебет, 15=сальдо кон кредит
    """
    if period_date is None:
        period_date = datetime.now().date().replace(day=1)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Ищем строки с данными: колонка 0 — число
    rows_data = []
    for row in all_rows:
        if not row:
            continue
        try:
            int(str(row[0]).strip())
            acc_num = _clean_str(row[1])
            if acc_num and acc_num not in ('счета', 'лицевого', 'п/п'):
                rows_data.append(row)
        except (ValueError, TypeError):
            continue

    result = []

    for row in rows_data:
        if not row:
            continue

        acc_num = _clean_str(row[1]) if len(row) > 1 else ''
        if not acc_num or acc_num in ('счета', 'лицевого'):
            continue

        full_name = _clean_str(row[2]) if len(row) > 2 else ''
        raw_address = _clean_str(row[5]) if len(row) > 5 else ''

        def _f(val):
            """Конвертирует в float, возвращает 0 при None/пусто/ошибке."""
            if val is None or val == '':
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        saldo_start = _f(row[6]) - _f(row[7]) if len(row) > 7 else 0
        accrued = _f(row[8]) if len(row) > 8 else 0
        paid = _f(row[12]) if len(row) > 12 else 0
        saldo_end = _f(row[14]) - _f(row[15]) if len(row) > 15 else 0

        parsed = parse_address(raw_address)

        result.append(make_unified_row('payment',
            personal_account=acc_num,
            full_name=full_name,
            city=parsed['city'],
            district=parsed['district'],
            street=parsed['street'],
            house=parsed['house'],
            building=parsed['building'],
            apartment=parsed['apartment'],
            period=str(period_date),
            balance_start=str(saldo_start),
            charged=str(accrued),
            paid=str(paid),
            balance_end=str(saldo_end),
            raw_address=raw_address,
        ))

    return result


# ══════════════════════════════════════════════════════════════════
# Конвертер №4: ЕРЦ Агалатово (форма № 03.01.04б)
# ══════════════════════════════════════════════════════════════════

def convert_erc_agalatovo(file_bytes, period_date=None):
    """
    Агалатово: 19 колонок, адрес разбит на отдельные колонки.
    Структура:
      0=№п/п, 1=л/с, 2=ФИО,
      5=Населённый пункт, 6=Улица, 7=Дом, 8=Кв.,
      9=Сальдо нач, 10=Сальдо пени нач,
      11=Начислено, 12=Начислено пени,
      13=Оплачено, 14=Оплачено пени,
      15=Сальдо кон, 18=Сальдо пени кон
    Данные с 14-й строки (после заголовков).
    """
    if period_date is None:
        period_date = datetime.now().date().replace(day=1)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Ищем строку с '№', '№' в колонках 0 и 1 (строка 9 в файле)
    data_start = 14
    for i, row in enumerate(all_rows):
        if row and len(row) > 1 and str(row[0]).strip() == '№' and '№' in str(row[1]):
            data_start = i + 6  # +5 строк подзаголовков, i 0-based → +6 для 1-based
            break

    rows_data = all_rows[data_start - 1:]
    result = []

    for row in rows_data:
        if not row:
            continue

        acc_num = _clean_str(row[1]) if len(row) > 1 else ''
        if not acc_num or acc_num in ('счета', 'лицевого'):
            continue

        full_name = _clean_str(row[2]) if len(row) > 2 else ''

        # Адрес в отдельных колонках: населённый пункт, улица, дом, кв.
        city_part = _clean_str(row[5]) if len(row) > 5 else ''
        street_part = _clean_str(row[6]) if len(row) > 6 else ''
        house_part = _clean_str(row[7]) if len(row) > 7 else ''
        apartment_part = _clean_str(row[8]) if len(row) > 8 else ''

        # Определяем город: Агалатово, Агалатово д (деревня) и т.д.
        city_clean = city_part
        # Убираем суффиксы "д", "п" (деревня, посёлок) для города
        city_clean = re.sub(r'\s+[дп]$', '', city_clean).strip()
        if not city_clean:
            city_clean = city_part

        # Собираем raw_address и parsed из частей
        raw_address_parts = []
        if city_part:
            raw_address_parts.append(city_part)
        if street_part and street_part not in ('-', ''):
            raw_address_parts.append(street_part)
        if house_part and house_part not in ('-', ''):
            raw_address_parts.append(f'д. {house_part}')
        if apartment_part and apartment_part not in ('-', ''):
            raw_address_parts.append(f'кв. {apartment_part}')
        raw_address = ', '.join(raw_address_parts) if raw_address_parts else city_part

        # Используем части напрямую, без parse_address
        parsed = {
            'city': city_clean,
            'district': '',
            'street': street_part if street_part not in ('-', '') else '',
            'house': house_part if house_part not in ('-', '') else '',
            'building': '',
            'apartment': apartment_part if apartment_part not in ('-', '') else '',
        }

        def _f(val):
            """Конвертирует в float, возвращает 0 при None/пусто/ошибке."""
            if val is None or val == '':
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        saldo_start = _f(row[9]) if len(row) > 9 else 0
        accrued = _f(row[11]) if len(row) > 11 else 0
        paid = _f(row[13]) if len(row) > 13 else 0
        saldo_end = _f(row[15]) if len(row) > 15 else 0

        result.append(make_unified_row('payment',
            personal_account=acc_num,
            full_name=full_name,
            city=parsed['city'] or city_part,
            district=parsed['district'],
            street=parsed['street'] or street_part,
            house=parsed['house'] or house_part,
            building=parsed['building'],
            apartment=parsed['apartment'] or apartment_part,
            period=str(period_date),
            balance_start=str(saldo_start),
            charged=str(accrued),
            paid=str(paid),
            balance_end=str(saldo_end),
            raw_address=raw_address,
        ))

    return result


# ══════════════════════════════════════════════════════════════════
# Конвертер №5: Красное Село (оборотно-сальдовая)
# ══════════════════════════════════════════════════════════════════

def convert_krasnoe_selo(file_bytes, period_date=None):
    """
    Красное Село: 12 колонок, адрес дома указан в строке 5,
    квартиры с 8-й строки: 0=кв, 3=ФИО, 4=начДт, 5=начКт, 6=начислено, 8=оплата, 9=конДт, 10=конКт
    """
    if period_date is None:
        period_date = datetime.now().date().replace(day=1)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Адрес дома: строка 5, колонка 1  (формат: "ул.Гатчинское д.9 к.1")
    building_address = ''
    if len(all_rows) > 4 and all_rows[4]:
        building_address = _clean_str(all_rows[4][1])

    # Парсим адрес дома — формат "ул.Гатчинское д.9 к.1"
    # Разделяем слипшиеся улица+дом+корпус
    building_parsed = {}
    if building_address:
        # Нормализуем: "ул.Гатчинское д.9 к.1" → "ул. Гатчинское, д. 9, корп. 1"
        fixed = building_address
        fixed = re.sub(r'\bул\.\s*', 'ул. ', fixed)
        fixed = re.sub(r'\bд\.\s*(\d+)', r', д. \1', fixed)
        fixed = re.sub(r'\bк\.\s*(\d+)', r', корп. \1', fixed)
        fixed = re.sub(r'\s+', ' ', fixed).strip()
        fixed = re.sub(r'\s*,\s*', ', ', fixed)
        fixed = fixed.lstrip(', ')
        # Убираем дубли: "ул. ул." → "ул."
        fixed = re.sub(r'\bул\.\s+ул\.', 'ул.', fixed)
        building_parsed = parse_address(fixed)

    # Данные: с 8-й строки
    rows_data = all_rows[7:]
    result = []

    for row in rows_data:
        if not row:
            continue

        apartment = _clean_str(row[0]) if len(row) > 0 else ''
        full_name = _clean_str(row[3]) if len(row) > 3 else ''

        if not apartment and not full_name:
            continue

        def _f(val):
            """Конвертирует в float, возвращает 0 при None/пусто/ошибке."""
            if val is None or val == '':
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        saldo_start = _f(row[4]) - _f(row[5]) if len(row) > 5 else 0
        accrued = _f(row[6]) if len(row) > 6 else 0
        paid = _f(row[8]) if len(row) > 8 else 0
        saldo_end = _f(row[9]) - _f(row[10]) if len(row) > 10 else 0

        result.append(make_unified_row('payment',
            personal_account='',
            full_name=full_name,
            city=building_parsed.get('city', ''),
            district=building_parsed.get('district', ''),
            street=building_parsed.get('street', ''),
            house=building_parsed.get('house', ''),
            building=building_parsed.get('building', ''),
            apartment=apartment,
            period=str(period_date),
            balance_start=str(saldo_start),
            charged=str(accrued),
            paid=str(paid),
            balance_end=str(saldo_end),
            raw_address=building_address,
        ))

    return result


# ══════════════════════════════════════════════════════════════════
# Конвертер №6: Стр.6-3 (ТО домофонов)
# ══════════════════════════════════════════════════════════════════

def convert_str63(file_bytes, period_date=None):
    """
    Стр.6-3: 8 колонок. 0=№п/п, 1=л/с, 2=ФИО, 3=Адрес,
    4=вх.сальдо, 5=начислено, 6=платежи, 7=исх.сальдо
    Адрес в формате: "Красное Село,Стрельнинское,д.6 к.3кв.1"
    """
    if period_date is None:
        period_date = datetime.now().date().replace(day=1)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_data = list(ws.iter_rows(min_row=2, values_only=True))

    result = []
    for row in rows_data:
        if not row:
            continue

        acc_num = _clean_str(row[1]) if len(row) > 1 else ''
        full_name = _clean_str(row[2]) if len(row) > 2 else ''
        raw_address = _clean_str(row[3]) if len(row) > 3 else ''

        if not raw_address and not full_name:
            continue

        # Исправляем адрес: "д.6 к.3кв.1" → "д. 6, корп. 3, кв. 1"
        raw_address = re.sub(r'к\.(\d+)кв', r'корп. \1, кв.', raw_address)
        raw_address = re.sub(r'д\.(\d+)', r'д. \1', raw_address)
        raw_address = re.sub(r'кв\.\.', 'кв. ', raw_address)  # фикс двойной точки
        raw_address = re.sub(r'\.{2,}', '.', raw_address)  # все двойные точки → одна
        raw_address = re.sub(r',\s*,', ', ', raw_address)
        raw_address = re.sub(r'\s{2,}', ' ', raw_address)

        def _f(val):
            """Конвертирует в float, возвращает 0 при None/пусто/ошибке."""
            if val is None or val == '':
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        balance_start = _f(row[4]) if len(row) > 4 else 0
        charged = _f(row[5]) if len(row) > 5 else 0
        paid = _f(row[6]) if len(row) > 6 else 0
        balance_end = _f(row[7]) if len(row) > 7 else 0

        parsed = parse_address(raw_address)

        result.append(make_unified_row('payment',
            personal_account=acc_num,
            full_name=full_name,
            city=parsed['city'],
            district=parsed['district'],
            street=parsed['street'],
            house=parsed['house'],
            building=parsed['building'],
            apartment=parsed['apartment'],
            period=str(period_date),
            balance_start=str(balance_start),
            charged=str(charged),
            paid=str(paid),
            balance_end=str(balance_end),
            raw_address=raw_address,
        ))

    return result


# ══════════════════════════════════════════════════════════════════
# Главная функция конвертации (автоопределение формата)
# ══════════════════════════════════════════════════════════════════

def auto_convert(file_bytes, period_date=None):
    """
    Автоматически определяет формат файла и конвертирует в унифицированные строки.
    Возвращает (rows, format_name).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))

    # Собираем текст первых строк для анализа
    header_text = ' '.join(
        ' '.join(str(c) for c in row if c is not None)
        for row in all_rows if row
    ).lower()

    # Красное Село (оборотно-сальдовая)
    if 'оборотно-сальдовая' in header_text:
        return convert_krasnoe_selo(file_bytes, period_date), 'krasnoe_selo'

    # ЕРЦ ЛО (форма 03.01.04б — Агалатово)
    if '03.01.04б' in header_text:
        return convert_erc_agalatovo(file_bytes, period_date), 'erc_agalatovo'

    # ЕРЦ ЛО (форма 30.01.01в)
    if '30.01.01в' in header_text:
        return convert_erc_lo(file_bytes, period_date), 'erc_lo'

    # ЕРЦ СПб (форма 30.01.01)
    if '30.01.01' in header_text and 'еирц спб' in header_text:
        return convert_erc_spb(file_bytes, period_date), 'erc_spb'

    # ЕРЦ Коммунар (форма 30.01.01, 15 колонок — как СПб)
    if '30.01.01' in header_text:
        return convert_erc_spb(file_bytes, period_date), 'erc_spb'

    # Стр.6-3: "Тех.обслуж.домофонов" в заголовках
    if 'тех.обслуж.домофонов' in header_text or 'домофон' in header_text:
        return convert_str63(file_bytes, period_date), 'str63'

    # ТСЖ: "№ п/п", "№ лицевого счета", "ФИО", "Адрес"
    if '№ лицевого счета' in header_text or '№ п/п' in header_text:
        return convert_tszh(file_bytes), 'tszh'

    # По умолчанию — пробуем как ТСЖ
    return convert_tszh(file_bytes), 'tszh'


# ══════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════

def convert_file(input_path, output_path=None, period_date=None):
    """
    Конвертирует один файл в унифицированный CSV.
    
    Args:
        input_path: путь к .xlsx файлу
        output_path: путь к выходному .csv (если None — заменяет .xlsx на .csv)
        period_date: дата периода для ЕРЦ (по умолчанию — 1-е число текущего месяца)
    
    Returns:
        dict: {rows, format_name, output_path}
    """
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.with_suffix('.csv')

    with open(input_path, 'rb') as f:
        file_bytes = f.read()

    rows, fmt = auto_convert(file_bytes, period_date)
    count = write_csv(rows, str(output_path))

    print(f'Конвертировано: {input_path.name} → {output_path.name}')
    print(f'  Формат: {fmt}')
    print(f'  Строк:  {count}')

    return {'rows': count, 'format': fmt, 'output_path': str(output_path)}


def convert_directory(dir_path, output_dir=None, period_date=None):
    """
    Конвертирует все .xlsx файлы в директории.
    """
    dir_path = Path(dir_path)
    if output_dir is None:
        output_dir = dir_path / 'converted'
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    results = []
    for f in sorted(dir_path.glob('*.xlsx')):
        out = output_dir / f.with_suffix('.csv').name
        r = convert_file(str(f), str(out), period_date)
        results.append(r)
        print()

    total = sum(r['rows'] for r in results)
    print(f'Всего сконвертировано: {len(results)} файлов, {total} строк')
    return results
