"""
Импорт данных Beward из двух Excel-файлов (data_only=True).
1. адрес панели-ipвсе.xlsx — справочник IP-адресов
2. Все коды Бевард-Спутник-Рубитек.xlsx — коды в BuildingEntrance + BewardDevice
"""
import openpyxl, re
from django.core.management.base import BaseCommand
from django.utils import timezone
from main.models import BewardDevice, BuildingEntrance, Building


def parse_entrance_number(val):
    """Парсит номер подъезда: '1.0' -> 1, '1' -> 1, формула -> цифры."""
    if val is None:
        return None
    s = str(val).strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        nums = re.findall(r'\d+', s)
        return int(nums[0]) if nums else None


def extract_address_parts(addr):
    """Извлекает (street_name, house_number, building_number) из адреса."""
    if not addr:
        return '', '', ''
    addr = str(addr).strip()

    # Нормализация: добавляем пробелы после запятых, где их нет
    addr = re.sub(r',(?!\s)', ', ', addr)
    # Убираем двойные запятые, лишние пробелы
    addr = re.sub(r',\s*,', ',', addr)
    addr = re.sub(r'\s+', ' ', addr)

    # Короткий формат: "Александровская ул., д. 31"
    if not addr.lower().startswith('санкт-петербург'):
        m = re.search(r'^(.+?)\s*,?\s*д\.\s*(\d+[а-яА-Я/]?\d*[а-яА-Я]?)', addr)
        if m:
            return m.group(1).strip(), m.group(2), ''
        return addr, '', ''

    # Полный формат: убираем "Санкт-Петербург, "
    addr = addr[len('Санкт-Петербург, '):]
    # Убираем "город X, ", "поселок X, ", "территория X, "
    addr = re.sub(r'(?:город|поселок|посёлок|деревня|территория|микрорайон)\s+[^,]+,?\s*', '', addr)

    parts = [p.strip() for p in addr.split(',')]
    house = ''
    building = ''
    street_parts = []
    found_house = False

    for p in parts:
        if not found_house:
            # "дом 49", "дсм 37", "д..10", "д. 31" или "улица Лермонтова дом 7"
            m = re.search(r'(?:дом|дсм|д\.\.|д\.)\s*(\d+[а-яА-Я/]?\d*[а-яА-Я]?)', p.lower())
            if m:
                house = m.group(1)
                # Улица — всё до слова «дом» в этой же части
                before_house = p[:m.start()].strip()
                if before_house:
                    street_parts.append(before_house)
                found_house = True
                continue
            # Адрес без "дом": просто число в конце
            if re.match(r'^\d+[а-яА-Я/]?\d*[а-яА-Я]?$', p):
                house = p
                found_house = True
                continue
            street_parts.append(p)
        else:
            if 'корпус' in p.lower() or 'корп' in p.lower() or 'литер' in p.lower():
                m = re.search(r'(\d+[а-яА-Я]?)', p)
                if m:
                    building = m.group(1)

    street = ' '.join(street_parts).strip()
    for prefix in ['улица ', 'проспект ', 'переулок ', 'шоссе ', 'бульвар ',
                    'площадь ', 'аллея ', 'набережная ', 'территория ', 'дорога ',
                    'пр-кт ', 'пр. ']:
        if street.lower().startswith(prefix.lower()):
            street = street[len(prefix):]
            break

    return street.strip(), house, building


def find_building(street_name, house_number, building_number=''):
    """Ищет Building по улице, дому, корпусу."""
    if not street_name or not house_number:
        return None

    qs = Building.objects.filter(house_number=house_number)

    # 1. Прямой icontains
    candidates = qs.filter(street_name__icontains=street_name)

    # 2. По первому слову
    if not candidates.exists() and street_name.split():
        candidates = qs.filter(street_name__icontains=street_name.split()[0])

    # 3. Обратный: улица из БД содержится в нашей
    if not candidates.exists():
        for b in qs:
            if b.street_name and b.street_name.lower() in street_name.lower():
                candidates = qs.filter(id=b.id)
                break

    # 4. Наша улица содержится в БД-улице
    if not candidates.exists():
        for b in qs:
            if street_name.lower() in (b.street_name or '').lower():
                candidates = qs.filter(id=b.id)
                break

    if candidates.exists() and building_number:
        with_b = candidates.filter(building_number=building_number)
        if with_b.exists():
            return with_b.first()
        with_b = candidates.filter(building_number__icontains=building_number)
        if with_b.exists():
            return with_b.first()

    return candidates.first() if candidates.exists() else None


class Command(BaseCommand):
    help = 'Импорт IP-адресов и кодов Beward из Excel'

    def add_arguments(self, parser):
        parser.add_argument('--ip-file', default='/tmp/ip_addresses.xlsx')
        parser.add_argument('--codes-file', default='/tmp/codes.xlsx')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        ip_file = options['ip_file']
        codes_file = options['codes_file']
        dry_run = options['dry_run']

        # === Этап 1: IP-адреса -> BewardDevice ===
        self.stdout.write('=== Этап 1: IP-адреса ===')
        wb1 = openpyxl.load_workbook(ip_file, data_only=True)
        ws1 = wb1['Лист1']

        ip_total = ip_no_match = 0
        for row in ws1.iter_rows(min_row=2, values_only=True):
            region = str(row[0]).strip() if row[0] else ''
            address = str(row[1]).strip() if row[1] else ''
            entrance_raw = str(row[2]).strip() if row[2] else ''
            ip_addr = str(row[3]).strip() if row[3] else ''

            if not ip_addr or not address:
                continue

            ent_num = parse_entrance_number(entrance_raw)
            street, house, bldg = extract_address_parts(address)
            building = find_building(street, house, bldg)

            if dry_run:
                ip_total += 1
                continue

            BewardDevice.objects.update_or_create(
                ip_address=ip_addr,
                defaults={
                    'region': region,
                    'address': address,
                    'entrance_number': str(ent_num) if ent_num else entrance_raw,
                    'building': building,
                }
            )
            ip_total += 1
            if building is None:
                ip_no_match += 1

        self.stdout.write(f'  BewardDevice: {ip_total}')
        self.stdout.write(f'  Без привязки к Building: {ip_no_match}')

        # === Этап 2: Коды -> BuildingEntrance + BewardDevice ===
        self.stdout.write('\n=== Этап 2: Коды в подъезды ===')
        wb2 = openpyxl.load_workbook(codes_file, data_only=True)
        ws2 = wb2['Лист1']

        ent_created = ent_updated = dev_updated = no_match = 0

        for row in ws2.iter_rows(min_row=2, values_only=True):
            region = str(row[2]).strip() if row[2] else ''
            address = str(row[3]).strip() if row[3] else ''
            entrance_raw = str(row[4]).strip() if row[4] else ''
            apartment_range = str(row[5]).strip() if row[5] else ''
            access_code = str(row[6]).strip() if row[6] else ''
            ip_addr = str(row[7]).strip() if row[7] else ''
            programming_code = str(row[8]).strip() if row[8] else ''
            notes = str(row[9]).strip() if row[9] else ''
            date_issued = row[1]

            if not address:
                continue

            ent_num = parse_entrance_number(entrance_raw)
            street, house, bldg = extract_address_parts(address)
            building = find_building(street, house, bldg)

            if dry_run:
                if ent_num and building:
                    ent_created += 1
                else:
                    no_match += 1
                continue

            entrance = None
            if building and ent_num:
                entrance, created = BuildingEntrance.objects.get_or_create(
                    building=building,
                    number=ent_num,
                    defaults={'apartments_count': 0}
                )
                entrance.ip_address = ip_addr or entrance.ip_address
                entrance.access_code = access_code or entrance.access_code
                entrance.programming_code = programming_code or entrance.programming_code

                if apartment_range and '-' in apartment_range:
                    try:
                        p = apartment_range.split('-')
                        entrance.apartment_from = int(p[0].strip())
                        entrance.apartment_to = int(p[-1].strip())
                        entrance.apartments_count = entrance.apartment_to - entrance.apartment_from + 1
                    except (ValueError, IndexError):
                        pass

                if notes:
                    existing = entrance.notes or ''
                    if notes not in existing:
                        entrance.notes = (existing + '\n' + notes).strip()
                entrance.save()

                if created:
                    ent_created += 1
                else:
                    ent_updated += 1
            elif ent_num:
                no_match += 1

            if ip_addr and not dry_run:
                dt = None
                if date_issued and hasattr(date_issued, 'strftime'):
                    dt = timezone.make_aware(
                        date_issued.replace(tzinfo=None),
                        timezone.get_current_timezone()
                    ) if date_issued and date_issued.tzinfo is None else date_issued

                BewardDevice.objects.update_or_create(
                    ip_address=ip_addr,
                    defaults={
                        'region': region,
                        'address': address,
                        'entrance_number': str(ent_num) if ent_num else entrance_raw,
                        'access_code': access_code,
                        'programming_code': programming_code,
                        'apartment_range': apartment_range,
                        'date_issued': dt,
                        'notes': (notes or '')[:1000],
                        'building': building,
                        'entrance': entrance,
                    }
                )
                dev_updated += 1

        self.stdout.write(f'  Подъездов создано: {ent_created}')
        self.stdout.write(f'  Подъездов обновлено: {ent_updated}')
        self.stdout.write(f'  Не найдено Building: {no_match}')
        self.stdout.write(f'  BewardDevice дополнено: {dev_updated}')
        self.stdout.write('\nГотово!')
