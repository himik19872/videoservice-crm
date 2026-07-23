"""
Прямой импорт из XLSX-файлов ЕРЦ с нормализацией адресов через Dadata.

Обрабатывает 4 формата:
  1. ЕИРЦ (Коммунар, СПб, Агалатово) — колонки в разных местах
  2. ЛО — расширенный формат с дебет/кредит
  3. Красное Село — формат ТСЖ
  4. Стр.6-3 — формат ТСЖ Красное Село
  5. ТСЖ Битрикс — реестр клиентов (без начислений)

Правила:
  - Уникальность клиента: personal_account_number
  - Если нет подъезда → подъезд №1
  - При повторном импорте: клиент обновляется, начисления ЕРЦ обновляются
  - Адрес нормализуется через Dadata
"""

import os
import re
import logging
from datetime import datetime, date
from typing import Optional, Tuple, List, Dict, Any

from django.conf import settings
from openpyxl import load_workbook

from .models import (
    Client, Building, BuildingEntrance, ManagementCompany,
    ErcAccount, ErcBillingRecord,
)
from .dadata_service import normalize_address

logger = logging.getLogger(__name__)


# ─── Детекция формата XLSX ───────────────────────────────────────

def _detect_format(ws) -> str:
    """Определяет формат файла по заголовкам."""
    row6 = [str(ws.cell(6, c).value or '') for c in range(1, 13)]
    row7 = [str(ws.cell(7, c).value or '') for c in range(1, 13)]
    row9 = [str(ws.cell(9, c).value or '') for c in range(1, 13)]

    # Красное Село (ТСЖ): row1 содержит "Оборотно-сальдовая"
    r1 = str(ws.cell(1, 3).value or '')
    if 'Оборотно-сальдовая' in r1:
        return 'krasnoe'

    # Стр.6-3: row1 содержит заголовки "л/счет", "ФИО"
    r1_all = ' '.join([str(ws.cell(1, c).value or '') for c in range(1, 9)])
    if 'л/счет' in r1_all and 'ФИО' in r1_all:
        return 'str63'

    # ТСЖ Битрикс: row1 содержит "№ лицевого счета", "ТСЖ"
    if '№ лицевого счета' in r1_all or 'ТСЖ' in r1_all:
        return 'tszh'

    # Агалатово: row9 col6 = "Населенный пункт"
    if 'Населенный пункт' in row9[5]:
        return 'agalatovo'

    # ЛО: row6 col7 = "Сальдо на на"
    if 'Сальдо на на' in row6[6]:
        return 'lo'

    # По умолчанию: Коммунар / СПб
    return 'erc'


# ─── Извлечение данных из строки ──────────────────────────────────

def _extract_row(ws, row_num: int, fmt: str) -> Optional[Dict[str, Any]]:
    """Извлекает поля клиента из строки Excel в зависимости от формата."""
    def cell(c):
        val = ws.cell(row_num, c).value
        if val is None:
            return ''
        return str(val).strip()

    if fmt == 'krasnoe':
        apartment = cell(1)
        name = cell(4)
        balance_start = cell(5).replace(',', '.')
        charged = cell(7).replace(',', '.')
        paid = cell(9).replace(',', '.')
        balance_end = cell(10).replace(',', '.')
        personal_account = ''  # нет л/с в этом формате
        raw_address = ''  # адрес в заголовке группы
        return {
            'apartment': apartment,
            'name': name,
            'personal_account': personal_account,
            'raw_address': raw_address,
            'phone': '',
            'source': 'Красное Село',
            'source_file': '',
            'balance_start': _safe_float(balance_start),
            'charged': _safe_float(charged),
            'paid': _safe_float(paid),
            'balance_end': _safe_float(balance_end),
        }

    if fmt == 'str63':
        personal_account = cell(2)
        name = cell(3)
        raw_address = cell(4)
        balance_start = cell(5).replace(',', '.')
        charged = cell(6).replace(',', '.')
        paid = cell(7).replace(',', '.')
        balance_end = cell(8).replace(',', '.')
        return {
            'apartment': '',
            'name': name,
            'personal_account': personal_account,
            'raw_address': raw_address,
            'phone': '',
            'source': 'Красное Село',
            'source_file': '',
            'balance_start': _safe_float(balance_start),
            'charged': _safe_float(charged),
            'paid': _safe_float(paid),
            'balance_end': _safe_float(balance_end),
        }

    if fmt == 'tszh':
        personal_account = cell(2)
        name = cell(3)
        raw_address = cell(4)
        entrance = cell(5)
        source = cell(6)

        parsed = _parse_erc_address(raw_address)
        return {
            'apartment': parsed['apartment'],
            'name': name,
            'personal_account': personal_account,
            'raw_address': raw_address,
            'phone': '',
            'source': source,
            'source_file': '',
            'entrance': entrance,
            '_city': parsed['city'],
            '_street_name': parsed['street'],
            '_street_type': parsed['street_type'],
            '_house_number': parsed['house'],
            '_building_number': parsed['building'],
            '_apartment': parsed['apartment'],
            '_district': parsed['district'],
            '_region': parsed['region'],
            'balance_start': 0,
            'charged': 0,
            'paid': 0,
            'balance_end': 0,
        }

    if fmt == 'agalatovo':
        personal_account = cell(2)
        name = cell(3)
        city_hint = cell(6).replace(' д', '').replace(' г', '').replace(' п', '').strip()   # «Агалатово д» → «Агалатово»
        street_hint = cell(7) # улица («-» = нет)
        house_hint = cell(8)  # дом
        apartment = cell(9)   # квартира
        balance_start = cell(10).replace(',', '.')
        paid = cell(14).replace(',', '.')           # колонка 14 = Оплачено
        charged = cell(12).replace(',', '.')        # колонка 12 = Начислено
        balance_end = '0'
        return {
            'apartment': apartment,
            'name': name,
            'personal_account': personal_account,
            'raw_address': '',
            'phone': '',
            'source': 'Агалатово',
            'source_file': '',
            # Пред-распарсенные поля — НЕ перепарсивать через normalize_address
            '_city': city_hint,
            '_street_name': street_hint if street_hint and street_hint != '-' else '',
            '_house_number': house_hint,
            '_apartment': apartment,
            '_district': 'Всеволожский р-н',
            '_region': 'Ленинградская обл',
            'balance_start': _safe_float(balance_start),
            'charged': _safe_float(charged),
            'paid': _safe_float(paid),
            'balance_end': _safe_float(balance_end),
        }

    if fmt == 'lo':
        personal_account = cell(2)
        name = cell(3)
        raw_address = cell(6)
        balance_start_dt = cell(7).replace(',', '.')
        balance_start_kt = cell(8).replace(',', '.')
        charged = cell(12).replace(',', '.') if cell(12) else cell(9).replace(',', '.')
        paid = cell(13).replace(',', '.')           # колонка 13 = Оплата
        balance_end = cell(15).replace(',', '.')    # колонка 15 = Сальдо на конец
        bs = _safe_float(balance_start_dt) - _safe_float(balance_start_kt)

        # ЛО — такой же формат адреса как Коммунар/СПб
        parsed = _parse_erc_address(raw_address)
        return {
            'apartment': parsed['apartment'],
            'name': name,
            'personal_account': personal_account,
            'raw_address': raw_address,
            'phone': '',
            'source': 'ЕРЦ ЛО',
            'source_file': '',
            '_city': parsed['city'],
            '_street_name': parsed['street'],
            '_street_type': parsed['street_type'],
            '_house_number': parsed['house'],
            '_building_number': parsed['building'],
            '_apartment': parsed['apartment'],
            '_district': parsed['district'],
            '_region': parsed['region'],
            'balance_start': bs,
            'charged': _safe_float(charged),
            'paid': _safe_float(paid),
            'balance_end': _safe_float(balance_end),
        }

    # erc (Коммунар / СПб)
    personal_account = cell(2)
    name = cell(3)
    raw_address = cell(6)
    balance_start = cell(9).replace(',', '.')
    charged = cell(11).replace(',', '.')  # фактически начислено
    paid = cell(12).replace(',', '.')
    balance_end = '0'

    # Прямой парсер формата «Город г, Улица ул, д Дом, кв Квартира»
    # или «Санкт-Петербург, Улица ул, д..Дом лит. А, Квартира»
    parsed = _parse_erc_address(raw_address)
    return {
        'apartment': parsed['apartment'],
        'name': name,
        'personal_account': personal_account,
        'raw_address': raw_address,
        'phone': '',
        'source': '',
        'source_file': '',
        '_city': parsed['city'],
        '_street_name': parsed['street'],
        '_street_type': parsed['street_type'],
        '_house_number': parsed['house'],
        '_building_number': parsed['building'],
        '_apartment': parsed['apartment'],
        '_district': parsed['district'],
        '_region': parsed['region'],
        'balance_start': _safe_float(balance_start),
        'charged': _safe_float(charged),
        'paid': _safe_float(paid),
        'balance_end': _safe_float(balance_end),
    }


def _safe_float(s: str) -> float:
    try:
        return float(s.replace(',', '.')) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


# ─── Прямой парсер адресов ERC (Коммунар / СПб) ───────────────────

def _parse_erc_address(raw_address: str) -> dict:
    """
    Парсит форматы:
      - «Коммунар г, Бумажников ул, 2, 2»
      - «Санкт-Петербург, Победы (Ломоносов) ул, д..1 лит. А, 1»
      - «Аннино п, 10-й пятилетки ул, 1, 1»
      - «Горбунки д, 40, 1» (деревня без улицы)

    Возвращает {city, street, street_type, house, building, apartment, district, region}
    """
    import re
    s = raw_address.strip()
    city = ''
    street = ''
    street_type = 'street'
    house = ''
    building = ''
    apartment = ''
    district = ''
    region = ''

    # Разбиваем по запятым
    parts = [p.strip() for p in s.split(',')]

    # ── Определяем регион ──
    if 'Санкт-Петербург' in s:
        region = 'Санкт-Петербург'
    else:
        region = 'Ленинградская обл'

    # ── Определяем формат адреса ──
    # Формат 1 (ERC): «Город г/п/д, Улица ул, Дом, Квартира»
    # Формат 2 (ЛО обратный): «Улица ул, Дом, Город г/п/д»
    # Отличие: в 1-й части формата 1 — город с суффиксом; в формате 2 — улица с суффиксом
    is_reverse = False
    if parts:
        first = parts[0].strip()
        # Если 1-я часть заканчивается на «ул», «шоссе», «ш» и т.п. — это обратный формат
        if re.search(r'\s+(ул|шоссе|пр-кт|пер|ш|наб|проезд|б-р|пл|аллея)\.?$', first):
            is_reverse = True

    # ── Парсим ──
    if is_reverse:
        # Обратный формат: «Улица ул, Дом, Город г/п/д»
        if parts:
            street_raw = parts[0].strip()
            st = re.sub(r'\s+(ул|шоссе|пр-кт|пер|ш|наб|проезд|б-р|пл|аллея)\.?$', '', street_raw).strip()
            street = st
            m = re.search(r'\s+(ул|шоссе|пр-кт|пер|ш|наб|проезд|б-р|пл|аллея)\.?$', street_raw)
            if m:
                type_map = {
                    'ул': 'street', 'пр-кт': 'avenue', 'пер': 'lane',
                    'б-р': 'boulevard', 'ш': 'highway', 'пл': 'square',
                    'наб': 'embankment', 'проезд': 'passage', 'аллея': 'alley',
                    'шоссе': 'highway',
                }
                street_type = type_map.get(m.group(1), 'street')
        if len(parts) >= 2:
            # Дом
            house = parts[1].strip()
        if len(parts) >= 3:
            # Город
            city_raw = parts[2].strip()
            city = re.sub(r'\s+(г|п|д|гп|сп|рп|с|ст)$', '', city_raw).strip()
            if not city:
                city = city_raw
        if len(parts) >= 4:
            apartment = parts[3].strip()
    else:
        # ── Город (1-й элемент) ──
        if parts:
            city_raw = parts[0]
            city = re.sub(r'\s+(г|п|д|гп|сп|рп|с|ст)$', '', city_raw).strip()
            if not city:
                city = city_raw

        # ── Улица (2-й элемент) ──
        if len(parts) >= 2:
            street_raw = parts[1].strip()
            if street_raw in ('-', '', '—'):
                street = ''
            elif re.match(r'^\d+[а-яА-ЯA-Za-z]*$', street_raw):
                street = ''
                house = street_raw
            else:
                st = street_raw
                for suffix in ['ул', 'шоссе', 'ш', 'пр-кт', 'пер', 'наб', 'проезд', 'б-р', 'пл', 'аллея', 'мкр']:
                    st = re.sub(rf'\s+{suffix}\.?$', '', st).strip()
                if st and st != street_raw:
                    street = st
                    m = re.search(rf'\s+(ул|шоссе|ш|пр-кт|пер|наб|проезд|б-р|пл|аллея|мкр)\.?$', street_raw)
                    if m:
                        type_map = {
                            'ул': 'street', 'пр-кт': 'avenue', 'пер': 'lane',
                            'б-р': 'boulevard', 'ш': 'highway', 'пл': 'square',
                            'наб': 'embankment', 'проезд': 'passage', 'аллея': 'alley',
                            'мкр': 'microdistrict', 'шоссе': 'highway',
                        }
                        street_type = type_map.get(m.group(1), 'street')
                else:
                    street = street_raw

    # ── Определяем где дом и квартира ──
    remaining = parts[2:] if len(parts) >= 3 else []

    for p in remaining:
        p = p.strip()

        # Дом: «д..1 лит. А», «д 1А», «11 корп. 1 стр. 1»
        dm = re.search(r'д\.\.?\s*(\d+[а-яА-ЯA-Za-z]*)', p)
        if dm:
            house = dm.group(1)
            # Литера
            lm = re.search(r'лит\.?\s*([А-ЯA-Z])', p)
            if lm:
                building = lm.group(1)
            # Корпус
            km = re.search(r'корп\.?\s*(\d+)', p)
            if km:
                building = km.group(1)
            continue

        # Просто число/число+буква — дом (если ещё нет) или квартира
        num_match = re.match(r'^(\d+[а-яА-ЯA-Za-z]*)\s*$', p)
        if num_match:
            val = num_match.group(1)
            if not house:
                house = val
            else:
                apartment = val
        else:
            # Может быть последним — квартира
            num_end = re.search(r'(\d+[а-яА-ЯA-Za-z]*)$', p)
            if num_end:
                apartment = num_end.group(1)

    # ── Определяем район ──
    if 'Гатчин' in raw_address or 'Коммунар' in raw_address or 'Тайцы' in raw_address:
        district = 'Гатчинский р-н'
    elif 'Ломоносов' in raw_address or 'Петергоф' in raw_address or 'Стрельна' in raw_address:
        district = 'Петродворцовый р-н'
    elif raw_address.startswith('Санкт-Петербург,'):
        district = 'Петродворцовый р-н'  # по умолчанию для СПб пригородов
    elif 'Пушкин' in raw_address:
        district = 'Пушкинский р-н'
    elif 'Колпин' in raw_address:
        district = 'Колпинский р-н'
    elif 'Аннино' in raw_address or 'Горбунки' in raw_address or 'Виллози' in raw_address:
        district = 'Ломоносовский р-н'
    elif 'Кипень' in raw_address or 'Келози' in raw_address or 'Карлино' in raw_address:
        district = 'Ломоносовский р-н'
    elif 'Яльгелево' in raw_address or 'Разбегаево' in raw_address or 'Пеники' in raw_address:
        district = 'Ломоносовский р-н'
    elif 'Ижора' in raw_address:
        district = 'Ломоносовский р-н'
    elif 'Сертолово' in raw_address or 'Агалат' in raw_address:
        district = 'Всеволожский р-н'
    elif 'Горелово' in raw_address or 'Красное Село' in raw_address:
        district = 'Красносельский р-н'

    return {
        'city': city,
        'street': street,
        'street_type': street_type,
        'house': house,
        'building': building,
        'apartment': apartment,
        'district': district,
        'region': region,
    }


# ─── Поиск первой строки с данными ────────────────────────────────

def _find_data_start(ws, fmt: str) -> int:
    """Возвращает номер первой строки с данными."""
    for r in range(1, ws.max_row + 1):
        if fmt == 'krasnoe':
            val = str(ws.cell(r, 4).value or '')
            if val and len(val) > 3 and not any(w in val.lower() for w in ['фио', 'ф.и.о', 'фамилия']):
                return r
        elif fmt == 'str63':
            val = str(ws.cell(r, 2).value or '')
            if val and val.isdigit() and len(val) >= 4:
                return r
        elif fmt == 'tszh':
            val = str(ws.cell(r, 2).value or '')
            if val and val.isdigit() and len(val) >= 8:
                return r
        elif fmt in ('erc', 'lo', 'agalatovo'):
            val = str(ws.cell(r, 2).value or '')
            if val and val.isdigit() and len(val) >= 8:
                return r
    return 1


# ─── Главная функция импорта ─────────────────────────────────────

def _get_or_create_region(city: str, dadata_region: str):
    """Определяет код региона по городу/области."""
    from .models import Region
    # Маппинг: название → код
    REGION_MAP = {
        'санкт-петербург': ('78', 'Санкт-Петербург'),
        'спб': ('78', 'Санкт-Петербург'),
        'ленинградская': ('47', 'Ленинградская область'),
        'москва': ('77', 'Москва'),
        'московская': ('50', 'Московская область'),
    }
    region_name = dadata_region or 'Санкт-Петербург'
    key = (city + ' ' + region_name).lower()
    code = '78'
    for k, (c, n) in REGION_MAP.items():
        if k in key:
            code = c
            region_name = n
            break
    region, _ = Region.objects.get_or_create(code=code, defaults={'name': region_name, 'country': 'Россия'})
    return region


def import_xlsx_file(file_path: str, source_filename: str = '', period_date: date = None):
    """
    Импорт XLSX ЕРЦ с полной адресной иерархией:
      Страна(Россия) → Регион(78/47) → Район → Город → Улица → Дом → Корпус → Квартира
    Общежития: если в одной квартире несколько л/с → is_dormitory=True.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    fmt = _detect_format(ws)
    data_start = _find_data_start(ws, fmt)

    from .models import Region

    stats = {
        'total_rows': 0, 'clients_created': 0, 'clients_updated': 0,
        'buildings_created': 0, 'entrances_created': 0,
        'erc_created': 0, 'erc_updated': 0, 'dormitories': 0,
        'skipped': 0, 'errors': [],
    }

    if period_date is None:
        period_date = date(2026, 5, 1)

    for r in range(data_start, ws.max_row + 1):
        try:
            row_data = _extract_row(ws, r, fmt)
            if not row_data:
                stats['skipped'] += 1
                continue

            personal_account = row_data['personal_account']
            name = row_data['name']
            raw_address = row_data['raw_address']

            # Пропуск служебных строк
            if not personal_account or len(personal_account) < 5:
                if not name or len(name) < 3:
                    stats['skipped'] += 1
                    continue
            if not personal_account.isdigit():
                stats['skipped'] += 1
                continue
            skip_words = ['лицевого', 'счета', 'лицевой', 'номер', 'фамилия', 'фио']
            if any(w in name.lower() for w in skip_words):
                stats['skipped'] += 1
                continue

            stats['total_rows'] += 1

            # ── Нормализация через Dadata + fallback ──
            # Если row_data содержит пред-распарсенные поля (_city, _street_name...) — используем их
            pre_city = row_data.get('_city', '')
            pre_street = row_data.get('_street_name', '')
            pre_house = row_data.get('_house_number', '')
            pre_apt = row_data.get('_apartment', '')

            if pre_city or pre_house:
                # Агалатово: собираем адрес из кусочков → нормализуем через DaData
                parts = []
                rgn = row_data.get('_region', '')
                dst = row_data.get('_district', '')
                if rgn:
                    parts.append(rgn)
                if dst:
                    parts.append(dst)
                if pre_city:
                    parts.append(pre_city)
                if pre_street and pre_street != '-':
                    parts.append(pre_street)
                if pre_house:
                    parts.append(f'д {pre_house}')
                if pre_apt:
                    parts.append(f'кв {pre_apt}')
                raw_for_dadata = ', '.join(parts)

                addr = normalize_address(raw_for_dadata)
                city = addr.get('city', '') or pre_city or 'Санкт-Петербург'
                region_str = addr.get('region', '') or rgn
                district = addr.get('district', '') or dst
                street_name = addr.get('street_name', '') or (pre_street if pre_street != '-' else '')
                street_type = addr.get('street_type', '') or 'street'
                house_number = addr.get('house_number', '') or pre_house
                building_number = addr.get('building_number', '') or ''
                apartment = pre_apt or row_data.get('apartment', '') or addr.get('apartment', '')
            else:
                addr = normalize_address(raw_address) if raw_address else {'success': False}
                city = addr.get('city', '') or 'Санкт-Петербург'
                region_str = addr.get('region', '')
                district = addr.get('district', '')
                street_name = addr.get('street_name', '')
                street_type = addr.get('street_type', 'street')
                house_number = addr.get('house_number', '')
                building_number = addr.get('building_number', '')
                apartment = row_data.get('apartment', '') or addr.get('apartment', '')

            # ── Регион с кодом ──
            region_obj = _get_or_create_region(city, region_str)

            # ── Здание (уникально: region + city + street_name + house_number + building_number) ──
            building_kwargs = {
                'city__icontains': city,
                'house_number': house_number,
            }
            if street_name:
                building_kwargs['street_name__icontains'] = street_name
            else:
                building_kwargs['street_name'] = ''

            if building_number:
                building_kwargs['building_number'] = building_number

            building = Building.objects.filter(**building_kwargs).first()
            if not building:
                building = Building.objects.create(
                    region=region_obj,
                    city=city,
                    district=district,
                    street_name=street_name,
                    street_type=street_type,
                    house_number=house_number,
                    building_number=building_number,
                    management_company=row_data.get('source', '') or region_str,
                )
                stats['buildings_created'] += 1
            elif not building.region:
                building.region = region_obj
                building.save(update_fields=['region'])

            # ── Подъезд (умолч. 1) ──
            entrance_obj = None
            if building:
                ent_num = 1
                ent_str = row_data.get('entrance', '') or ''
                if ent_str:
                    try:
                        ent_num = int(float(ent_str))
                    except (ValueError, TypeError):
                        ent_num = 1
                entrance_obj, ent_created = BuildingEntrance.objects.get_or_create(
                    building=building, number=ent_num,
                    defaults={'apartments_count': 0}
                )
                if ent_created:
                    stats['entrances_created'] += 1

            # ── Общежитие: одна квартира + разные л/с ──
            if apartment:
                same_apt = Client.objects.filter(
                    building=building, apartment=apartment
                ).exclude(personal_account_number=personal_account).exists()
                if same_apt and not building.is_dormitory:
                    building.is_dormitory = True
                    building.save(update_fields=['is_dormitory'])
                    stats['dormitories'] += 1

            # ── УК ──
            source = row_data.get('source', '')
            mc_obj = None
            if source:
                mc_obj, _ = ManagementCompany.objects.get_or_create(name=source)
                if building and mc_obj and not building.management_company_fk:
                    building.management_company_fk = mc_obj
                    building.save(update_fields=['management_company_fk'])

            # ── Адресная строка ──
            type_label = {
                'street': 'ул.', 'avenue': 'пр-кт', 'lane': 'пер.',
                'boulevard': 'б-р', 'highway': 'ш.', 'square': 'пл.',
                'embankment': 'наб.', 'passage': 'проезд', 'alley': 'аллея',
                'microdistrict': 'мкр.',
            }.get(street_type, 'ул.')

            parts = [f'г. {city}']
            if region_str:
                parts.append(region_str)
            if district:
                parts.append(district)
            if street_name:
                parts.append(f'{type_label} {street_name}')
            if house_number:
                h = f'д. {house_number}'
                if building_number:
                    h += f' корп. {building_number}'
                parts.append(h)
            if apartment:
                parts.append(f'кв. {apartment}')
            full_address = ', '.join(parts)

            # ── Квартира (объект) ──
            apartment_obj = None
            if building and apartment:
                from .models import Apartment
                apartment_obj, _ = Apartment.objects.get_or_create(
                    building=building, number=apartment,
                    defaults={'entrance': entrance_obj}
                )
                if entrance_obj and not apartment_obj.entrance:
                    apartment_obj.entrance = entrance_obj
                    apartment_obj.save(update_fields=['entrance'])

            # ── Клиент (КЛЮЧ = personal_account) ──
            existing = Client.objects.filter(personal_account_number=personal_account).first()

            if existing:
                if building and not existing.building:
                    existing.building = building
                if entrance_obj and not existing.entrance:
                    existing.entrance = entrance_obj
                if apartment_obj and not existing.apartment_obj:
                    existing.apartment_obj = apartment_obj
                if not existing.address or existing.address == 'г. , д. ':
                    existing.address = full_address
                if name and existing.name == 'Не определено':
                    existing.name = name
                if not existing.apartment and apartment:
                    existing.apartment = apartment
                if not existing.region:
                    existing.region = region_obj
                existing.save()
                stats['clients_updated'] += 1
            else:
                Client.objects.create(
                    name=name or 'Не определено',
                    address=full_address,
                    building=building,
                    entrance=entrance_obj,
                    apartment=apartment,
                    apartment_obj=apartment_obj,
                    management_company=mc_obj,
                    personal_account_number=personal_account,
                    region=region_obj,
                    district=district,
                    source='erc',
                )
                stats['clients_created'] += 1

            # ── ЕРЦ запись ──
            charged = row_data.get('charged', 0)
            paid = row_data.get('paid', 0)
            balance_start = row_data.get('balance_start', 0)
            balance_end = row_data.get('balance_end', 0)

            if charged or paid or balance_start:
                erc_account, _ = ErcAccount.objects.get_or_create(
                    account_number=personal_account,
                    defaults={'full_name': name, 'address': full_address}
                )
                _, created = ErcBillingRecord.objects.update_or_create(
                    account=erc_account,
                    period=period_date,
                    defaults={'balance_start': balance_start, 'charged': charged, 'paid': paid, 'balance_end': balance_end}
                )
                if created:
                    stats['erc_created'] += 1
                else:
                    stats['erc_updated'] += 1

        except Exception as e:
            stats['errors'].append(f'Строка {r}: {str(e)}')

    logger.warning(
        f'[XLSX IMPORT] rows={stats["total_rows"]} +{stats["clients_created"]} ~{stats["clients_updated"]} '
        f'bld={stats["buildings_created"]} entr={stats["entrances_created"]} '
        f'dorm={stats["dormitories"]} erc:{stats["erc_created"]}+~{stats["erc_updated"]} sk={stats["skipped"]}'
    )
    return stats


def import_erc_update_only(file_path: str, period_date: date = None):
    """
    Режим обновления ЕРЦ: обновляет ТОЛЬКО ErcBillingRecord по номеру лицевого счёта.
    Не создаёт клиентов, не парсит адреса. Только: л/с → сальдо, начислено, оплачено.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    fmt = _detect_format(ws)
    data_start = _find_data_start(ws, fmt)

    if period_date is None:
        period_date = date(2026, 5, 1)

    stats = {
        'total_rows': 0,
        'erc_updated': 0,
        'erc_created': 0,
        'skipped': 0,
        'errors': [],
    }

    for r in range(data_start, ws.max_row + 1):
        try:
            row_data = _extract_row(ws, r, fmt)
            if not row_data:
                stats['skipped'] += 1
                continue

            personal_account = row_data.get('personal_account', '')
            if not personal_account or len(personal_account) < 5 or not personal_account.isdigit():
                stats['skipped'] += 1
                continue

            name = row_data.get('name', '')
            skip_words = ['лицевого', 'счета', 'лицевой', 'номер', 'фамилия', 'фио']
            if any(w in name.lower() for w in skip_words):
                stats['skipped'] += 1
                continue

            stats['total_rows'] += 1

            # Только ЕРЦ запись — без клиента, без адреса
            charged = row_data.get('charged', 0)
            paid = row_data.get('paid', 0)
            balance_start = row_data.get('balance_start', 0)
            balance_end = row_data.get('balance_end', 0)

            if charged or paid or balance_start:
                erc_account, _ = ErcAccount.objects.get_or_create(
                    account_number=personal_account,
                    defaults={'full_name': name}
                )
                _, created = ErcBillingRecord.objects.update_or_create(
                    account=erc_account,
                    period=period_date,
                    defaults={
                        'balance_start': balance_start,
                        'charged': charged,
                        'paid': paid,
                        'balance_end': balance_end,
                    }
                )
                if created:
                    stats['erc_created'] += 1
                else:
                    stats['erc_updated'] += 1

        except Exception as e:
            stats['errors'].append(f'Строка {r}: {str(e)}')

    logger.warning(
        f'[ERC UPDATE] rows={stats["total_rows"]} +{stats["erc_created"]} ~{stats["erc_updated"]} sk={stats["skipped"]}'
    )
    return stats


def import_xlsx_directory(directory: str, period_date: date = None):
    """
    Импортирует ВСЕ XLSX-файлы из директории.
    Возвращает сводную статистику.
    """
    all_stats = []
    for fname in sorted(os.listdir(directory)):
        if not fname.endswith('.xlsx'):
            continue
        if fname.startswith('~'):
            continue

        filepath = os.path.join(directory, fname)
        stats = import_xlsx_file(filepath, source_filename=fname, period_date=period_date)
        all_stats.append(stats)

    # Сводка
    total = {
        'files': len(all_stats),
        'total_rows': sum(s['total_rows'] for s in all_stats),
        'clients_created': sum(s['clients_created'] for s in all_stats),
        'clients_updated': sum(s['clients_updated'] for s in all_stats),
        'buildings_created': sum(s['buildings_created'] for s in all_stats),
        'entrances_created': sum(s['entrances_created'] for s in all_stats),
        'erc_created': sum(s['erc_created'] for s in all_stats),
        'erc_updated': sum(s['erc_updated'] for s in all_stats),
        'skipped': sum(s['skipped'] for s in all_stats),
        'details': all_stats,
    }
    return total
