"""
Сервис импорта данных из Excel-файлов.

Поддерживает два формата:
1. База клиентов (ТСЖ/УК): № п/п, № лицевого счета, ФИО, Адрес, № парадной, ТСЖ
   Адрес нормализуется через DaData API.
2. Оборотная ведомость ЕРЦ (форма № 30.01.01): автоопределение колонок по заголовкам.
"""

import io
import re
import requests
from datetime import datetime
from collections import defaultdict
import openpyxl


# ══════════════════════════════════════════════════════════════════
# Нормализация адреса через DaData
# ══════════════════════════════════════════════════════════════════

def get_dadata_token():
    """Получить DaData API токен из настроек."""
    try:
        from main.models import SystemSettings
        s = SystemSettings.objects.first()
        return s.dadata_token if s else ''
    except Exception:
        return ''


def normalize_address_dadata(raw_address):
    """
    Нормализует адрес через DaData API.
    Возвращает dict с полями из ответа DaData или None при ошибке.
    """
    token = get_dadata_token()
    if not token:
        return None

    try:
        resp = requests.post(
            'https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/address',
            json={'query': raw_address, 'count': 1},
            headers={
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'Authorization': f'Token {token}'
            },
            timeout=5
        )
        data = resp.json()
        suggestions = data.get('suggestions', [])
        if not suggestions:
            return None

        s = suggestions[0]
        d = s.get('data', {})

        return {
            'unrestricted': s.get('unrestricted_value', ''),
            'value': s.get('value', ''),
            'city': d.get('city', '') or d.get('settlement', ''),
            'city_district': d.get('city_district', ''),
            'street': d.get('street', ''),
            'house': d.get('house', ''),
            'block': d.get('block', ''),
            'flat': d.get('flat', ''),
            'postal_code': d.get('postal_code', ''),
            'region': d.get('region', ''),
            'area': d.get('area', ''),
            'settlement': d.get('settlement', ''),
        }
    except Exception:
        return None


def parse_address(raw_address):
    """
    Разбирает адрес на составные части с использованием DaData для нормализации.
    Если DaData недоступен — использует упрощённый регекс-парсер как fallback.
    
    Возвращает dict: city, street, district, house, building, apartment, full.
    """
    result = {
        'city': '',
        'street': '',
        'district': '',
        'house': '',
        'building': '',
        'apartment': '',
        'full': raw_address.strip() if raw_address else '',
    }

    if not raw_address or not raw_address.strip():
        return result

    raw = raw_address.strip()
    result['full'] = raw

    # Пробуем нормализовать через DaData
    dd = normalize_address_dadata(raw)
    if dd:
        # Город: приоритет city > settlement
        result['city'] = dd['city'] or ''
        # Если есть city_district (район города, напр. Петергоф, Ломоносов) — добавляем
        if dd['city_district']:
            if result['city']:
                result['district'] = dd['city_district']
            else:
                result['city'] = dd['city_district']

        result['street'] = dd['street'] or ''
        result['house'] = dd['house'] or ''
        result['building'] = dd['block'] or ''
        result['apartment'] = dd['flat'] or ''

        # Формируем полный адрес
        address_parts = []
        if result['city']:
            prefix = 'г. ' if not result['city'].startswith(('г.', 'г ', 'пос.', 'д.')) else ''
            address_parts.append(f'{prefix}{result["city"]}')
        if result['district'] and result['district'] != result['city']:
            address_parts[-1] = f'{address_parts[-1]}, {result["district"]}' if address_parts else result['district']
        if result['street']:
            address_parts.append(result['street'])
        if result['house']:
            house_str = f'д. {result["house"]}'
            if result['building']:
                house_str += f' корп. {result["building"]}'
            address_parts.append(house_str)
        if result['apartment']:
            address_parts.append(f'кв. {result["apartment"]}')
        
        result['full'] = ', '.join(address_parts) if address_parts else raw
        return result

    # ── Fallback: упрощённый парсер (если DaData недоступен) ──
    text = raw
    district_match = re.search(r'\(([^)]+)\)', text)
    if district_match:
        result['district'] = district_match.group(1).strip()
        text = text.replace(district_match.group(0), '').strip()
        text = re.sub(r',\s*,', ',', text)

    parts = [p.strip() for p in text.split(',') if p.strip()]

    if parts:
        first = parts[0]
        if any(c in first for c in ['Санкт-Петербург', 'Москва', 'СПб', 'Спб']):
            result['city'] = first.replace('СПб', 'Санкт-Петербург').replace('Спб', 'Санкт-Петербург')
            parts = parts[1:]

    for i, p in enumerate(parts):
        if any(kw in p.lower() for kw in [' ул', 'ул.', ' ул.', 'пр-кт', 'пр.', 'пер.', 'наб.', 'шоссе', 'проезд', 'аллея', 'бульвар']):
            result['street'] = p
            remaining = parts[i+1:]
            for rp in remaining:
                rp = rp.strip()
                hm = re.search(r'д\.?\.?\s*(\d+[а-яА-Яa-zA-Z]*)', rp, re.IGNORECASE)
                if hm and not result['house']:
                    result['house'] = hm.group(1)
                cm = re.search(r'корп\.?\s*(\S+)', rp, re.IGNORECASE)
                if cm:
                    result['building'] = cm.group(1).rstrip(',')
                am = re.search(r'(\d+)$', rp.strip())
                if am and result['house'] and not result['apartment']:
                    result['apartment'] = am.group(1)
            break

    return result


# ══════════════════════════════════════════════════════════════════
# Импорт базы клиентов (ТСЖ)
# ══════════════════════════════════════════════════════════════════

def import_clients_from_excel(file_bytes, user):
    """
    Импорт клиентов из Excel-файла (ТСЖ/УК).
    Формат: № п/п, № лицевого счета (игнорируется), ФИО, Адрес, № парадной, ТСЖ

    Адрес парсится на: город, улица, район, дом, корпус, квартира.
    Возвращает: dict с результатами
    """
    from .models import Client

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # пропускаем заголовок
    total = 0
    created = 0
    updated = 0
    errors = []

    for row in rows:
        # Нет адреса и нет ФИО — пропускаем
        if not row:
            continue

        total += 1
        try:
            # Колонки: 0=№п/п, 1=№лицевого (игнорируем), 2=ФИО, 3=Адрес, 4=№парадной, 5=ТСЖ
            full_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else 'Не определено'
            raw_address = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            entrance = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            management_company = str(row[5]).strip() if len(row) > 5 and row[5] else ''

            # Пропускаем только полностью пустые строки (нет ни адреса, ни ФИО, ни УК)
            if not full_name and not raw_address and not management_company:
                continue

            # Парсим и нормализуем адрес через DaData
            parsed = parse_address(raw_address)

            # Используем нормализованный полный адрес от DaData
            address = parsed['full'] if parsed['full'] else raw_address

            # Пытаемся найти существующего клиента по адресу
            existing = Client.objects.filter(
                address=address, name=full_name
            ).first() if address else None

            if existing:
                # Обновляем существующего
                existing.entrance_number = entrance or existing.entrance_number
                existing.management_company = management_company or existing.management_company
                existing.district = parsed['district'] or existing.district
                existing.source = 'excel_import'
                existing.save()
                updated += 1
            else:
                Client.objects.create(
                    name=full_name if full_name else 'Не определено',
                    address=address,
                    phone='',
                    entrance_number=entrance,
                    management_company=management_company,
                    district=parsed['district'],
                    source='excel_import',
                )
                created += 1

        except Exception as e:
            errors.append(f'Строка {total}: {str(e)}')

    return {
        'success': True,
        'total': total,
        'created': created,
        'updated': updated,
        'errors': errors,
    }


# ══════════════════════════════════════════════════════════════════
# Импорт ЕРЦ — универсальный (4 формата)
# ══════════════════════════════════════════════════════════════════

def _find_column_indices(header_rows, ws_max_column):
    """
    Ищет индексы колонок по заголовкам. Поддерживает 4 формата ЕРЦ.
    Возвращает dict с детектированным форматом и маппингом колонок.
    """
    max_cols = max(len(row) for row in header_rows if row)
    combined = [''] * max_cols
    
    for row in header_rows:
        if not row:
            continue
        for i in range(len(row)):
            if i < len(combined) and row[i] is not None:
                combined[i] += ' ' + str(row[i]).lower().strip()

    def has(idx, *words):
        if idx >= len(combined):
            return False
        return any(w in combined[idx] for w in words)

    mapping = {}
    fmt = 'standard'  # standard | lo | agalatovo

    # 1. Лицевой счёт — вторая колонка
    mapping['account_number'] = 1
    
    # 2. ФИО — третья колонка
    mapping['full_name'] = 2

    # Определяем формат
    has_city_col = any('населенный пункт' in combined[i] for i in range(len(combined)))
    has_credit_col = any('кредит' in combined[i] and i > 5 for i in range(len(combined)))
    has_paid_pct = any('% оплаты' in combined[i] for i in range(len(combined)))

    if has_city_col:
        # Формат Агалатово: адрес разбит по колонкам
        fmt = 'agalatovo'
        mapping['address_city'] = 5      # Населенный пункт
        mapping['address_street'] = 6     # Улица
        mapping['address_house'] = 7      # Дом
        mapping['address_apt'] = 8        # Кв.
        mapping['balance_start'] = 9      # Сальдо на 01.05 (колонка 10 Excel)
        mapping['charged'] = 11           # Начислено (колонка 12 Excel)
        mapping['paid'] = 13              # Оплачено (колонка 14 Excel)
        mapping['paid_percent'] = None
        mapping['balance_end'] = None
    elif has_credit_col and not has_paid_pct:
        # Формат ЛО: дебет/кредит, но без % оплаты
        fmt = 'lo'
        mapping['format'] = 'lo'
        mapping['address'] = 5
        mapping['balance_start'] = 6      # дебет
        mapping['balance_start_credit'] = 7  # кредит
        mapping['charged'] = 8            # Начислено
        mapping['charged_total'] = 11     # Всего начислено
        mapping['paid'] = 12              # Оплата
        mapping['balance_end'] = 14       # дебет
        mapping['balance_end_credit'] = 15  # кредит
        mapping['paid_percent'] = None
        mapping['residents'] = None
    else:
        # Стандартный формат (СПб, Коммунар)
        fmt = 'standard'
        mapping['format'] = 'standard'
        mapping['address'] = 5
        # Ищем колонки по заголовкам
        for i in range(6, max_cols):
            if 'жильцов' in combined[i]:
                mapping['residents'] = i
            elif 'сальдо' in combined[i] and any(d in combined[i] for d in ['01.', 'начало']):
                mapping['balance_start'] = i
            elif 'без льгот' in combined[i]:
                mapping['charged_no_benefits'] = i
            elif 'фактически' in combined[i]:
                mapping['charged'] = i
            elif 'начислено' in combined[i] and 'льгот' not in combined[i] and 'charged' not in mapping:
                mapping['charged'] = i
            elif 'оплачено' in combined[i] and '%' not in combined[i]:
                mapping['paid'] = i
            elif '% оплаты' in combined[i]:
                mapping['paid_percent'] = i
            elif 'сальдо' in combined[i] and any(d in combined[i] for d in ['конец', '31.', 'конечн']):
                mapping['balance_end'] = i
            elif 'кредит' in combined[i]:
                mapping['credit'] = i
        
        # Defaults для стандартного формата
        mapping.setdefault('balance_start', 8)
        mapping.setdefault('charged', 10)
        mapping.setdefault('paid', 11)
        mapping.setdefault('paid_percent', 12)
        mapping.setdefault('balance_end', 13)
        mapping.setdefault('credit', 14)
        mapping.setdefault('residents', 7)
        mapping.setdefault('charged_no_benefits', 9)

    return mapping, fmt


def import_erc_from_excel(file_bytes, user, period_date=None):
    """Универсальный импорт ЕРЦ. Автоопределение формата."""
    from .models import ErcAccount, ErcBillingRecord, Client
    from django.utils import timezone

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Ищем строку заголовка с «фамилия» + «лицевого»
    header_start = None
    for r in range(1, min(40, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, min(ws.max_column+1, 20))]
        combined = ' '.join(row_vals)
        if ('фамилия' in combined or 'фио' in combined) and ('лицевого' in combined or 'счета' in combined or 'адрес' in combined):
            header_start = r
            break

    if header_start is None:
        header_start = 6

    # Берём 4 строки заголовка
    header_rows = []
    for r in range(header_start, min(header_start + 4, ws.max_row + 1)):
        header_rows.append(list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0])

    col, fmt = _find_column_indices(header_rows, ws.max_column)

    # Ищем первую строку данных
    data_start_row = header_start + 4
    for r in range(header_start + 3, min(header_start + 15, ws.max_row + 1)):
        val = ws.cell(row=r, column=2).value  # л/с всегда в колонке 2
        if val is not None and str(val).strip():
            data_start_row = r
            break

    rows = list(ws.iter_rows(min_row=data_start_row, values_only=True))

    # Период
    if period_date is None:
        today = timezone.localdate()
        month = today.month - 1 if today.day <= 10 else today.month
        year = today.year
        if month == 0:
            month = 12; year -= 1
        period_date = today.replace(year=year, month=month, day=1)

    total = created_accounts = updated_accounts = created_records = updated_records = 0
    errors = []

    def parse_decimal(val):
        if val is None or str(val).strip() in ('', '—', '-'):
            return 0
        try:
            return float(str(val).replace(',', '.').replace('\xa0', '').replace(' ', ''))
        except: return 0

    def get_row(row, idx):
        if idx is None or idx >= len(row): return ''
        return row[idx] if row[idx] is not None else ''

    for row in rows:
        if not row: continue
        account_number = str(get_row(row, 1)).strip()
        if not account_number: continue

        total += 1
        try:
            full_name = str(get_row(row, 2)).strip()

            # Собираем адрес в зависимости от формата
            if fmt == 'agalatovo':
                city = str(get_row(row, col.get('address_city', 5))).strip()
                street = str(get_row(row, col.get('address_street', 6))).strip()
                house = str(get_row(row, col.get('address_house', 7))).strip()
                apt = str(get_row(row, col.get('address_apt', 8))).strip()
                parts = []
                if city and city != '-': parts.append(city)
                if street and street != '-': parts.append(street)
                if house and house != '-': parts.append(f'д. {house}')
                if apt and apt != '-': parts.append(f'кв. {apt}')
                address = ', '.join(parts) if parts else ''
                
                balance_start = parse_decimal(get_row(row, col.get('balance_start', 10)))
                charged = parse_decimal(get_row(row, col.get('charged', 12)))
                paid = parse_decimal(get_row(row, col.get('paid', 13)))
                paid_percent = 0
                balance_end = 0
                credit = 0
            elif fmt == 'lo':
                address = str(get_row(row, col.get('address', 5))).strip()
                balance_start = parse_decimal(get_row(row, col.get('balance_start', 6)))
                charged = parse_decimal(get_row(row, col.get('charged_total', 11)) or 
                                       get_row(row, col.get('charged', 8)))
                paid = parse_decimal(get_row(row, col.get('paid', 12)))
                balance_end = parse_decimal(get_row(row, col.get('balance_end', 14)))
                paid_percent = round(paid / charged * 100, 1) if charged else 0
                credit = parse_decimal(get_row(row, col.get('balance_end_credit', 15)))
            else:
                address = str(get_row(row, col.get('address', 5))).strip()
                residents = parse_decimal(get_row(row, col.get('residents', 7)))
                balance_start = parse_decimal(get_row(row, col.get('balance_start', 8)))
                charged_no_benefits = parse_decimal(get_row(row, col.get('charged_no_benefits', 9)))
                charged = parse_decimal(get_row(row, col.get('charged', 10)))
                paid = parse_decimal(get_row(row, col.get('paid', 11)))
                paid_percent = parse_decimal(get_row(row, col.get('paid_percent', 12)))
                balance_end = parse_decimal(get_row(row, col.get('balance_end', 13)))
                credit = parse_decimal(get_row(row, col.get('credit', 14)))

            account, is_new = ErcAccount.objects.update_or_create(
                account_number=account_number,
                defaults={'full_name': full_name, 'address': address, 'residents_count': 0, 'is_active': True}
            )
            if is_new: created_accounts += 1
            else: updated_accounts += 1

            Client.objects.update_or_create(
                personal_account_number=account_number,
                defaults={'name': full_name or 'Не определено', 'address': address, 'phone': '', 'source': 'erc'}
            )

            ErcBillingRecord.objects.update_or_create(
                account=account, period=period_date,
                defaults={
                    'balance_start': balance_start, 'charged': charged,
                    'charged_no_benefits': 0, 'paid': paid,
                    'paid_percent': paid_percent, 'balance_end': balance_end, 'credit': credit,
                }
            )
            created_records += 1

        except Exception as e:
            errors.append(f'Строка {total} (счёт {account_number}): {str(e)}')

    return {
        'success': True, 'total': total, 'period': period_date.strftime('%Y-%m-%d'),
        'column_mapping': col, 'format': fmt,
        'accounts': {'created': created_accounts, 'updated': updated_accounts},
        'billing_records': {'created': created_records, 'updated': updated_records},
        'errors': errors,
    }


# ══════════════════════════════════════════════════════════════════
# Preview (предпросмотр первых строк)
# ══════════════════════════════════════════════════════════════════

def preview_excel(file_bytes, max_rows=10):
    """
    Возвращает первые N строк Excel-файла для предпросмотра.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    preview = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        preview.append([str(cell) if cell is not None else '' for cell in row])

    total_rows = ws.max_row - 1 if ws.max_row else 0  # минус заголовок

    return {
        'headers': preview[0] if preview else [],
        'rows': preview[1:] if len(preview) > 1 else [],
        'total_rows': total_rows,
    }
