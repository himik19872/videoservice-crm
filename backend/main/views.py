from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.db.models import Sum
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from datetime import timedelta
from .models import Region, Master, Client, Equipment, Order, OrderHistory, Report, Building, TraccarSettings, TraccarDevice, SystemSettings, UserProfile, WorkShift, OrderMedia, PushToken
from .models import MaxBotSettings, MaxUserLink
from .models import InventoryItem, InventoryMovement, Payment, MasterSalary
from .models import MasterCashDebt, MasterInventoryDebt, OrderMaterial, Message
from .models import LegalEntity, EstimateService, CommercialEstimate, EstimateItem
from .models import Supplier, SupplyInvoice, SupplyInvoiceItem
from .models import IssueOrder, IssueOrderItem, PurchaseRequest, PurchaseRequestItem
from .models import OrderComment
from .models import ErcAccount, ErcBillingRecord
from .models import StorageLocation
from .models import OutgoingInvoice, OutgoingInvoiceItem
from .models import CallLog
from .models import AsteriskSipPeer, AsteriskTrunk, AsteriskRoute, AsteriskIvr, AsteriskIvrOption
from .models import AsteriskVoicemail, AsteriskCallRecording
from .models import BuildingEntrance, ManagementCompany, Tariff, PaymentRecord, BewardDevice, BuildingSystem
from .models import MCContact, MCPayment, MCComment
from django.db.models import Q
from .serializers import (
    RegionSerializer, MasterSerializer, ClientSerializer,
    EquipmentSerializer, OrderSerializer, OrderCreateSerializer,
    OrderStatusUpdateSerializer, ReportSerializer, UserSerializer, LoginSerializer,
    BuildingSerializer, BuildingDetailSerializer, TraccarSettingsSerializer, TraccarDeviceSerializer,
    SystemSettingsSerializer, UserProfileSerializer, WorkShiftSerializer, OrderMediaSerializer, PushTokenSerializer
)
from .serializers import InventoryItemSerializer, InventoryMovementSerializer, PaymentSerializer, MasterSalarySerializer, MessageSerializer
from .serializers import LegalEntitySerializer, EstimateServiceSerializer, CommercialEstimateSerializer, EstimateItemSerializer
from .serializers import SupplierSerializer, SupplyInvoiceSerializer, SupplyInvoiceCreateSerializer, SupplyInvoiceReceiveSerializer
from .serializers import SupplyInvoiceItemSerializer
from .serializers import IssueOrderSerializer, IssueOrderCreateSerializer, IssueOrderItemSerializer
from .serializers import PurchaseRequestSerializer, PurchaseRequestItemSerializer
from .serializers import OrderCommentSerializer
from .serializers import ErcAccountSerializer, ErcBillingRecordSerializer
from .serializers import StorageLocationSerializer, StorageLocationDetailSerializer
from .serializers import OutgoingInvoiceSerializer, OutgoingInvoiceCreateSerializer
from .serializers import CallLogSerializer
from .serializers import AsteriskSipPeerSerializer, AsteriskTrunkSerializer, AsteriskRouteSerializer
from .serializers import AsteriskIvrSerializer, AsteriskIvrOptionSerializer
from .serializers import AsteriskVoicemailSerializer, AsteriskCallRecordingSerializer
from .serializers import BuildingEntranceSerializer, ManagementCompanySerializer, TariffSerializer, PaymentRecordSerializer, BewardDeviceSerializer
from .serializers import BuildingSystemSerializer
from .serializers import MCContactSerializer, MCPaymentSerializer, MCCommentSerializer


class RegionViewSet(viewsets.ModelViewSet):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description']


class MasterViewSet(viewsets.ModelViewSet):
    queryset = Master.objects.all()
    serializer_class = MasterSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['region', 'is_available']
    search_fields = ['user__username', 'user__first_name', 'user__last_name', 'phone']

    def get_queryset(self):
        queryset = Master.objects.all()
        user = self.request.user

        # Мастер видит только свои данные
        if user.is_authenticated and not user.is_staff:
            try:
                master = user.master_profile
                return queryset.filter(id=master.id)
            except Master.DoesNotExist:
                return Master.objects.none()

        return queryset

    def create(self, request, *args, **kwargs):
        from django.contrib.auth.models import User
        import random

        full_name = request.data.get('full_name', '')
        name_parts = full_name.strip().split(' ', 1) if full_name else ['', '']
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ''

        # Берём пароль из запроса или генерируем
        password = request.data.get('password', '').strip()
        if not password:
            password = f"master{random.randint(100, 999)}"

        # Генерируем username из ФИО или случайный
        username = request.data.get('username', '').strip()
        if not username:
            base = (first_name[:1] + last_name).lower() if last_name else first_name.lower()
            username = base
            # Уникализируем, если такой username уже есть
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{base}{counter}"
                counter += 1

        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=first_name,
            last_name=last_name
        )

        # Создаём UserProfile для роли
        UserProfile.objects.get_or_create(
            user=user,
            defaults={'role': 'master', 'phone': request.data.get('phone', '')}
        )

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        master = serializer.save(user=user)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=False, methods=['get'])
    def locations(self, request):
        """Координаты всех мастеров (телефон + Traccar-маяк)"""
        masters = Master.objects.select_related('traccar_device', 'user').all()
        user = request.user
        if not user.is_staff:
            try:
                masters = masters.filter(id=user.master_profile.id)
            except Master.DoesNotExist:
                masters = Master.objects.none()

        result = []
        for master in masters:
            lat = lon = speed = None
            is_online = False
            last_update = None
            loc_source = None

            try:
                device = master.traccar_device
                if device and device.last_latitude is not None:
                    lat, lon, speed = device.last_latitude, device.last_longitude, device.last_speed
                    is_online = device.is_online
                    last_update = device.last_update.isoformat() if device.last_update else None
                    loc_source = 'traccar'
            except Exception:
                pass

            if loc_source is None:
                last_hist = OrderHistory.objects.filter(
                    order__master=master, master_lat__isnull=False, master_lon__isnull=False
                ).order_by('-changed_at').first()
                if last_hist:
                    lat, lon = last_hist.master_lat, last_hist.master_lon
                    last_update = last_hist.changed_at.isoformat() if last_hist.changed_at else None
                    loc_source = 'phone'
                    is_online = (timezone.now() - last_hist.changed_at).total_seconds() < 3600

            if lat is not None and lon is not None:
                result.append({
                    'master_id': master.id, 'master_name': str(master),
                    'lat': lat, 'lon': lon, 'speed': speed,
                    'is_online': is_online, 'last_update': last_update, 'source': loc_source,
                })

        return Response(result)

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """Статистика по мастеру за месяц"""
        master = self.get_object()
        from datetime import date

        month = request.query_params.get('month', date.today().strftime('%Y-%m'))
        year, month_num = map(int, month.split('-'))

        orders = Order.objects.filter(
            master=master,
            created_at__year=year,
            created_at__month=month_num
        )

        completed = orders.filter(status='completed')
        now = timezone.now()
        overdue = orders.filter(status__in=['assigned', 'accepted', 'in_progress', 'paused', 'need_help'], deadline__lt=now)

        # Среднее время выполнения (в часах)
        avg_hours = 0
        count_completed = completed.count()
        if count_completed > 0:
            total_seconds = sum(
                (o.completed_at - o.started_at).total_seconds()
                for o in completed
                if o.started_at and o.completed_at
            )
            avg_hours = round(total_seconds / 3600 / count_completed, 1)

        total_cost = completed.aggregate(s=Sum('cost'))['s'] or 0

        by_type = {}
        for t in ['repair', 'connection', 'sale']:
            by_type[t] = orders.filter(order_type=t).count()

        return Response({
            'master_id': master.id,
            'master_name': str(master),
            'total_orders': orders.count(),
            'completed_orders': count_completed,
            'overdue_orders': overdue.count(),
            'avg_completion_hours': avg_hours,
            'total_cost': float(total_cost),
            'by_type': by_type,
            'month': month,
        })

    @action(detail=True, methods=['post'])
    def update_gps(self, request, pk=None):
        """Запросить актуальные GPS-координаты мастера из Traccar"""
        import requests
        master = self.get_object()
        try:
            device = master.traccar_device
        except Exception:
            return Response({'error': 'У мастера не привязан GPS-трекер'}, status=400)

        settings = TraccarSettings.objects.first()
        if not settings or not settings.is_active:
            return Response({'error': 'Интеграция Traccar не активна'}, status=400)

        try:
            resp = requests.get(
                f"{settings.server_url.rstrip('/')}/api/positions",
                auth=(settings.username, settings.password),
                params={'deviceId': device.internal_device_id},
                timeout=10
            )
            if resp.status_code == 200 and resp.json():
                pos = resp.json()[0]
                device.last_latitude = pos.get('latitude')
                device.last_longitude = pos.get('longitude')
                device.last_speed = round(pos.get('speed', 0) * 1.852, 1) if pos.get('speed') else None
                device.last_update = timezone.now()
                device.is_online = True
                device.save()
                return Response({
                    'ok': True,
                    'latitude': device.last_latitude,
                    'longitude': device.last_longitude,
                    'speed': device.last_speed,
                    'is_online': device.is_online,
                    'last_update': device.last_update.isoformat() if device.last_update else None,
                })
            return Response({'error': 'Нет данных с устройства', 'is_online': False}, status=404)
        except Exception as e:
            return Response({'error': f'Ошибка связи с Traccar: {e}'}, status=500)


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['region', 'is_legal', 'legal_type']
    search_fields = ['name', 'phone', 'email', 'address', 'inn']

    def get_queryset(self):
        queryset = Client.objects.all()
        user = self.request.user

        # Фильтр по источнику
        source = self.request.query_params.get('source', '').strip()
        if source:
            queryset = queryset.filter(source=source)

        # Фильтр «без дома»
        no_building = self.request.query_params.get('no_building', '').strip()
        if no_building == 'true':
            queryset = queryset.filter(building__isnull=True)

        # Если не администратор, показываем только клиентов своего региона
        if user.is_authenticated and not user.is_staff:
            try:
                master = user.master_profile
                if master.region:
                    return queryset.filter(region=master.region)
            except Master.DoesNotExist:
                pass

        return queryset

    @action(detail=False, methods=['get'])
    def autocomplete(self, request):
        """
        Быстрый поиск клиентов для автокомплита.
        Ищет по: адресу, названию юрлица, ИНН, ФИО, телефону.
        Параметр: ?q=поисковый_запрос (минимум 2 символа)
        Возвращает до 15 результатов.
        """
        q = request.query_params.get('q', '').strip()
        if len(q) < 2:
            return Response([])

        from django.db.models import Q

        clients = Client.objects.filter(
            Q(name__icontains=q) |
            Q(phone__icontains=q) |
            Q(address__icontains=q) |
            Q(inn__icontains=q) |
            Q(legal_address__icontains=q)
        ).order_by('name')[:15]

        # Возвращаем только нужные поля для автокомплита
        data = []
        for c in clients:
            label = c.name
            sub = c.address or c.phone or ''
            if c.is_legal:
                label = f'{c.name} (ИНН {c.inn})' if c.inn else c.name
                sub = c.legal_address or c.address or ''

            data.append({
                'id': c.id,
                'label': label,
                'sub': sub,
                'name': c.name,
                'phone': c.phone,
                'address': c.address,
                'inn': c.inn,
                'is_legal': c.is_legal,
                'personal_account_number': c.personal_account_number,
            })

        return Response(data)
    @action(detail=False, methods=['get'])
    def street_autocomplete(self, request):
        """Умный поиск: улица → номера домов из базы.
        ?q=Стрельнин — вернёт улицы; ?q=Стрельнинское&house=6 — дома."""
        q = request.query_params.get('q', '').strip()
        house_filter = request.query_params.get('house', '').strip()
        if not q or len(q) < 2:
            return Response([])

        if house_filter:
            buildings = Building.objects.filter(
                street_name__icontains=q, house_number__icontains=house_filter
            ).order_by('street_name', 'house_number').distinct('street_name', 'house_number', 'building_number')[:20]
            result, seen = [], set()
            for b in buildings:
                key = f'{b.house_number}|{b.building_number}'
                if key not in seen:
                    seen.add(key)
                    label = f'д. {b.house_number}'
                    if b.building_number:
                        label += f' корп. {b.building_number}'
                    result.append({'id': b.id, 'label': label, 'sub': f'{b.street_name}'})
            return Response(result)

        buildings = Building.objects.filter(street_name__icontains=q).order_by('street_name').distinct('street_name')[:12]
        result, seen = [], set()
        for b in buildings:
            if b.street_name not in seen:
                seen.add(b.street_name)
                cnt = Building.objects.filter(street_name=b.street_name).count()
                result.append({'id': b.id, 'label': b.street_name, 'sub': f'{cnt} домов', 'street': b.street_name})
        return Response(result)




    @action(detail=True, methods=['get'])
    def erc_payments(self, request, pk=None):
        """Возвращает историю платежей ЕРЦ для клиента (по personal_account_number)."""
        client = self.get_object()
        if not client.personal_account_number:
            return Response([])

        from .models import ErcBillingRecord
        records = ErcBillingRecord.objects.filter(
            account__account_number=client.personal_account_number
        ).order_by('-period')

        from .serializers import ErcBillingRecordSerializer
        return Response(ErcBillingRecordSerializer(records, many=True).data)


    @action(detail=False, methods=['get'])
    def address_suggest(self, request):
        """Адресные подсказки через DaData. Параметр: ?q=улица Ленина"""
        import requests
        q = request.query_params.get('q', '').strip()
        if len(q) < 3:
            return Response([])

        settings = SystemSettings.objects.first()
        token = settings.dadata_token if settings else ''
        if not token:
            return Response([])

        try:
            resp = requests.post(
                'https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/address',
                json={'query': q, 'count': 7},
                headers={
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                    'Authorization': f'Token {token}'
                },
                timeout=5
            )
            data = resp.json()
            suggestions = []
            for s in data.get('suggestions', []):
                suggestions.append({
                    'value': s.get('value', ''),
                    'unrestricted_value': s.get('unrestricted_value', ''),
                    'data': {
                        'city': s.get('data', {}).get('city', ''),
                        'street': s.get('data', {}).get('street', ''),
                        'house': s.get('data', {}).get('house', ''),
                        'flat': s.get('data', {}).get('flat', ''),
                    }
                })
            return Response(suggestions)
        except Exception:
            return Response([])



class EquipmentViewSet(viewsets.ModelViewSet):
    queryset = Equipment.objects.all()
    serializer_class = EquipmentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['client', 'status', 'equipment_type']
    search_fields = ['name', 'serial_number']

    def get_queryset(self):
        queryset = Equipment.objects.all()
        user = self.request.user

        # Если не администратор, показываем только оборудование клиентов своего региона
        if user.is_authenticated and not user.is_staff:
            try:
                master = user.master_profile
                if master.region:
                    return queryset.filter(client__region=master.region)
            except Master.DoesNotExist:
                pass

        return queryset


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'order_type', 'priority', 'master', 'region', 'client']
    search_fields = ['number', 'address', 'description', 'city', 'street_name', 'client__first_name', 'client__last_name']
    ordering_fields = ['created_at', 'priority', 'status', 'cost', 'master__user__last_name', 'region__name', 'address', 'city', 'street_name', 'number', 'order_type']

    def get_queryset(self):
        queryset = Order.objects.all()
        user = self.request.user

        if not user.is_authenticated:
            return Order.objects.none()

        try:
            profile = user.profile
        except UserProfile.DoesNotExist:
            return Order.objects.none()

        role = profile.role

        # Админ, диспетчер, главный инженер, техдиректор, исполнительный, гендиректор — видят все заявки
        if role in ('admin', 'dispatcher', 'chief_engineer', 'tech_director', 'executive_director', 'general_director'):
            return queryset

        # Мастер, монтажник — только свои
        if role in ('master', 'installer'):
            try:
                master = user.master_profile
                return queryset.filter(Q(master=master) | Q(helpers=user)).distinct()
            except Master.DoesNotExist:
                return Order.objects.none()

        # Инженер, начальник сервисной службы — заявки где нужна помощь или они помощники
        if role in ('engineer', 'supervisor'):
            return queryset.filter(
                Q(status='need_help') | Q(master__user=user) | Q(helpers=user)
            )

        # Помощники в любых ролях видят свои заявки
        return queryset.filter(Q(helpers=user) | Q(master__user=user)).distinct()

    def create(self, request, *args, **kwargs):
        serializer = OrderCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        order = serializer.save()

        # Создаем запись в истории
        OrderHistory.objects.create(
            order=order,
            changed_by=request.user,
            old_status='',
            new_status=order.status
        )

        # Уведомляем диспетчеров и админов о новой заявке
        from django.contrib.auth.models import User
        staff_users = User.objects.filter(
            profile__role__in=['admin', 'dispatcher', 'secretary', 'operator']
        ).exclude(id=request.user.id)
        for u in staff_users:
            try:
                send_push_notification(
                    u.id,
                    '🆕 Новая заявка',
                    f'#{order.number} — {order.get_order_type_display()}, {order.address or order.full_address}'
                )
            except Exception:
                pass

        return Response(OrderSerializer(order).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = OrderStatusUpdateSerializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        new_status = request.data.get('status', instance.status)
        user = request.user

        # Мастер не может подтверждать заявку — только диспетчер/админ
        if new_status == 'confirmed' and not user.is_staff:
            return Response(
                {'error': 'Только диспетчер или администратор может подтвердить выполнение заявки'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Мастер не может редактировать выполненную заявку
        if instance.status == 'completed' and not user.is_staff:
            return Response(
                {'error': 'Выполненная заявка недоступна для редактирования. Ожидайте подтверждения диспетчера.'},
                status=status.HTTP_403_FORBIDDEN
            )

        old_status = instance.status
        old_master_id = instance.master_id
        notes = request.data.get('notes', '')
        order = serializer.save()

        # Если статус изменился, создаем запись в истории
        if old_status != order.status:
            # Авто-захват GPS из трекера мастера, если не переданы явно
            master_lat = request.data.get('master_lat')
            master_lon = request.data.get('master_lon')
            if (not master_lat or not master_lon) and order.master:
                try:
                    device = order.master.traccar_device
                    if device and device.last_latitude is not None:
                        master_lat = device.last_latitude
                        master_lon = device.last_longitude
                except Exception:
                    pass

            OrderHistory.objects.create(
                order=order,
                changed_by=request.user,
                old_status=old_status,
                new_status=order.status,
                notes=notes,
                master_lat=master_lat,
                master_lon=master_lon,
            )

            now = timezone.now()
            status_time_map = {
                'assigned': 'assigned_at',
                'accepted': 'accepted_at',
                'in_progress': 'started_at',
                'paused': 'paused_at',
                'completed': 'completed_at',
                'confirmed': 'confirmed_at',
            }

            # Если переводим в completed и foto_report_required — проверяем медиа
            if order.status == 'completed' and order.photo_report_required:
                if not order.media.exists():
                    return Response(
                        {'error': 'Для этой заявки требуется фото/видео отчёт. Прикрепите файлы перед завершением.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            field = status_time_map.get(order.status)
            if field and not getattr(order, field):
                setattr(order, field, now)

            # Если на паузе или завершена — обязательный комментарий
            if order.status == 'paused' and not notes:
                return Response(
                    {'error': 'Для статуса "На паузе" обязательно укажите причину (notes)'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if order.status == 'completed' and not notes:
                return Response(
                    {'error': 'Для завершения заявки обязательно оставьте комментарий о проделанной работе'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # GPS-позиция мастера при смене статуса
            master_lat = request.data.get('master_lat')
            master_lon = request.data.get('master_lon')

            # Если заявка подтверждена — запоминаем, кто подтвердил
            if order.status == 'confirmed' and not order.confirmed_by:
                order.confirmed_by = request.user

            order.save()

        # Каскадное закрытие: если заявка завершена/отменена/подтверждена → дочерние тоже
        if order.status in ('completed', 'confirmed', 'cancelled') and order.linked_orders.exists():
            linked = order.linked_orders.exclude(status__in=('completed', 'confirmed', 'cancelled'))
            for child in linked:
                child.status = order.status
                now = timezone.now()
                if order.status == 'completed':
                    child.completed_at = now
                elif order.status == 'confirmed':
                    child.confirmed_at = now
                    child.confirmed_by = request.user
                child.save()
                OrderHistory.objects.create(
                    order=child, changed_by=request.user,
                    old_status=old_status, new_status=order.status,
                    notes=f'Автоматически: главная заявка #{order.number} закрыта'
                )
            print(f'[Orders] Cascaded close: {linked.count()} children for #{order.number}')

        # Если переназначили мастера
        if order.master_id and order.master_id != old_master_id:
            old_master_name = str(Master.objects.get(id=old_master_id)) if old_master_id else '—'
            OrderHistory.objects.create(
                order=order,
                changed_by=request.user,
                old_status=old_status,
                new_status=order.status,
                notes=f'Мастер переназначен: {old_master_name} → {order.master}'
            )

        return Response(OrderSerializer(order).data)

    @action(detail=False, methods=['get'])
    def calendar(self, request):
        """Календарь: заявки с датой scheduled_at"""
        user = request.user
        qs = Order.objects.filter(scheduled_at__isnull=False).exclude(status__in=['cancelled', 'confirmed'])
        if not user.is_staff:
            try:
                master = user.master_profile
                qs = qs.filter(master=master)
            except Master.DoesNotExist:
                qs = qs.filter(helpers=user)
        return Response([{
            'id': o.id, 'number': o.number, 'order_type': o.order_type,
            'status': o.status,
            'address': o.address or o.full_address,
            'master': o.master.user.get_full_name() or o.master.user.username if o.master else None,
            'client': o.client.name if o.client else None,
            'scheduled_at': o.scheduled_at.isoformat() if o.scheduled_at else None,
            'priority': o.priority,
        } for o in qs.order_by('scheduled_at')])

    @action(detail=True, methods=['post'])
    def link_orders(self, request, pk=None):
        """Объединить дочерние заявки в эту (главную)."""
        order = self.get_object()
        child_ids = request.data.get('child_ids', [])
        reason = request.data.get('reason', '')
        if not child_ids:
            return Response({'error': 'Укажите child_ids — список ID заявок для объединения'}, status=400)

        # Привязываем заявки к этой
        linked = Order.objects.filter(id__in=child_ids).exclude(id=order.id).exclude(
            parent_order__isnull=False  # уже привязанные не трогаем
        )

        # Пишем историю в каждую дочернюю заявку
        for child in linked:
            OrderHistory.objects.create(
                order=child, changed_by=request.user,
                old_status=child.status, new_status=child.status,
                notes=f'🔗 Объединена с заявкой #{order.number}' + (f': {reason}' if reason else '')
            )
        # История в главной
        child_nums = ', '.join(f'#{o.number}' for o in linked)
        OrderHistory.objects.create(
            order=order, changed_by=request.user,
            old_status=order.status, new_status=order.status,
            notes=f'🔗 Объединены заявки: {child_nums}' + (f' — {reason}' if reason else '')
        )

        count = linked.update(parent_order=order)

        # Назначаем того же мастера, если у главной есть
        if order.master:
            linked.filter(master__isnull=True).update(master=order.master)

        return Response({'ok': True, 'linked_count': count})

    @action(detail=True, methods=['post'])
    def unlink_orders(self, request, pk=None):
        """Отвязать все дочерние заявки от этой."""
        order = self.get_object()
        children = list(order.linked_orders.all())
        for child in children:
            OrderHistory.objects.create(
                order=child, changed_by=request.user,
                old_status=child.status, new_status=child.status,
                notes=f'🔗 Отвязана от заявки #{order.number}'
            )
        OrderHistory.objects.create(
            order=order, changed_by=request.user,
            old_status=order.status, new_status=order.status,
            notes=f'🔗 Дочерние заявки отвязаны ({len(children)} шт.)'
        )
        count = order.linked_orders.update(parent_order=None)
        return Response({'ok': True, 'unlinked_count': count})

    @action(detail=False, methods=['get'])
    def find_similar(self, request):
        """Поиск похожих открытых заявок по дому/подъезду."""
        building_id = request.query_params.get('building_id')
        entrance = request.query_params.get('entrance', '')
        city = request.query_params.get('city', '')
        street = request.query_params.get('street_name', '')
        house = request.query_params.get('house_number', '')

        if not building_id and not (city and street and house):
            return Response([], status=200)

        qs = Order.objects.exclude(status__in=['completed', 'confirmed', 'cancelled']).exclude(
            parent_order__isnull=False
        )

        if building_id:
            qs = qs.filter(building_id=building_id)
        elif city and street and house:
            qs = qs.filter(city__iexact=city, street_name__iexact=street, house_number__iexact=house)

        if entrance:
            qs = qs.filter(entrance=entrance)

        results = qs.order_by('-created_at')[:10]
        return Response([{
            'id': o.id, 'number': o.number, 'status': o.status,
            'order_type': o.get_order_type_display(), 'address': o.address,
            'description': o.description[:100], 'apartment': o.apartment,
            'created_at': o.created_at.isoformat() if o.created_at else None,
        } for o in results])

    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """Назначить заявку сотруднику (мастер, монтажник, инженер и др.)"""
        order = self.get_object()
        master_id = request.data.get('master_id')  # старый параметр
        user_id = request.data.get('user_id')  # новый параметр — любой сотрудник
        scheduled_at = request.data.get('scheduled_at')

        master = None
        assigned_name = 'Неизвестный'

        # Поддержка user_id (любой сотрудник)
        if user_id:
            try:
                assigned_user = User.objects.get(id=user_id)
                # Ищем или создаём Master для этого пользователя
                master, _ = Master.objects.get_or_create(
                    user=assigned_user,
                    defaults={'phone': getattr(assigned_user.profile, 'phone', ''), 'region': Region.objects.first()}
                )
                assigned_name = assigned_user.get_full_name() or assigned_user.username
            except User.DoesNotExist:
                return Response({'error': 'Сотрудник не найден'}, status=404)
        elif master_id:
            try:
                master = Master.objects.get(id=master_id)
                assigned_name = str(master)
            except Master.DoesNotExist:
                return Response({'error': 'Мастер не найден'}, status=404)
        else:
            return Response({'error': 'Не указан ID сотрудника (user_id или master_id)'}, status=400)

        old_status = order.status
        order.master = master
        order.status = 'assigned'
        order.assigned_at = timezone.now()

        if scheduled_at:
            order.scheduled_at = scheduled_at
        order.save()

        OrderHistory.objects.create(
            order=order, changed_by=request.user,
            old_status=old_status, new_status=order.status,
            notes=f'Заявка назначена: {assigned_name}'
        )

        # Уведомление сотруднику
        if master:
            send_push_notification(master.user_id,
                'Новая заявка',
                f'#{order.number} — {order.get_order_type_display()}, {order.address}')

        if order.client:
            try:
                from .max_service import notify_client_order_assigned
                notify_client_order_assigned(
                    client_id=order.client_id,
                    order_number=order.number,
                    order_type=order.get_order_type_display(),
                    address=order.full_address,
                    master_name=assigned_name,
                    master_phone=master.phone if master else 'не указан',
                )
            except Exception:
                pass

        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=['post'])
    def notify_master(self, request, pk=None):
        """Повторно отправить уведомление мастеру (напоминание)."""
        order = self.get_object()
        if not order.master:
            return Response({'error': 'У заявки нет назначенного сотрудника'}, status=400)
        if order.status in ('completed', 'confirmed', 'cancelled'):
            return Response({'error': 'Заявка уже закрыта'}, status=400)

        send_push_notification(
            order.master.user_id,
            '📣 Напоминание о заявке',
            f'#{order.number} — {order.get_order_type_display()}, {order.address}',
            data={'order_id': order.id},
        )
        return Response({'ok': True, 'message': f'Уведомление отправлено мастеру {order.master}'})

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Начать выполнение заявки"""
        order = self.get_object()

        if order.status != 'assigned':
            return Response(
                {'error': 'Нельзя начать заявку, которая не назначена'},
                status=status.HTTP_400_BAD_REQUEST
            )

        old_status = order.status
        order.status = 'in_progress'
        order.started_at = timezone.now()
        order.save()

        OrderHistory.objects.create(
            order=order,
            changed_by=request.user,
            old_status=old_status,
            new_status=order.status,
            notes='Начато выполнение заявки'
        )

        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Завершить заявку"""
        order = self.get_object()

        if order.status not in ['assigned', 'in_progress']:
            return Response({'error': 'Нельзя завершить заявку в текущем статусе'}, status=400)

        old_status = order.status
        order.status = 'completed'
        order.completed_at = timezone.now()
        order.save()

        OrderHistory.objects.create(
            order=order, changed_by=request.user,
            old_status=old_status, new_status=order.status,
            notes=request.data.get('notes', 'Заявка выполнена')
        )

        # Клиенту НЕ отправляем — ждём подтверждения диспетчером (confirm)

        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=['post'])
    def receive_payment(self, request, pk=None):
        """Мастер принимает оплату от клиента (нал/безнал/перевод)"""
        order = self.get_object()
        amount = request.data.get('amount')
        method = request.data.get('payment_method', 'cash')
        if not amount:
            return Response({'error': 'Укажите amount'}, status=400)
        try:
            master = request.user.master_profile
        except Exception:
            return Response({'error': 'Только мастер может принимать оплату'}, status=403)
        # Создаём платёж
        payment = Payment.objects.create(
            order=order, amount=float(amount), payment_method=method,
            collected_by_master=master, collected_by_master_at=timezone.now(),
            is_submitted_to_office=False, received_by=request.user
        )
        # Если наличные — создаём долг мастеру
        if method == 'cash':
            MasterCashDebt.objects.create(
                master=master, order=order, amount=float(amount),
                notes=f'Наличные от клиента по заявке {order.number}'
            )
        return Response({'ok': True, 'payment_id': payment.id})

    @action(detail=True, methods=['post'])
    def submit_cash(self, request, pk=None):
        """Мастер сдал наличные в кассу"""
        debt = MasterCashDebt.objects.filter(order_id=pk, is_paid_to_office=False).first()
        if not debt:
            return Response({'error': 'Нет неоплаченного долга по этой заявке'}, status=404)
        debt.is_paid_to_office = True
        debt.paid_to_office_at = timezone.now()
        debt.accepted_by = request.user
        debt.save()
        return Response({'ok': True})

    @action(detail=True, methods=['post'])
    def return_item(self, request, pk=None):
        """Мастер возвращает сломанное/старое оборудование на склад"""
        debt = MasterInventoryDebt.objects.filter(order_id=pk, is_returned=False).first()
        if not debt:
            return Response({'error': 'Нет невозвращённого оборудования по этой заявке'}, status=404)
        debt.is_returned = True
        debt.returned_at = timezone.now()
        debt.accepted_by = request.user
        debt.condition = request.data.get('condition', 'broken')
        debt.save()
        return Response({'ok': True})

    @action(detail=True, methods=['post'])
    def helpers(self, request, pk=None):
        """Добавить/убрать помощников к заявке"""
        order = self.get_object()
        helper_ids = request.data.get('helper_ids', [])
        if not helper_ids:
            return Response({'error': 'Укажите helper_ids (список ID пользователей)'}, status=400)
        from django.contrib.auth.models import User as AuthUser
        users = AuthUser.objects.filter(id__in=helper_ids)
        order.helpers.set(users)
        return Response({
            'ok': True,
            'helpers': [{'id': u.id, 'username': u.username, 'full_name': u.get_full_name() or u.username} for u in users]
        })

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Отменить заявку"""
        order = self.get_object()
        if order.status == 'cancelled':
            return Response({'error': 'Заявка уже отменена'}, status=status.HTTP_400_BAD_REQUEST)

        old_status = order.status
        order.status = 'cancelled'
        order.save()
        OrderHistory.objects.create(
            order=order, changed_by=request.user,
            old_status=old_status, new_status=order.status,
            notes=request.data.get('notes', 'Заявка отменена')
        )
        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """Диспетчер/админ подтверждает выполнение заявки"""
        order = self.get_object()
        if not request.user.is_staff:
            return Response({'error': 'Только диспетчер или администратор может подтвердить заявку'}, status=403)

        if order.status != 'completed':
            return Response({'error': 'Подтвердить можно только выполненную заявку'}, status=400)

        order.status = 'confirmed'
        order.confirmed_at = timezone.now()
        order.confirmed_by = request.user
        order.save()
        OrderHistory.objects.create(
            order=order, changed_by=request.user,
            old_status='completed', new_status='confirmed',
            notes=request.data.get('notes', 'Заявка подтверждена диспетчером')
        )

        # Списываем материалы с мастера — закрываем расходные ордера по заявке
        from .models import IssueOrder, IssueOrderItem, InventoryItem, InventoryMovement
        for io in order.issue_orders.filter(status__in=['pending', 'received', 'partially_used']):
            for ioi in io.items.all():
                if ioi.quantity_used == 0:
                    ioi.quantity_used = ioi.quantity_issued
                    ioi.save()
                if ioi.quantity_used > 0:
                    ioi.inventory_item.status = 'installed'
                    ioi.inventory_item.save(update_fields=['status', 'updated_at'])
                    InventoryMovement.objects.create(
                        item=ioi.inventory_item,
                        movement_type='installed',
                        quantity=ioi.quantity_used,
                        master=io.master,
                        order=order,
                        performed_by=request.user,
                        notes=f'Списание по заявке #{order.number}'
                    )
            io.status = 'fully_used'
            io.completed_at = timezone.now()
            io.save()

        # Max-уведомление клиенту: заявка подтверждена диспетчером
        if order.client:
            try:
                from .max_service import notify_client_order_confirmed
                media_files = order.media.all()
                has_photos = any(m.file_type == 'image' for m in media_files)
                has_videos = any(m.file_type == 'video' for m in media_files)
                master_name = order.master.user.get_full_name() or order.master.user.username if order.master else '—'
                notify_client_order_confirmed(
                    client_id=order.client_id,
                    order_number=order.number,
                    order_type=order.get_order_type_display(),
                    address=order.full_address,
                    has_photos=has_photos,
                    has_videos=has_videos,
                    master_name=master_name,
                )
            except Exception:
                pass

        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=['post'])
    def rework(self, request, pk=None):
        """Диспетчер/админ возвращает заявку в работу (клиент недоволен)"""
        order = self.get_object()
        if not request.user.is_staff:
            return Response({'error': 'Только диспетчер или администратор может вернуть заявку'}, status=403)

        if order.status != 'completed':
            return Response({'error': 'Вернуть можно только выполненную заявку'}, status=400)

        notes = request.data.get('notes', '')
        if not notes:
            return Response({'error': 'Укажите причину возврата (notes) — что нужно доделать'}, status=400)

        order.status = 'in_progress'
        order.completed_at = None
        order.save()
        OrderHistory.objects.create(
            order=order, changed_by=request.user,
            old_status='completed', new_status='in_progress',
            notes=f'Возвращено в работу. Причина: {notes}'
        )
        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=['get'])
    def gps_history(self, request, pk=None):
        """GPS-история мастера по заявке (координаты на момент смены статусов)"""
        order = self.get_object()
        if not order.master:
            return Response({'error': 'У заявки нет назначенного мастера'}, status=400)
        try:
            device = order.master.traccar_device
        except Exception:
            return Response({'error': 'У мастера не привязан GPS-трекер'}, status=400)

        # Собираем точки: координаты мастера на каждый момент истории
        history = OrderHistory.objects.filter(order=order).order_by('changed_at')
        result = []
        for h in history:
            result.append({
                'status': h.new_status,
                'changed_at': h.changed_at.isoformat() if h.changed_at else None,
                'lat': h.master_lat,
                'lon': h.master_lon,
            })

        # Текущая позиция мастера
        current = {
            'lat': device.last_latitude,
            'lon': device.last_longitude,
            'speed': device.last_speed,
            'is_online': device.is_online,
            'last_update': device.last_update.isoformat() if device.last_update else None,
        }

        return Response({'history': result, 'current': current})

    @action(detail=True, methods=['get', 'post'])
    def comments(self, request, pk=None):
        """Комментарии/диалог внутри заявки"""
        order = self.get_object()

        if request.method == 'GET':
            qs = order.comments.select_related('author').order_by('created_at')
            return Response(OrderCommentSerializer(qs, many=True).data)

        # POST — добавить комментарий
        text = request.data.get('text', '').strip()
        if not text:
            return Response({'error': 'Текст комментария обязателен'}, status=status.HTTP_400_BAD_REQUEST)

        event_type = request.data.get('event_type', 'comment')
        comment = OrderComment.objects.create(
            order=order,
            author=request.user,
            text=text,
            event_type=event_type
        )

        # Уведомляем всех, кто вовлечён в заявку (кроме автора)
        recipients = set()
        if order.master and order.master.user_id != request.user.id:
            recipients.add(order.master.user_id)
        for h in order.helpers.exclude(id=request.user.id):
            recipients.add(h.id)
        # Диспетчеры и админы тоже получают уведомление
        from django.contrib.auth.models import User as AuthUser
        staff = AuthUser.objects.filter(
            profile__role__in=['admin', 'dispatcher']
        ).exclude(id=request.user.id).values_list('id', flat=True)
        recipients.update(staff)

        for uid in recipients:
            try:
                send_push_notification(
                    uid,
                    f'💬 Заявка #{order.number}',
                    f'{request.user.get_full_name() or request.user.username}: {text[:100]}'
                )
            except Exception:
                pass

        return Response(OrderCommentSerializer(comment).data, status=status.HTTP_201_CREATED)


class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['report_type', 'status']
    search_fields = ['title']

    def get_queryset(self):
        queryset = Report.objects.all()
        user = self.request.user

        # Администратор и диспетчер видят все отчёты, мастер — только свои
        if user.is_authenticated:
            if user.is_staff:
                return queryset
            return queryset.filter(created_by=user)

        return queryset.none()

    @action(detail=False, methods=['post'])
    def generate_daily(self, request):
        """Сгенерировать ежедневный отчет"""
        today = timezone.localdate()
        yesterday = today - timedelta(days=1)

        # Подсчет статистики
        orders_today = Order.objects.filter(created_at__date=today)
        orders_yesterday = Order.objects.filter(created_at__date=yesterday)
        completed_today = Order.objects.filter(status='completed', completed_at__date=today)
        new_orders = Order.objects.filter(status='new')
        in_progress = Order.objects.filter(status='in_progress')

        report_data = {
            'date': today.isoformat(),
            'summary': {
                'total_orders_today': orders_today.count(),
                'completed_today': completed_today.count(),
                'new_orders': new_orders.count(),
                'in_progress': in_progress.count(),
            },
            'orders': {
                'today': [
                    {
                        'number': order.number,
                        'type': order.get_order_type_display(),
                        'client': order.client.name,
                        'status': order.get_status_display()
                    }
                    for order in orders_today[:10]
                ],
                'completed_today': [
                    {
                        'number': order.number,
                        'completed_at': order.completed_at.isoformat() if order.completed_at else None
                    }
                    for order in completed_today
                ],
            }
        }

        report = Report.objects.create(
            title=f'Ежедневный отчет {today}',
            report_type='daily',
            period_start=today,
            period_end=today,
            data=report_data,
            created_by=request.user
        )

        return Response(ReportSerializer(report).data)

    @action(detail=False, methods=['post'])
    def generate_master_performance(self, request):
        """Сгенерировать отчет по производительности мастеров"""
        from datetime import date

        period_start = request.data.get('period_start', date.today().isoformat())
        period_end = request.data.get('period_end', date.today().isoformat())

        masters = Master.objects.all()
        master_stats = []

        for master in masters:
            orders = Order.objects.filter(
                master=master,
                created_at__date__gte=period_start,
                created_at__date__lte=period_end
            )

            completed = orders.filter(status='completed').count()
            total = orders.count()
            avg_completion_time = 0

            if completed > 0:
                from django.db.models import Avg
                completed_orders = orders.filter(completed_at__isnull=False)
                avg_completion_time = completed_orders.aggregate(
                    avg_time=Avg('completed_at')
                )['avg_time']

            master_stats.append({
                'master_id': master.id,
                'master_name': str(master),
                'total_orders': total,
                'completed_orders': completed,
                'completion_rate': (completed / total * 100) if total > 0 else 0,
                'avg_completion_time_hours': avg_completion_time
            })

        report_data = {
            'period_start': period_start,
            'period_end': period_end,
            'master_stats': master_stats
        }

        report = Report.objects.create(
            title=f'Отчет по производительности мастеров ({period_start} - {period_end})',
            report_type='custom',
            period_start=period_start,
            period_end=period_end,
            data=report_data,
            created_by=request.user
        )

        return Response(ReportSerializer(report).data)

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Сводка: мастера + диспетчеры"""
        today = timezone.localdate()
        month_start = today.replace(day=1)

        masters = Master.objects.all()
        master_stats = []
        for master in masters:
            orders = Order.objects.filter(master=master)
            month_orders = orders.filter(created_at__date__gte=month_start)
            total = month_orders.count()
            accepted = month_orders.filter(status__in=['accepted', 'in_progress', 'paused', 'need_help', 'completed', 'confirmed']).count()
            completed = month_orders.filter(status__in=['completed', 'confirmed']).count()
            in_work = orders.filter(status__in=['assigned', 'accepted', 'in_progress', 'paused', 'need_help']).count()
            try:
                from .models import TraccarMileage
                mileage = TraccarMileage.objects.filter(master=master, synced_at__date__gte=month_start).aggregate(s=Sum('distance_km'))['s'] or 0
            except Exception:
                mileage = 0
            master_stats.append({
                'master_id': master.id, 'master_name': str(master),
                'total_month': total, 'accepted': accepted, 'completed': completed,
                'in_work': in_work, 'completion_rate': round(completed / total * 100, 1) if total > 0 else 0,
                'mileage_km': round(mileage, 1),
            })

        dps = UserProfile.objects.filter(role='dispatcher')
        dp_stats = []
        for dp in dps:
            created = Order.objects.filter(confirmed_by=dp.user, created_at__date__gte=month_start).count()
            confirmed = Order.objects.filter(confirmed_by=dp.user, status='confirmed', confirmed_at__date__gte=month_start).count()
            dp_stats.append({'dispatcher_id': dp.user.id, 'name': str(dp.user), 'created_orders': created, 'confirmed_orders': confirmed})

        return Response({'date': today.isoformat(), 'masters': master_stats, 'dispatchers': dp_stats})


class BuildingViewSet(viewsets.ModelViewSet):
    """Дома (обслуживаемые адреса)"""
    queryset = Building.objects.all()
    serializer_class = BuildingSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['region', 'city', 'street_type', 'equipment_type']
    search_fields = ['street_name', 'house_number', 'building_number', 'city', 'notes']
    ordering_fields = ['street_name', 'house_number', 'created_at']
    pagination_class = None  # все 1825 домов сразу

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return BuildingDetailSerializer
        return BuildingSerializer

    def get_queryset(self):
        queryset = Building.objects.all()
        user = self.request.user

        # Мастер видит дома своего региона
        if user.is_authenticated and not user.is_staff:
            try:
                master = user.master_profile
                if master.region:
                    return queryset.filter(region=master.region)
            except Master.DoesNotExist:
                pass
        return queryset

    @action(detail=True, methods=['post'])
    def set_management_company(self, request, pk=None):
        """Сменить УК/ТСЖ для дома."""
        building = self.get_object()
        mc_id = request.data.get('management_company_fk')
        if mc_id:
            mc = get_object_or_404(ManagementCompany, id=mc_id)
            building.management_company_fk = mc
            building.management_company = mc.name
        else:
            building.management_company_fk = None
            building.management_company = request.data.get('management_company', building.management_company)
        building.save()
        return Response(BuildingSerializer(building).data)

    @action(detail=True, methods=['get', 'post'])
    def systems(self, request, pk=None):
        """Системы дома (домофон, видеонаблюдение...). GET/POST/PATCH."""
        building = self.get_object()
        if request.method == 'GET':
            qs = building.systems.filter(is_active=True)
            return Response(BuildingSystemSerializer(qs, many=True).data)
        
        data = request.data.copy()
        data['building'] = building.id
        system_type = data.get('system_type')
        tariff_id = data.get('tariff')
        system_id = data.get('id')

        monthly = data.get('monthly_amount')
        if tariff_id and (not monthly or float(monthly) == 0):
            tariff = get_object_or_404(Tariff, id=tariff_id)
            monthly = float(tariff.amount) * building.apartments_count
        elif not monthly:
            monthly = 0

        if system_id:
            bs = get_object_or_404(BuildingSystem, id=system_id, building=building)
            if tariff_id:
                bs.tariff_id = tariff_id
            bs.monthly_amount = monthly
            if data.get('notes') is not None:
                bs.notes = data['notes']
            bs.save()
            return Response(BuildingSystemSerializer(bs).data)

        if system_type:
            bs, _ = BuildingSystem.objects.update_or_create(
                building=building, system_type=system_type,
                defaults={'tariff_id': tariff_id, 'monthly_amount': monthly,
                          'notes': data.get('notes', ''), 'is_active': True}
            )
            return Response(BuildingSystemSerializer(bs).data)
        return Response({'error': 'Укажите system_type или id'}, status=400)

    @action(detail=True, methods=['delete'], url_path='systems/(?P<system_id>[^/.]+)')
    def delete_system(self, request, pk=None, system_id=None):
        """Отключить систему дома."""
        building = self.get_object()
        bs = get_object_or_404(BuildingSystem, id=system_id, building=building)
        bs.is_active = False
        bs.save()
        return Response({'ok': True})

    @action(detail=True, methods=['post'])
    def auto_entrances(self, request, pk=None):
        """Авто-создание подъездов: разбивает квартиры поровну."""
        building = self.get_object()
        entrances_count = request.data.get('entrances_count', building.entrances_count)
        apartments_count = request.data.get('apartments_count', building.apartments_count)

        if entrances_count and apartments_count and entrances_count > 0:
            per_entrance = apartments_count // entrances_count
            remainder = apartments_count % entrances_count
            created = 0
            start = 1
            for i in range(1, entrances_count + 1):
                size = per_entrance + (1 if i <= remainder else 0)
                end = start + size - 1
                BuildingEntrance.objects.get_or_create(
                    building=building, number=i,
                    defaults={'apartment_from': start, 'apartment_to': end, 'apartments_count': size}
                )
                start = end + 1
                created += 1
            return Response({'ok': True, 'entrances_created': created})
        return Response({'error': 'Укажите entrances_count и apartments_count'}, status=400)

    @action(detail=True, methods=['post'])
    def apply_to_residents(self, request, pk=None):
        """
        Кнопка «Применить ко всем квартирам»: переносит адрес дома
        на всех привязанных клиентов (address, region, district).
        """
        building = self.get_object()
        residents = Client.objects.filter(building=building)
        count = 0
        for client in residents:
            changed = False
            # Формируем адрес как в _build_client_address
            parts = [f'г. {building.city}' if building.city else 'Санкт-Петербург']
            if building.district and building.district != building.city:
                parts.append(building.district)
            if building.street_name:
                street = building.get_street_type_display().lower()
                parts.append(f'{street} {building.street_name}'.strip())
            house = f'д. {building.house_number}'
            if building.building_number:
                house += f' корп. {building.building_number}'
            if building.liter:
                house += f' лит. {building.liter}'
            parts.append(house)
            if client.apartment:
                parts.append(f'кв. {client.apartment}')
            new_address = ', '.join(parts)

            if client.address != new_address:
                client.address = new_address
                changed = True
            if building.region_id and client.region_id != building.region_id:
                client.region = building.region
                changed = True
            if building.district and client.district != building.district:
                client.district = building.district
                changed = True

            if changed:
                client.save(update_fields=['address', 'region', 'district'])
                count += 1

        return Response({'ok': True, 'residents_updated': count, 'total_residents': residents.count()})

    @action(detail=True, methods=['post'])
    def dadata_verify(self, request, pk=None):
        """
        Проверить/нормализовать адрес дома через Dadata.
        Принимает опционально address (строка) или использует full_address дома.
        Возвращает нормализованные поля.
        """
        building = self.get_object()
        raw = request.data.get('address', '') or str(building)

        from .dadata_service import normalize_address as dadata_normalize
        result = dadata_normalize(raw)

        if result['success']:
            # Обновляем дом
            if result['region']:
                from .models import Region
                region_name = result['region']
                region_code = '78' if 'спб' in region_name.lower() or 'санкт' in region_name.lower() or 'петербург' in region_name.lower() else ('47' if 'ленин' in region_name.lower() else '')
                region_obj = None
                if region_name:
                    region_obj, _ = Region.objects.get_or_create(
                        name=region_name,
                        defaults={'code': region_code, 'country': 'Россия'}
                    )
                    building.region = region_obj
            if result['district']:
                building.district = result['district']
            if result['city']:
                building.city = result['city']
            if result['street_name']:
                building.street_name = result['street_name']
            if result['street_type']:
                building.street_type = result['street_type']
            if result['house_number']:
                building.house_number = result['house_number']
            if result['building_number']:
                building.building_number = result['building_number']
            building.dadata_verified = True
            building.save()
            return Response({
                'success': True,
                'building': BuildingSerializer(building).data,
                'dadata': {
                    'city': result['city'], 'region': result['region'],
                    'district': result['district'], 'street_name': result['street_name'],
                    'street_type': result['street_type'], 'house_number': result['house_number'],
                    'building_number': result['building_number'],
                    'full_address': result['full_address'],
                }
            })
        return Response({'success': False, 'error': 'Не удалось нормализовать адрес через Dadata'}, status=400)


class TraccarSettingsViewSet(viewsets.ModelViewSet):
    """Настройки интеграции с Traccar GPS (только один объект)"""
    queryset = TraccarSettings.objects.all()
    serializer_class = TraccarSettingsSerializer

    def list(self, request, *args, **kwargs):
        # Всегда возвращаем один объект настроек
        settings = TraccarSettings.objects.first()
        if not settings:
            settings = TraccarSettings.objects.create(
                server_url='http://localhost:8082',
                username='',
                password='',
                is_active=False,
                sync_interval_minutes=5
            )
        return Response(self.get_serializer(settings).data)

    @action(detail=False, methods=['post'])
    def test_connection(self, request):
        """Проверка соединения с сервером Traccar"""
        import requests
        settings = TraccarSettings.objects.first()
        if not settings:
            return Response({'error': 'Настройки не найдены'}, status=400)

        try:
            resp = requests.get(
                f"{settings.server_url.rstrip('/')}/api/server",
                auth=(settings.username, settings.password),
                timeout=5
            )
            return Response({'ok': resp.status_code == 200, 'status': resp.status_code, 'version': resp.json().get('version', '?')})
        except Exception as e:
            return Response({'ok': False, 'error': str(e)})

    @action(detail=False, methods=['post'])
    def discover(self, request):
        """Поиск устройств в Traccar и авто-заполнение internal_device_id"""
        import requests
        settings = TraccarSettings.objects.first()
        if not settings or not settings.is_active:
            return Response({'error': 'Интеграция не активна'}, status=400)

        try:
            resp = requests.get(
                f"{settings.server_url.rstrip('/')}/api/devices",
                auth=(settings.username, settings.password),
                timeout=10
            )
            if resp.status_code == 200:
                devices = resp.json()
                result = [{'id': d['id'], 'name': d['name'], 'uniqueId': d['uniqueId']} for d in devices]
                return Response({'count': len(result), 'devices': result})
            return Response({'error': f'Status {resp.status_code}'}, status=400)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TraccarDeviceViewSet(viewsets.ModelViewSet):
    """GPS-устройства мастеров (привязка к Traccar)"""
    queryset = TraccarDevice.objects.all()
    serializer_class = TraccarDeviceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['master', 'is_online']
    search_fields = ['device_name', 'master__user__first_name', 'master__user__last_name']

    def create(self, request, *args, **kwargs):
        """Привязка GPS-устройства: по unique_id авто-находим internal_device_id"""
        import requests
        unique_id = str(request.data.get('unique_id', '')).strip()
        master_id = request.data.get('master_id')

        if not unique_id or not master_id:
            return Response({'error': 'unique_id и master_id обязательны'}, status=400)

        settings = TraccarSettings.objects.first()
        internal_id = None
        device_name = request.data.get('device_name', '')

        if settings and settings.is_active:
            try:
                resp = requests.get(
                    f"{settings.server_url.rstrip('/')}/api/devices",
                    auth=(settings.username, settings.password),
                    timeout=10
                )
                if resp.status_code == 200:
                    for dev in resp.json():
                        if str(dev.get('uniqueId', '')) == unique_id:
                            internal_id = dev['id']
                            device_name = device_name or dev.get('name', '')
                            break
            except Exception as e:
                return Response({'error': f'Ошибка связи с Traccar: {e}'}, status=400)

        if not internal_id:
            return Response(
                {'error': f'Устройство с uniqueId={unique_id} не найдено в Traccar. Проверьте настройки подключения.'},
                status=400
            )

        device = TraccarDevice.objects.create(
            master=Master.objects.get(id=master_id),
            internal_device_id=internal_id,
            unique_id=unique_id,
            device_name=device_name,
        )
        return Response(self.get_serializer(device).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def sync_position(self, request, pk=None):
        """Синхронизировать позицию устройства из Traccar"""
        import requests
        device = self.get_object()
        settings = TraccarSettings.objects.first()
        if not settings or not settings.is_active:
            return Response({'error': 'Интеграция не активна'}, status=400)

        try:
            resp = requests.get(
                f"{settings.server_url.rstrip('/')}/api/positions",
                auth=(settings.username, settings.password),
                params={'deviceId': device.internal_device_id},
                timeout=10
            )
            if resp.status_code == 200 and resp.json():
                pos = resp.json()[0]
                device.last_latitude = pos.get('latitude')
                device.last_longitude = pos.get('longitude')
                device.last_speed = round(pos.get('speed', 0) * 1.852, 1) if pos.get('speed') else None
                device.last_update = timezone.now()
                device.is_online = True
                device.save()
                return Response(self.get_serializer(device).data)
            return Response({'error': 'Нет данных с устройства'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class SystemSettingsViewSet(viewsets.ViewSet):
    """Системные настройки и операции"""
    permission_classes = [IsAuthenticated]

    def list(self, request):
        settings, _ = SystemSettings.objects.get_or_create()
        return Response(SystemSettingsSerializer(settings).data)

    def update(self, request, pk=None):
        settings = SystemSettings.objects.first()
        if not settings:
            settings = SystemSettings.objects.create()
        serializer = SystemSettingsSerializer(settings, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def lookup_company_by_inn(self, request):
        """Поиск компании по ИНН через DaData API"""
        import requests
        inn = request.query_params.get('inn', '').strip()
        if not inn or len(inn) < 10:
            return Response({'error': 'Укажите ИНН (10 или 12 цифр)'}, status=400)

        settings = SystemSettings.objects.first()
        token = settings.dadata_token if settings else ''
        if not token:
            return Response({'error': 'Не настроен DaData API токен в системных настройках'}, status=400)

        try:
            resp = requests.post(
                'https://suggestions.dadata.ru/suggestions/api/4_1/rs/findById/party',
                json={'query': inn},
                headers={
                    'Content-Type': 'application/json',
                    'Accept': 'application/json',
                    'Authorization': f'Token {token}'
                },
                timeout=10
            )
            data = resp.json()
            suggestions = data.get('suggestions', [])
            if not suggestions:
                return Response({'found': False, 'message': 'Компания не найдена'})

            s = suggestions[0]
            d = s.get('data', {})
            return Response({
                'found': True,
                'name': d.get('name', {}).get('full_with_opf', '') or s.get('value', ''),
                'short_name': d.get('name', {}).get('short_with_opf', ''),
                'inn': d.get('inn', inn),
                'kpp': d.get('kpp', ''),
                'ogrn': d.get('ogrn', ''),
                'legal_address': d.get('address', {}).get('unrestricted_value', '') or d.get('address', {}).get('value', ''),
                'director': d.get('management', {}).get('name', ''),
                'director_post': d.get('management', {}).get('post', ''),
                'status': d.get('state', {}).get('status', ''),
            })
        except Exception as e:
            return Response({'error': f'Ошибка запроса к DaData: {str(e)}'}, status=500)

    @action(detail=False, methods=['get'])
    def check_update(self, request):
        """Проверить наличие обновлений: сравниваем origin/branch с GitHub."""
        import requests, re, os, subprocess
        settings = SystemSettings.objects.first()
        if not settings or not settings.git_repo_url:
            return Response({'has_update': False, 'error': 'Не настроен GitHub репозиторий'})

        repo_url = settings.git_repo_url.rstrip('/').replace('.git', '')
        branch = settings.git_branch or 'main'
        headers = {'Accept': 'application/vnd.github.v3+json'}
        if settings.git_token:
            headers['Authorization'] = f'Bearer {settings.git_token}'

        m = re.search(r'github\.com[:/]([^/]+)/([^/\s]+)', repo_url)
        if not m:
            return Response({'has_update': False, 'error': 'Неверный формат URL репозитория'})
        owner, repo = m.group(1), m.group(2)

        try:
            # Получаем последний коммит из GitHub API
            api_url = f'https://api.github.com/repos/{owner}/{repo}/commits/{branch}'
            resp = requests.get(api_url, headers=headers, timeout=15)
            if resp.status_code != 200:
                return Response({'has_update': False, 'error': f'GitHub API: {resp.status_code} {resp.json().get("message", "")}'})
            remote_full = resp.json().get('sha', '')
            remote_sha = remote_full[:7]

            # Получаем локальный origin/branch — то, что уже запушено
            local_sha = ''
            origin_ref = f'origin/{branch}'
            origin_path = os.path.join('/app/.git/refs/remotes/origin', branch)
            if os.path.exists(origin_path):
                with open(origin_path) as f:
                    local_sha = f.read().strip()[:7]
            else:
                # Пробуем через git rev-parse
                try:
                    r = subprocess.run(['git', 'rev-parse', '--short', origin_ref],
                        cwd='/app', capture_output=True, text=True, timeout=5)
                    if r.returncode == 0:
                        local_sha = r.stdout.strip()
                except:
                    pass

            has_update = bool(remote_full and local_sha and remote_full[:7] != local_sha[:7])

            settings.last_update_check = timezone.now()
            settings.latest_commit = local_sha or ''
            settings.save(update_fields=['last_update_check', 'latest_commit'])

            return Response({
                'has_update': has_update,
                'local': local_sha or '?',
                'remote': remote_sha,
                'checked_at': settings.last_update_check.isoformat(),
            })
        except Exception as e:
            return Response({'has_update': False, 'error': str(e)})

    @action(detail=False, methods=['post'])
    def update_now(self, request):
        """Обновить CRM через git pull + npm build (фронтенд)."""
        import subprocess, os, json

        settings = SystemSettings.objects.first()

        try:
            git_dir = '/app/.git'
            if not os.path.isdir(git_dir):
                return Response({'ok': False, 'error': '.git не найден в /app'})

            host_dir = '/app/host'
            has_host = os.path.isdir(host_dir) and os.path.exists(os.path.join(host_dir, 'frontend', 'package.json'))

            branch = settings.git_branch or 'main'
            steps = []

            # Шаг 1: git fetch
            steps.append('git fetch')
            r = subprocess.run(['git', 'fetch', 'origin'], cwd='/app', capture_output=True, text=True, timeout=60)
            if r.returncode != 0:
                return Response({'ok': False, 'error': f'git fetch failed: {r.stderr[-200:]}'})

            # Шаг 2: git reset --hard
            steps.append('git reset')
            r = subprocess.run(['git', 'reset', '--hard', f'origin/{branch}'], cwd='/app', capture_output=True, text=True, timeout=30)
            if r.returncode != 0:
                return Response({'ok': False, 'error': f'git reset failed: {r.stderr[-200:]}'})

            # Шаг 3: миграции
            steps.append('migrate')
            r = subprocess.run(['python', 'manage.py', 'migrate'], cwd='/app', capture_output=True, text=True, timeout=60)
            migrate_out = r.stdout[-300:]

            # Шаг 4: сборка фронтенда (через хост-директорию)
            build_out = ''
            if has_host:
                steps.append('npm install')
                r = subprocess.run(['npm', 'install'], cwd=os.path.join(host_dir, 'frontend'),
                    capture_output=True, text=True, timeout=180, env={**os.environ, 'NODE_ENV': 'production'})
                steps.append('npm build')
                r = subprocess.run(['npm', 'run', 'build'], cwd=os.path.join(host_dir, 'frontend'),
                    capture_output=True, text=True, timeout=300, env={**os.environ, 'NODE_ENV': 'production'})
                build_out = r.stdout[-500:] + '\n' + r.stderr[-300:]
                # Копируем build в /app/frontend/build если нужно
                src_build = os.path.join(host_dir, 'frontend', 'build')
                dst_build = '/app/frontend/build'
                if os.path.isdir(src_build) and os.path.isdir(os.path.dirname(dst_build)):
                    subprocess.run(['cp', '-r', src_build, dst_build], timeout=30)
                steps.append('build done')
            else:
                steps.append('no frontend dir')

            # Шаг 5: перезапуск Daphne
            subprocess.run(['touch', '/app/crm/asgi.py'], timeout=5)

            # Новый HEAD
            r = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'], cwd='/app', capture_output=True, text=True, timeout=5)
            new_commit = r.stdout.strip()

            settings.latest_commit = new_commit
            settings.last_update_check = timezone.now()
            settings.save(update_fields=['latest_commit', 'last_update_check'])

            return Response({
                'ok': True,
                'commit': new_commit,
                'steps': steps,
                'migrate_output': migrate_out.strip(),
                'build_output': build_out.strip()[-500:],
                'frontend_built': has_host,
            })
        except Exception as e:
            return Response({'ok': False, 'error': str(e)})

    @action(detail=False, methods=['post'])
    def backup_db(self, request):
        """Ручной бэкап базы данных"""
        import subprocess
        import os
        settings = SystemSettings.objects.first()
        path = settings.backup_path if settings else '/var/backups/crm/'
        os.makedirs(path, exist_ok=True)
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{path}crm_backup_{timestamp}.sql'

        try:
            result = subprocess.run(
                ['pg_dump', '-h', 'localhost', '-U', 'crm_user', '-d', 'crm', '-f', filename],
                env={**os.environ, 'PGPASSWORD': 'crm_pass'},
                capture_output=True, text=True, timeout=60
            )
            if result.returncode == 0:
                return Response({'ok': True, 'file': filename, 'message': 'Бэкап создан'})
            return Response({'ok': False, 'error': result.stderr}, status=500)
        except Exception as e:
            return Response({'ok': False, 'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def export_data(self, request):
        """
        Экспорт всей базы в JSON для переноса на другой сервер.
        ?sections=clients,orders,users,buildings,tariffs,erc,equipment,inventory,settings,all
        """
        import json
        from django.core.serializers import serialize
        from django.contrib.auth.models import User as AuthUser

        sections = request.query_params.get('sections', 'all')
        sections = [s.strip() for s in sections.split(',') if s.strip()]
        if 'all' in sections:
            sections = ['clients', 'orders', 'users', 'buildings', 'tariffs', 'erc', 'equipment', 'inventory', 'settings']

        dump = {'version': '3.2.0', 'exported_at': timezone.now().isoformat(), 'sections': {}}

        models_map = {
            'clients': [Client, Building],
            'orders': [Order, OrderHistory, OrderComment, OrderMedia, Payment, MasterCashDebt],
            'users': [AuthUser, UserProfile, Master, WorkShift, PushToken],
            'buildings': [BuildingEntrance, ManagementCompany],
            'tariffs': [Tariff, PaymentRecord],
            'erc': [ErcAccount, ErcBillingRecord],
            'equipment': [Equipment, InventoryItem, InventoryMovement, Supplier,
                          SupplyInvoice, SupplyInvoiceItem, IssueOrder, IssueOrderItem,
                          LegalEntity, EstimateService, CommercialEstimate, EstimateItem,
                          StorageLocation, OutgoingInvoice, OutgoingInvoiceItem,
                          PurchaseRequest, PurchaseRequestItem],
            'inventory': [],  # уже в equipment
            'settings': [SystemSettings, TraccarSettings, TraccarDevice,
                         AsteriskSipPeer, AsteriskTrunk, AsteriskRoute,
                         AsteriskIvr, AsteriskIvrOption, AsteriskVoicemail, AsteriskCallRecording],
        }

        for section in sections:
            dump['sections'][section] = {}
            models = models_map.get(section, [])
            for model in models:
                try:
                    data = json.loads(serialize('json', model.objects.all()))
                    dump['sections'][section][model.__name__] = [obj['fields'] for obj in data]
                except Exception:
                    dump['sections'][section][model.__name__] = []

        # User — особый случай: экспортируем с паролями (хешами)
        if 'users' in sections:
            users = AuthUser.objects.all()
            dump['sections']['users']['User'] = [
                {'username': u.username, 'email': u.email, 'first_name': u.first_name,
                 'last_name': u.last_name, 'is_staff': u.is_staff, 'is_active': u.is_active,
                 'date_joined': u.date_joined.isoformat() if u.date_joined else None,
                 'password': u.password}
                for u in users
            ]

        # Считаем размер
        resp_data = json.dumps(dump, ensure_ascii=False, indent=2)
        size_kb = round(len(resp_data) / 1024, 1)

        return Response({
            'success': True,
            'size_kb': size_kb,
            'sections_exported': list(dump['sections'].keys()),
            'dump': dump,  # полный дамп — фронтенд скачает как файл
        })

    @action(detail=False, methods=['post'])
    def import_data(self, request):
        """
        Импорт базы из JSON-дампа (multipart: file).
        """
        import json
        from django.contrib.auth.models import User as AuthUser
        from django.db import transaction

        if 'file' not in request.FILES:
            return Response({'ok': False, 'error': 'Файл дампа не прикреплён (поле file)'}, status=400)

        try:
            raw = request.FILES['file'].read().decode('utf-8')
            dump = json.loads(raw)
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            return Response({'ok': False, 'error': f'Ошибка чтения JSON: {str(e)}'}, status=400)

        sections_data = dump.get('sections', {})
        dry_run = request.data.get('dry_run', 'false') == 'true'

        stats = {'created': 0, 'updated': 0, 'errors': []}

        # Порядок важен: сначала справочники, потом зависимые
        import_order = [
            ('buildings', BuildingEntrance, ['building_id']),
            ('buildings', ManagementCompany, []),
            ('tariffs', Tariff, []),
            ('clients', Building, ['region_id']),
            ('clients', Client, ['region_id', 'building_id', 'management_company_id', 'entrance_id', 'tariff_id']),
            ('users', AuthUser, []),
            ('users', UserProfile, ['user_id']),
            ('users', Master, ['user_id', 'region_id']),
            ('users', WorkShift, ['user_id']),
            ('users', PushToken, ['user_id']),
            ('tariffs', PaymentRecord, ['client_id']),
            ('erc', ErcAccount, ['client_id']),
            ('erc', ErcBillingRecord, ['account_id']),
            ('equipment', Equipment, ['client_id']),
            ('equipment', InventoryItem, ['storage_location_id']),
            ('equipment', StorageLocation, []),
            ('equipment', Supplier, []),
            ('equipment', LegalEntity, []),
            ('equipment', EstimateService, []),
            ('equipment', CommercialEstimate, ['client_id', 'legal_entity_id', 'order_id']),
            ('equipment', EstimateItem, ['estimate_id', 'inventory_item_id', 'service_id']),
            ('orders', Order, ['client_id', 'master_id', 'region_id', 'building_id', 'confirmed_by_id']),
            ('orders', OrderHistory, ['order_id', 'changed_by_id']),
            ('orders', OrderComment, ['order_id', 'author_id']),
            ('orders', OrderMedia, ['order_id', 'uploaded_by_id']),
            ('orders', Payment, ['order_id', 'received_by_id']),
            ('orders', MasterCashDebt, ['master_id', 'order_id']),
            ('equipment', InventoryMovement, ['item_id', 'master_id', 'order_id', 'performed_by_id', 'supply_invoice_id']),
            ('equipment', SupplyInvoice, ['supplier_id', 'received_by_id']),
            ('equipment', SupplyInvoiceItem, ['invoice_id', 'inventory_item_id']),
            ('equipment', IssueOrder, ['order_id', 'master_id', 'issued_by_id']),
            ('equipment', IssueOrderItem, ['issue_order_id', 'inventory_item_id']),
            ('equipment', OutgoingInvoice, ['from_legal_id', 'to_client_id']),
            ('equipment', OutgoingInvoiceItem, ['invoice_id', 'inventory_item_id']),
            ('equipment', PurchaseRequest, ['estimate_id', 'order_id', 'created_by_id']),
            ('equipment', PurchaseRequestItem, ['purchase_request_id', 'inventory_item_id', 'supplier_id']),
            ('settings', SystemSettings, []),
            ('settings', TraccarSettings, []),
            ('settings', TraccarDevice, ['master_id']),
            ('settings', AsteriskSipPeer, []),
            ('settings', AsteriskTrunk, []),
            ('settings', AsteriskRoute, ['trunk_id']),
            ('settings', AsteriskIvr, []),
            ('settings', AsteriskIvrOption, ['ivr_id']),
            ('settings', AsteriskVoicemail, []),
            ('settings', AsteriskCallRecording, ['client_id']),
        ]

        with transaction.atomic():
            for section_name, model, fk_fields in import_order:
                rows = sections_data.get(section_name, {}).get(model.__name__, [])
                if not rows:
                    continue
                for row_data in rows:
                    try:
                        # Убираем поля, которые могут отсутствовать
                        clean_data = {}
                        for k, v in row_data.items():
                            # Пропускаем auto-поля
                            if k in ('id', 'created_at', 'updated_at', 'date_joined', 'last_login'):
                                continue
                            if k.endswith('_at') or k.endswith('_date'):
                                if v:
                                    clean_data[k] = v
                                continue
                            # FK-поля: None → пропускаем
                            if k in fk_fields:
                                if v is not None:
                                    clean_data[k] = v
                                continue
                            clean_data[k] = v

                        # User — особый случай (пароль уже хеширован)
                        if model is AuthUser:
                            pwd = clean_data.pop('password', None)
                            user, created = AuthUser.objects.get_or_create(
                                username=clean_data.get('username', ''),
                                defaults={**clean_data, 'password': pwd or ''}
                            )
                            if not created and pwd:
                                user.password = pwd
                                user.save()
                            if created:
                                stats['created'] += 1
                            else:
                                stats['updated'] += 1
                            continue

                        # ErcAccount — по account_number
                        if model is ErcAccount:
                            acc_num = clean_data.pop('account_number', None)
                            if acc_num:
                                obj, created = model.objects.update_or_create(
                                    account_number=acc_num, defaults=clean_data
                                )
                            else:
                                obj = model.objects.create(**clean_data)
                                created = True
                            if created: stats['created'] += 1
                            else: stats['updated'] += 1
                            continue

                        # SystemSettings — только один объект
                        if model is SystemSettings:
                            obj, created = model.objects.get_or_create(defaults=clean_data)
                            if not created:
                                for k, v in clean_data.items():
                                    setattr(obj, k, v)
                                obj.save()
                                stats['updated'] += 1
                            else:
                                stats['created'] += 1
                            continue

                        # ManagementCompany — по name
                        if model is ManagementCompany:
                            name = clean_data.pop('name', None)
                            if name:
                                obj, created = model.objects.update_or_create(name=name, defaults=clean_data)
                            else:
                                obj = model.objects.create(**clean_data)
                                created = True
                            if created: stats['created'] += 1
                            else: stats['updated'] += 1
                            continue

                        # По умолчанию: просто создаём
                        model.objects.create(**clean_data)
                        stats['created'] += 1
                    except Exception as e:
                        stats['errors'].append(f'{model.__name__}: {str(e)[:200]}')

            if dry_run:
                raise transaction.TransactionManagementError('DRY_RUN')

        # Пересчитываем счётчики
        try:
            from django.core.management import call_command
            call_command('update_counters', '--all')
        except Exception:
            pass

        return Response({
            'ok': True,
            'dry_run': dry_run,
            **stats,
            'message': f'Создано: {stats["created"]}, обновлено: {stats["updated"]}, ошибок: {len(stats["errors"])}',
        })


class UserProfileViewSet(viewsets.ModelViewSet):
    """Управление пользователями (админы/диспетчеры/мастера)"""
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['role', 'is_on_shift']
    search_fields = ['user__username', 'user__first_name', 'user__last_name']

    @action(detail=False, methods=['get'])
    def dispatcher_stats(self, request):
        dps = UserProfile.objects.filter(role='dispatcher')
        stats = []
        for dp in dps:
            confirmed = Order.objects.filter(confirmed_by=dp.user).count()
            stats.append({'id': dp.id, 'name': str(dp.user), 'confirmed_orders': confirmed, 'is_on_shift': dp.is_on_shift})
        return Response(stats)


class WorkShiftViewSet(viewsets.ModelViewSet):
    """Рабочие смены"""
    queryset = WorkShift.objects.all()
    serializer_class = WorkShiftSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['user', 'is_active']

    @action(detail=False, methods=['post'])
    def start(self, request):
        try:
            profile = request.user.profile
        except UserProfile.DoesNotExist:
            UserProfile.objects.create(user=request.user, role='master')
            profile = request.user.profile
        if profile.is_on_shift:
            return Response({'error': 'Смена уже начата'}, status=400)
        profile.is_on_shift = True
        profile.shift_started_at = timezone.now()
        profile.save()
        shift = WorkShift.objects.create(user=request.user, started_at=timezone.now())
        return Response(WorkShiftSerializer(shift).data, status=201)

    @action(detail=False, methods=['post'])
    def end(self, request):
        try:
            profile = request.user.profile
        except UserProfile.DoesNotExist:
            return Response({'error': 'Профиль не найден'}, status=400)
        shift = WorkShift.objects.filter(user=request.user, is_active=True).first()
        if not shift:
            return Response({'error': 'Нет активной смены'}, status=400)
        now = timezone.now()
        shift.ended_at = now
        shift.is_active = False
        orders = Order.objects.filter(master__user=request.user, updated_at__gte=shift.started_at, updated_at__lte=now)
        if not orders.exists():
            orders = Order.objects.filter(confirmed_by=request.user, updated_at__gte=shift.started_at, updated_at__lte=now)
        shift.orders_total = orders.count()
        shift.orders_completed = orders.filter(status__in=['completed', 'confirmed']).count()
        shift.total_cost = orders.aggregate(s=Sum('cost'))['s'] or 0
        try:
            from .models import TraccarMileage
            shift.total_mileage_km = TraccarMileage.objects.filter(master__user=request.user, synced_at__gte=shift.started_at, synced_at__lte=now).aggregate(s=Sum('distance_km'))['s'] or 0
        except Exception:
            pass
        shift.hours_worked = round((now - shift.started_at).total_seconds() / 3600, 1)
        shift.save()
        profile.is_on_shift = False
        profile.shift_started_at = None
        profile.save()
        return Response(WorkShiftSerializer(shift).data)

    @action(detail=False, methods=['get'])
    def my_today(self, request):
        today = timezone.localdate()
        shifts = WorkShift.objects.filter(user=request.user, started_at__date=today)
        return Response({
            'shifts': WorkShiftSerializer(shifts, many=True).data,
            'total_orders': sum(s.orders_total for s in shifts),
            'total_cost': float(sum(s.total_cost for s in shifts)),
            'total_mileage': round(sum(s.total_mileage_km for s in shifts), 1),
            'total_hours': round(sum(s.hours_worked for s in shifts), 1),
        })


class OrderMediaViewSet(viewsets.ModelViewSet):
    """Фото/видео отчёты по заявке"""
    queryset = OrderMedia.objects.all()
    serializer_class = OrderMediaSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['order', 'file_type']

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

    @action(detail=False, methods=['post'], url_path='upload')
    def upload(self, request):
        """Загрузка фото/видео для заявки (multipart: file, order_id, file_type)"""
        order_id = request.data.get('order_id')
        file_type = request.data.get('file_type', 'image')
        notes = request.data.get('notes', '')

        if not order_id:
            return Response({'error': 'order_id обязателен'}, status=400)

        try:
            order = Order.objects.get(id=order_id)
        except Order.DoesNotExist:
            return Response({'error': 'Заявка не найдена'}, status=404)

        if 'file' not in request.FILES:
            return Response({'error': 'Файл не прикреплён'}, status=400)

        media = OrderMedia.objects.create(
            order=order,
            file=request.FILES['file'],
            file_type=file_type,
            uploaded_by=request.user,
            notes=notes,
        )
        return Response(OrderMediaSerializer(media).data, status=201)


class PushTokenViewSet(viewsets.ModelViewSet):
    """Push-токены для уведомлений (Expo)"""
    queryset = PushToken.objects.all()
    serializer_class = PushTokenSerializer
    http_method_names = ['post', 'delete', 'get']

    def create(self, request, *args, **kwargs):
        """Регистрация push-токена"""
        token = request.data.get('token', '').strip()
        platform = request.data.get('platform', 'android')
        if not token:
            return Response({'error': 'token обязателен'}, status=400)
        obj, created = PushToken.objects.update_or_create(
            token=token,
            defaults={'user': request.user, 'platform': platform, 'is_active': True}
        )
        return Response(PushTokenSerializer(obj).data, status=201 if created else 200)

    def get_queryset(self):
        return PushToken.objects.filter(user=self.request.user)


def send_push_notification(user_id, title, body, data=None):
    """Отправить уведомление: Max (Telegram) + Expo Push + WebSocket"""
    from .max_service import send_notification_to_user
    print(f'[Push] Sending to user_id={user_id}: title="{title}"')
    # Max (Telegram-бот)
    result_max = send_notification_to_user(user_id, title, body)
    # Expo Push — мобильное приложение мастеров
    result_expo = _send_expo_push(user_id, title, body, data)
    # WebSocket — мгновенная доставка (Channels)
    result_ws = _send_ws_notification(user_id, title, body, data)
    print(f'[Push] Result for user_id={user_id}: via_max={result_max}, via_expo={result_expo}, via_ws={result_ws}')


def _send_ws_notification(user_id, title, body, data=None):
    """Отправить уведомление через WebSocket (Channels layer)."""
    try:
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        channel_layer = get_channel_layer()
        if channel_layer is None:
            return False
        async_to_sync(channel_layer.group_send)(
            f'user_{user_id}',
            {
                'type': 'notification_message',
                'title': title,
                'body': body,
                'data': data or {},
            }
        )
        return True
    except Exception as e:
        print(f'[WS] Layer error: {e}')
        return False


def _send_expo_push(user_id, title, body, data=None):
    """Отправить push-уведомление через Expo Push API на все активные токены пользователя."""
    import requests as expo_requests
    tokens = PushToken.objects.filter(user_id=user_id, is_active=True)
    if not tokens:
        return False

    messages = []
    for pt in tokens:
        msg = {
            'to': pt.token,
            'title': title,
            'body': body,
            'sound': 'default',
            'priority': 'high',
        }
        if data:
            msg['data'] = data
        messages.append(msg)

    if not messages:
        return False

    try:
        # Expo Push API может принимать до 100 сообщений за раз
        # Отправляем чанками по 50 для надёжности
        chunk_size = 50
        all_ok = True
        for i in range(0, len(messages), chunk_size):
            chunk = messages[i:i + chunk_size]
            resp = expo_requests.post(
                'https://exp.host/--/api/v2/push/send',
                json=chunk,
                headers={
                    'Accept': 'application/json',
                    'Accept-Encoding': 'gzip, deflate',
                    'Content-Type': 'application/json',
                },
                timeout=15,
            )
            if resp.status_code != 200:
                print(f'[ExpoPush] HTTP {resp.status_code}: {resp.text[:200]}')
                all_ok = False
                continue

            result = resp.json()
            if 'data' in result:
                for item in result['data']:
                    if item.get('status') == 'error':
                        err = item.get('message', '')
                        print(f'[ExpoPush] Token error: {err}')
                        # Деактивируем невалидные токены
                        if 'DeviceNotRegistered' in err:
                            expo_token = item.get('id') or messages[0].get('to', '')
                            PushToken.objects.filter(token=expo_token).update(is_active=False)
                            print(f'[ExpoPush] Deactivated token: {expo_token[:20]}...')
                        all_ok = False
        return all_ok
    except Exception as e:
        print(f'[ExpoPush] Error: {e}')
        return False


# ==================== Auth endpoints ====================

@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """Аутентификация пользователя, возвращает токен и данные пользователя"""
    serializer = LoginSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    username = serializer.validated_data['username']
    password = serializer.validated_data['password']

    user = authenticate(username=username, password=password)
    if not user:
        return Response(
            {'error': 'Неверные учетные данные'},
            status=status.HTTP_401_UNAUTHORIZED
        )

    token, _ = Token.objects.get_or_create(user=user)

    return Response({
        'token': token.key,
        'user': UserSerializer(user).data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    """Возвращает данные текущего пользователя"""
    return Response(UserSerializer(request.user).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def refresh_token_view(request):
    """Обновление токена (ротация)"""
    request.user.auth_token.delete()
    token = Token.objects.create(user=request.user)
    return Response({'token': token.key})


# ══════════════════════════════════════════════════════════════════
# Склад и оборудование
# ══════════════════════════════════════════════════════════════════

class InventoryItemViewSet(viewsets.ModelViewSet):
    """Склад: оборудование (приход, учёт, остатки)"""
    queryset = InventoryItem.objects.all()
    serializer_class = InventoryItemSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['item_type', 'status', 'storage_location']
    search_fields = ['name', 'serial_number', 'model_name', 'supplier', 'barcode']

    @action(detail=False, methods=['get'])
    def by_barcode(self, request):
        """Поиск товара по штрих-коду (для сканера)"""
        barcode = request.query_params.get('code', '').strip()
        if not barcode:
            return Response({'error': 'Передайте ?code=ШТРИХКОД'}, status=400)
        try:
            item = InventoryItem.objects.get(barcode=barcode)
            return Response(InventoryItemSerializer(item).data)
        except InventoryItem.DoesNotExist:
            return Response({'error': 'Товар с таким штрих-кодом не найден', 'barcode': barcode}, status=404)

    @action(detail=True, methods=['post'])
    def generate_barcode(self, request, pk=None):
        """Сгенерировать новый штрих-код для позиции"""
        item = self.get_object()
        import uuid
        item.barcode = f'SKU-{uuid.uuid4().hex[:8].upper()}'
        item.save(update_fields=['barcode', 'updated_at'])
        return Response(InventoryItemSerializer(item).data)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Сводка по складу"""
        items = InventoryItem.objects.all()
        total_items = items.count()
        in_stock = items.filter(status='in_stock').count()
        with_masters = items.filter(status='with_master').count()
        total_value = sum(
            (i.sale_price or 0) * i.quantity
            for i in items.filter(status__in=['in_stock', 'with_master'])
        )
        by_type = {}
        for t in InventoryItem.ITEM_TYPES:
            cnt = items.filter(item_type=t[0]).count()
            if cnt:
                by_type[str(t[1])] = cnt
        return Response({
            'total_items': total_items,
            'in_stock': in_stock,
            'with_masters': with_masters,
            'total_value': float(total_value),
            'by_type': by_type,
        })

    @action(detail=True, methods=['post'])
    def add_stock(self, request, pk=None):
        """Приход на склад (простой, без накладной)"""
        item = self.get_object()
        qty = int(request.data.get('quantity', 0))
        if qty <= 0:
            return Response({'error': 'Укажите количество > 0'}, status=400)
        InventoryMovement.objects.create(
            item=item, movement_type='in', quantity=qty,
            performed_by=request.user,
            notes=request.data.get('notes', 'Приход на склад')
        )
        return Response(InventoryItemSerializer(item).data)

    @action(detail=True, methods=['post'])
    def issue_to_master(self, request, pk=None):
        """Выдать мастеру"""
        item = self.get_object()
        master_id = request.data.get('master_id')
        qty = int(request.data.get('quantity', 1))
        if not master_id:
            return Response({'error': 'Укажите master_id'}, status=400)
        try:
            master = Master.objects.get(id=master_id)
        except Master.DoesNotExist:
            return Response({'error': 'Мастер не найден'}, status=404)
        if item.quantity < qty:
            return Response({'error': f'Недостаточно на складе (доступно: {item.quantity})'}, status=400)
        InventoryMovement.objects.create(
            item=item, movement_type='out_to_master', quantity=qty,
            master=master, performed_by=request.user,
            notes=request.data.get('notes', f'Выдано мастеру {master}')
        )
        return Response(InventoryItemSerializer(item).data)


class InventoryMovementViewSet(viewsets.ReadOnlyModelViewSet):
    """Движения оборудования (только чтение)"""
    queryset = InventoryMovement.objects.select_related('item', 'master', 'master__user', 'order', 'performed_by', 'supply_invoice', 'supply_invoice__supplier')
    serializer_class = InventoryMovementSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['movement_type', 'master', 'item']

    def get_queryset(self):
        qs = super().get_queryset()
        master_id = self.request.query_params.get('master_id')
        if master_id:
            qs = qs.filter(master_id=master_id)
        return qs


# ══════════════════════════════════════════════════════════════════
# Поставщики и накладные
# ══════════════════════════════════════════════════════════════════

class SupplierViewSet(viewsets.ModelViewSet):
    """Поставщики оборудования"""
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'inn', 'phone', 'email']


class SupplyInvoiceViewSet(viewsets.ModelViewSet):
    """Накладные поставщиков"""
    queryset = SupplyInvoice.objects.select_related('supplier', 'received_by').prefetch_related('items__inventory_item')
    serializer_class = SupplyInvoiceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status', 'supplier']
    search_fields = ['invoice_number', 'supplier__name']

    def get_serializer_class(self):
        if self.action == 'create':
            return SupplyInvoiceCreateSerializer
        if self.action in ('receive', 'partial_update'):
            return SupplyInvoiceReceiveSerializer
        return SupplyInvoiceSerializer

    def create(self, request, *args, **kwargs):
        """Создать накладную с позициями одним запросом"""
        serializer = SupplyInvoiceCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        supplier = get_object_or_404(Supplier, id=data['supplier_id'])

        invoice = SupplyInvoice.objects.create(
            supplier=supplier,
            invoice_number=data['invoice_number'],
            invoice_date=data['invoice_date'],
            notes=data.get('notes', ''),
            status='draft'
        )

        items_data = data.get('items', [])
        for item_data in items_data:
            inv_item = get_object_or_404(InventoryItem, id=item_data['inventory_item_id'])
            SupplyInvoiceItem.objects.create(
                invoice=invoice,
                inventory_item=inv_item,
                quantity_ordered=item_data.get('quantity_ordered', 0),
                quantity_received=item_data.get('quantity_received', 0),
                unit_price=item_data.get('unit_price', inv_item.cost_price or 0),
                notes=item_data.get('notes', ''),
            )
        invoice.recalculate_totals()

        # Если все позиции уже приняты — сразу применяем
        if invoice.items.filter(quantity_received__gt=0).exists():
            invoice.apply_received(request.user)
            invoice.refresh_from_db()

        return Response(SupplyInvoiceSerializer(invoice).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Принять товар по накладной: обновить quantity_received и оприходовать"""
        invoice = self.get_object()
        if invoice.status == 'cancelled':
            return Response({'error': 'Накладная отменена'}, status=400)

        serializer = SupplyInvoiceReceiveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        items_data = {it['inventory_item_id']: it for it in serializer.validated_data['items']}

        for item in invoice.items.all():
            if item.inventory_item_id in items_data:
                new_qty = items_data[item.inventory_item_id].get('quantity_received', item.quantity_received)
                item.quantity_received = min(new_qty, item.quantity_ordered)  # не больше заказанного
                item.save(update_fields=['quantity_received'])

        invoice.apply_received(request.user)
        invoice.recalculate_totals()
        invoice.refresh_from_db()

        return Response(SupplyInvoiceSerializer(invoice).data)

    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        """Добавить позицию в существующую накладную"""
        invoice = self.get_object()
        if invoice.status in ('received', 'cancelled'):
            return Response({'error': 'Нельзя редактировать принятую/отменённую накладную'}, status=400)

        inv_item = get_object_or_404(InventoryItem, id=request.data.get('inventory_item_id'))
        item, created = SupplyInvoiceItem.objects.update_or_create(
            invoice=invoice,
            inventory_item=inv_item,
            defaults={
                'quantity_ordered': request.data.get('quantity_ordered', 1),
                'quantity_received': request.data.get('quantity_received', 0),
                'unit_price': request.data.get('unit_price', inv_item.cost_price or 0),
                'notes': request.data.get('notes', ''),
            }
        )
        invoice.recalculate_totals()
        return Response(SupplyInvoiceSerializer(invoice).data)

    @action(detail=True, methods=['post'])
    def remove_item(self, request, pk=None):
        """Удалить позицию из накладной"""
        invoice = self.get_object()
        if invoice.status in ('received', 'cancelled'):
            return Response({'error': 'Нельзя редактировать принятую/отменённую накладную'}, status=400)

        item_id = request.data.get('item_id')
        SupplyInvoiceItem.objects.filter(id=item_id, invoice=invoice).delete()
        invoice.recalculate_totals()
        return Response(SupplyInvoiceSerializer(invoice).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Отменить накладную"""
        invoice = self.get_object()
        if invoice.status == 'received':
            return Response({'error': 'Нельзя отменить уже принятую накладную'}, status=400)
        invoice.status = 'cancelled'
        invoice.save()
        return Response(SupplyInvoiceSerializer(invoice).data)


# ══════════════════════════════════════════════════════════════════
# Склад v3: Расходные ордера и заявки на закупку
# ══════════════════════════════════════════════════════════════════

class IssueOrderViewSet(viewsets.ModelViewSet):
    """Расходные ордера — выдача материалов со склада под заявку"""
    queryset = IssueOrder.objects.select_related('order', 'master__user', 'issued_by').prefetch_related('items__inventory_item')
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['order', 'master', 'status']

    def get_serializer_class(self):
        if self.action == 'create':
            return IssueOrderCreateSerializer
        return IssueOrderSerializer

    def create(self, request):
        serializer = IssueOrderCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        order = get_object_or_404(Order, id=data['order_id'])
        master = get_object_or_404(Master, id=data['master_id'])
        issue_order = IssueOrder.objects.create(order=order, master=master, issued_by=request.user, notes=data.get('notes', ''))
        for item_data in data.get('items', []):
            inv_item = get_object_or_404(InventoryItem, id=item_data['inventory_item_id'])
            qty = item_data.get('quantity_issued', 1)
            IssueOrderItem.objects.create(
                issue_order=issue_order, inventory_item=inv_item,
                quantity_issued=qty,
                need_return_old=item_data.get('need_return_old', False),
                old_item_description=item_data.get('old_item_description', ''),
                notes=item_data.get('notes', ''),
            )
            InventoryMovement.objects.create(item=inv_item, movement_type='out_to_master', quantity=qty, master=master, order=order, performed_by=request.user, notes=f'Ордер №{issue_order.id}')
        return Response(IssueOrderSerializer(issue_order).data, status=201)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Сотрудник подтверждает получение материалов"""
        issue_order = self.get_object()
        if issue_order.status != 'pending':
            return Response({'error': 'Ордер уже получен или закрыт'}, status=400)
        issue_order.status = 'received'
        issue_order.received_at = timezone.now()
        issue_order.save()
        return Response(IssueOrderSerializer(issue_order).data)

    @action(detail=True, methods=['post'])
    def report_usage(self, request, pk=None):
        """Сотрудник отчитывается: сколько использовал, сколько вернул"""
        issue_order = self.get_object()
        items_data = {it['item_id']: it for it in request.data.get('items', [])}
        for item in issue_order.items.all():
            if item.inventory_item_id in items_data:
                d = items_data[item.inventory_item_id]
                item.quantity_used = d.get('quantity_used', item.quantity_used)
                item.quantity_returned = d.get('quantity_returned', item.quantity_returned)
                item.old_item_returned = d.get('old_item_returned', item.old_item_returned)
                item.save()
                if item.quantity_returned > 0:
                    inv_item = item.inventory_item
                    inv_item.quantity += item.quantity_returned
                    inv_item.status = 'in_stock'
                    inv_item.save()
                    InventoryMovement.objects.create(item=inv_item, movement_type='return_from_master', quantity=item.quantity_returned, master=issue_order.master, order=issue_order.order, performed_by=request.user, notes=f'Возврат по ордеру №{issue_order.id}')
        from django.db.models import F as _F
        all_used = not issue_order.items.filter(quantity_issued__gt=_F('quantity_used') + _F('quantity_returned')).exists()
        issue_order.status = 'fully_used' if all_used else 'partially_used'
        issue_order.completed_at = timezone.now()
        issue_order.save()
        return Response(IssueOrderSerializer(issue_order).data)


class PurchaseRequestViewSet(viewsets.ModelViewSet):
    """Заявки на закупку"""
    queryset = PurchaseRequest.objects.select_related('estimate', 'order', 'created_by').prefetch_related('items')
    serializer_class = PurchaseRequestSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status']

    def create(self, request):
        data = request.data
        pr = PurchaseRequest.objects.create(
            estimate_id=data.get('estimate_id'),
            order_id=data.get('order_id'),
            created_by=request.user,
            notes=data.get('notes', ''),
            status='pending',
        )
        for item_data in data.get('items', []):
            PurchaseRequestItem.objects.create(
                purchase_request=pr,
                inventory_item_id=item_data.get('inventory_item_id'),
                name=item_data.get('name', 'Без названия'),
                quantity=item_data.get('quantity', 1),
                unit=item_data.get('unit', 'шт.'),
                estimated_price=item_data.get('estimated_price'),
                supplier_id=item_data.get('supplier_id'),
                notes=item_data.get('notes', ''),
            )
        return Response(PurchaseRequestSerializer(pr).data, status=201)

    @action(detail=True, methods=['post'])
    def mark_ordered(self, request, pk=None):
        pr = self.get_object()
        pr.status = 'ordered'
        pr.save()
        return Response(PurchaseRequestSerializer(pr).data)

    @action(detail=True, methods=['post'])
    def mark_received(self, request, pk=None):
        pr = self.get_object()
        pr.status = 'received'
        pr.save()
        return Response(PurchaseRequestSerializer(pr).data)


# ══════════════════════════════════════════════════════════════════
# Финансы
# ══════════════════════════════════════════════════════════════════

class PaymentViewSet(viewsets.ModelViewSet):
    """Оплаты по заявкам"""
    queryset = Payment.objects.select_related('order', 'order__client', 'received_by')
    serializer_class = PaymentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['payment_method', 'is_received', 'order']

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Финансовая сводка за период"""
        days = int(request.query_params.get('days', 30))
        since = timezone.now() - timedelta(days=days)
        payments = Payment.objects.filter(paid_at__gte=since, is_received=True)
        total = payments.aggregate(s=Sum('amount'))['s'] or 0
        by_method = {}
        for m in Payment.PAYMENT_METHODS:
            s = payments.filter(payment_method=m[0]).aggregate(s=Sum('amount'))['s'] or 0
            if s:
                by_method[str(m[1])] = float(s)
        return Response({
            'total': float(total),
            'count': payments.count(),
            'by_method': by_method,
            'days': days,
        })


class MasterSalaryViewSet(viewsets.ModelViewSet):
    """Зарплаты мастеров"""
    queryset = MasterSalary.objects.select_related('master', 'master__user')
    serializer_class = MasterSalarySerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['master', 'status']

    @action(detail=False, methods=['post'])
    def calculate(self, request):
        """Рассчитать зарплату мастера за период"""
        master_id = request.data.get('master_id')
        period_start = request.data.get('period_start')
        period_end = request.data.get('period_end')
        commission = float(request.data.get('commission_percent', 30))

        if not master_id or not period_start or not period_end:
            return Response({'error': 'Укажите master_id, period_start, period_end'}, status=400)

        try:
            master = Master.objects.get(id=master_id)
        except Master.DoesNotExist:
            return Response({'error': 'Мастер не найден'}, status=404)

        orders = Order.objects.filter(
            master=master,
            confirmed_at__date__gte=period_start,
            confirmed_at__date__lte=period_end,
            status='confirmed',
        )

        total_orders = orders.count()
        total_revenue = orders.aggregate(s=Sum('cost'))['s'] or 0
        base_salary = float(total_revenue) * commission / 100
        bonus = float(request.data.get('bonus', 0))
        deduction = float(request.data.get('deduction', 0))
        total_salary = base_salary + bonus - deduction

        salary, _ = MasterSalary.objects.update_or_create(
            master=master, period_start=period_start, period_end=period_end,
            defaults={
                'orders_total': total_orders,
                'orders_completed': orders.count(),
                'total_revenue': total_revenue,
                'commission_percent': commission,
                'bonus': bonus,
                'deduction': deduction,
                'total_salary': max(0, total_salary),
                'notes': request.data.get('notes', ''),
            }
        )
        return Response(MasterSalarySerializer(salary).data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        salary = self.get_object()
        salary.status = 'approved'
        salary.save(update_fields=['status'])
        return Response(MasterSalarySerializer(salary).data)

    @action(detail=False, methods=['get'])
    def master_debts(self, request):
        """Долги мастеров по наличным и оборудованию"""
        master_id = request.query_params.get('master_id')
        cash_qs = MasterCashDebt.objects.select_related('master', 'master__user', 'order')
        inv_qs = MasterInventoryDebt.objects.select_related('master', 'master__user', 'order', 'item')
        if master_id:
            cash_qs = cash_qs.filter(master_id=master_id, is_paid_to_office=False)
            inv_qs = inv_qs.filter(master_id=master_id, is_returned=False)
        return Response({
            'cash_debts': [{'id': d.id, 'master': str(d.master), 'order': d.order.number, 'amount': float(d.amount), 'is_paid': d.is_paid_to_office} for d in cash_qs],
            'inventory_debts': [{'id': d.id, 'master': str(d.master), 'order': d.order.number, 'description': d.description, 'is_returned': d.is_returned, 'condition': d.condition} for d in inv_qs],
        })



class MessageViewSet(viewsets.ModelViewSet):
    """Чат между сотрудниками"""
    queryset = Message.objects.all()
    serializer_class = MessageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['sender', 'recipient']
    ordering_fields = ['created_at']

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Message.objects.none()
        # Видим свои сообщения, адресованные нам, и broadcast
        return Message.objects.filter(
            Q(sender=user) | Q(recipient=user) | Q(is_broadcast=True)
        ).distinct()

    def perform_create(self, serializer):
        serializer.save(sender=self.request.user)

    @action(detail=True, methods=['post'])
    def read(self, request, pk=None):
        msg = self.get_object()
        msg.read_by.add(request.user)
        return Response({'status': 'ok'})

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        user = request.user
        count = Message.objects.filter(
            Q(recipient=user) | Q(is_broadcast=True)
        ).exclude(read_by=user).count()
        return Response({'unread': count})

    @action(detail=False, methods=['get'])
    def users(self, request):
        """Список сотрудников для выбора получателя"""
        from .models import User
        users = User.objects.filter(is_active=True).values('id', 'username', 'first_name', 'last_name')
        return Response([{
            'id': u['id'],
            'username': u['username'],
            'full_name': f"{u['first_name']} {u['last_name']}".strip() or u['username']
        } for u in users])


# ══════════════════════════════════════════════════════════════════
# Сметы и КП — viewsets
# ══════════════════════════════════════════════════════════════════

class LegalEntityViewSet(viewsets.ModelViewSet):
    """Юрлица компании"""
    queryset = LegalEntity.objects.all()
    serializer_class = LegalEntitySerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'short_name', 'inn']


class EstimateServiceViewSet(viewsets.ModelViewSet):
    """Справочник услуг и работ для смет"""
    queryset = EstimateService.objects.all()
    serializer_class = EstimateServiceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'is_active']
    search_fields = ['name']
    ordering_fields = ['name', 'category', 'sale_price', 'cost_price']


class CommercialEstimateViewSet(viewsets.ModelViewSet):
    """Сметы и коммерческие предложения"""
    queryset = CommercialEstimate.objects.all()
    serializer_class = CommercialEstimateSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'client', 'legal_entity', 'order']
    search_fields = ['number', 'name', 'client__name']
    ordering_fields = ['created_at', 'total', 'status', 'number']

    def perform_create(self, serializer):
        obj = serializer.save(created_by=self.request.user)
        obj.recalculate()

    def perform_update(self, serializer):
        obj = serializer.save()
        obj.recalculate()

    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        """Добавить позицию в смету"""
        estimate = self.get_object()
        data = request.data.copy()
        data['estimate'] = estimate.id

        # Если материал со склада
        if data.get('item_type') == 'material' and data.get('inventory_item_id'):
            try:
                inv = InventoryItem.objects.get(id=data['inventory_item_id'])
                data['name'] = inv.name
                data['unit'] = inv.unit
                data['cost_price'] = inv.cost_price or 0
                data['sale_price'] = inv.sale_price or 0
            except InventoryItem.DoesNotExist:
                pass

        # Если услуга из справочника
        if data.get('item_type') == 'service' and data.get('service_id'):
            try:
                svc = EstimateService.objects.get(id=data['service_id'])
                data['name'] = svc.name
                data['unit'] = svc.unit
                data['cost_price'] = svc.cost_price
                data['sale_price'] = svc.sale_price
                data['installer_salary'] = svc.installer_salary
            except EstimateService.DoesNotExist:
                pass

        serializer = EstimateItemSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        estimate.recalculate()
        return Response(CommercialEstimateSerializer(estimate).data, status=201)

    @action(detail=True, methods=['post'])
    def remove_item(self, request, pk=None):
        """Удалить позицию из сметы"""
        estimate = self.get_object()
        item_id = request.data.get('item_id')
        if not item_id:
            return Response({'error': 'item_id обязателен'}, status=400)
        try:
            item = estimate.items.get(id=item_id)
            item.delete()
            estimate.recalculate()
            return Response(CommercialEstimateSerializer(estimate).data)
        except EstimateItem.DoesNotExist:
            return Response({'error': 'Позиция не найдена'}, status=404)

    @action(detail=True, methods=['post'])
    def update_item(self, request, pk=None):
        """Обновить позицию сметы"""
        estimate = self.get_object()
        item_id = request.data.pop('item_id', None)
        if not item_id:
            return Response({'error': 'item_id обязателен'}, status=400)
        try:
            item = estimate.items.get(id=item_id)
            for key, value in request.data.items():
                setattr(item, key, value)
            item.save()
            estimate.recalculate()
            return Response(CommercialEstimateSerializer(estimate).data)
        except EstimateItem.DoesNotExist:
            return Response({'error': 'Позиция не найдена'}, status=404)

    @action(detail=True, methods=['post'])
    def recalc(self, request, pk=None):
        """Принудительный пересчёт сметы"""
        estimate = self.get_object()
        estimate.recalculate()
        return Response(CommercialEstimateSerializer(estimate).data)

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """Генерация PDF/печатной формы коммерческого предложения"""
        estimate = self.get_object()
        settings_qs = SystemSettings.objects.first()
        settings_data = {}
        if settings_qs:
            from .serializers import SystemSettingsPublicSerializer
            settings_data = SystemSettingsPublicSerializer(settings_qs).data

        le = estimate.legal_entity
        le_data = None
        if le:
            le_data = {
                'name': le.name, 'short_name': le.short_name or le.name,
                'inn': le.inn, 'kpp': le.kpp, 'ogrn': le.ogrn,
                'legal_address': le.legal_address,
                'phone': le.phone, 'email': le.email,
                'bank_name': le.bank_name, 'bik': le.bik,
                'corr_account': le.corr_account,
                'settlement_account': le.settlement_account,
                'director': le.director,
            }

        client = estimate.client
        client_data = None
        if client:
            client_data = {
                'name': client.name, 'phone': client.phone, 'email': client.email,
                'inn': getattr(client, 'inn', ''),
                'legal_address': getattr(client, 'legal_address', ''),
                'director_name': getattr(client, 'director_name', ''),
            }

        items = estimate.items.all()
        from datetime import date, timedelta
        validity_date = (date.today() + timedelta(days=settings_data.get('cp_validity_days', 7))).strftime('%d.%m.%Y')

        color = settings_data.get('cp_color', '#1a3e60')
        logo_tag = ''
        if settings_data.get('cp_show_logo') and settings_data.get('cp_logo_url'):
            logo_tag = f'<img src="{settings_data["cp_logo_url"]}" style="max-height:60px;" />'

        items_html = ''
        for idx, item in enumerate(items, 1):
            items_html += f'''
            <tr>
                <td style="text-align:center;">{idx}</td>
                <td>{item.name}</td>
                <td style="text-align:center;">{item.unit}</td>
                <td style="text-align:center;">{item.quantity:.1f}</td>
                <td style="text-align:right;">{item.sale_price:,.2f} ₽</td>
                <td style="text-align:right;">{item.total_price:,.2f} ₽</td>
            </tr>'''

        html = f'''<!DOCTYPE html>
<html lang="ru">
<head><meta charset="utf-8"><title>КП №{estimate.number}</title>
<style>
@page {{ size: A4; margin: 15mm; }}
@page :first {{ margin-top: 25mm; }}
@media print {{
  body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
  .no-print {{ display: none !important; }}
}}
body {{ font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 11pt; color: #333; }}
.header {{ background: {color}; color: #fff; padding: 15px 20px; border-radius: 6px; margin-bottom: 20px; }}
.header h1 {{ margin: 0; font-size: 18pt; }}
.header .sub {{ font-size: 10pt; opacity: .85; margin-top: 5px; }}
.company-info {{ margin-bottom: 20px; }}
.company-info table {{ width: 100%; border-collapse: collapse; }}
.company-info td {{ vertical-align: top; padding: 8px; font-size: 10pt; }}
.company-info .label {{ color: #888; font-size: 9pt; }}
.client-info {{ margin-bottom: 20px; border: 1px solid #ddd; padding: 12px; border-radius: 6px; }}
.client-info h3 {{ margin: 0 0 10px; color: {color}; }}
.items-table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
.items-table th {{ background: {color}; color: #fff; padding: 8px; font-size: 10pt; }}
.items-table td {{ padding: 6px 8px; border-bottom: 1px solid #eee; font-size: 10pt; }}
.items-table tr:nth-child(even) {{ background: #f9f9f9; }}
.totals {{ float: right; width: 300px; margin-top: 15px; }}
.totals table {{ width: 100%; border-collapse: collapse; }}
.totals td {{ padding: 4px 8px; font-size: 10pt; }}
.totals .total {{ font-weight: bold; font-size: 14pt; color: {color}; border-top: 2px solid {color}; }}
.footer {{ margin-top: 40px; font-size: 10pt; color: #666; }}
.validity {{ margin-top: 15px; font-size: 10pt; color: #888; }}
.print-btn {{ position: fixed; top: 10px; right: 10px; padding: 10px 20px; background: {color}; color: #fff; border: none; border-radius: 6px; cursor: pointer; font-size: 14px; z-index: 999; }}
.print-btn:hover {{ opacity: .9; }}
</style></head>
<body>
<button class="print-btn no-print" onclick="window.print()">🖨️ Печать / Сохранить PDF</button>

<div class="header">
    {logo_tag}
    <h1>{settings_data.get('cp_header_text', 'Коммерческое предложение')} №{estimate.number}</h1>
    <div class="sub">от {estimate.created_at.strftime('%d.%m.%Y')}</div>
</div>

<div class="company-info">
    <table>
        <tr>
            <td width="50%">
                <div class="label">Исполнитель:</div>
                <strong>{le_data['short_name'] or le_data['name'] if le_data else 'Наша компания'}</strong><br>
                {'ИНН ' + le_data['inn'] + '<br>' if le_data and le_data.get('inn') else ''}
                {'КПП ' + le_data['kpp'] + '<br>' if le_data and le_data.get('kpp') else ''}
                {le_data['legal_address'] + '<br>' if le_data and le_data.get('legal_address') else ''}
                {'📞 ' + le_data['phone'] + '<br>' if le_data and le_data.get('phone') else ''}
            </td>
            <td width="50%">
                <div class="label">Заказчик:</div>
                <strong>{client_data['name'] if client_data else '—'}</strong><br>
                {'ИНН ' + client_data['inn'] + '<br>' if client_data and client_data.get('inn') else ''}
                {'📞 ' + client_data['phone'] + '<br>' if client_data and client_data.get('phone') else ''}
            </td>
        </tr>
    </table>
</div>

<table class="items-table">
    <tr><th>№</th><th>Наименование</th><th>Ед.</th><th>Кол-во</th><th>Цена</th><th>Сумма</th></tr>
    {items_html}
</table>

<div class="totals">
    <table>
        <tr><td>Материалы:</td><td style="text-align:right;">{estimate.total_materials:,.2f} ₽</td></tr>
        <tr><td>Услуги:</td><td style="text-align:right;">{estimate.total_services:,.2f} ₽</td></tr>
        <tr><td>Подытог:</td><td style="text-align:right;">{estimate.subtotal:,.2f} ₽</td></tr>
        {f'<tr><td>Скидка:</td><td style="text-align:right;">-{estimate.discount}%</td></tr>' if estimate.discount and estimate.discount > 0 else ''}
        {f'<tr><td>Доставка:</td><td style="text-align:right;">{estimate.delivery_cost:,.2f} ₽</td></tr>' if estimate.delivery_cost and estimate.delivery_cost > 0 else ''}
        <tr class="total"><td>ИТОГО:</td><td style="text-align:right;">{estimate.total:,.2f} ₽</td></tr>
    </table>
</div>

<div style="clear:both;"></div>

<div class="validity">
    💡 Предложение действительно до: <strong>{validity_date}</strong>
</div>

<div class="footer">
    <p>{settings_data.get('cp_footer_text', 'С уважением, команда Видео Сервис')}</p>
    {f'<p>_________________ {settings_data.get("cp_signature_name", "")}<br><em>{settings_data.get("cp_signature_title", "")}</em></p>' if settings_data.get('cp_signature_name') else ''}
</div>
</body></html>'''

        from django.http import HttpResponse
        return HttpResponse(html, content_type='text/html; charset=utf-8')


class EstimateItemViewSet(viewsets.ModelViewSet):
    """Позиции сметы (для прямого редактирования)"""
    queryset = EstimateItem.objects.all()
    serializer_class = EstimateItemSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['estimate', 'item_type']


# ══════════════════════════════════════════════════════════════════
# Места хранения (StorageLocation)
# ══════════════════════════════════════════════════════════════════

class StorageLocationViewSet(viewsets.ModelViewSet):
    """Физические места хранения товаров на складе (ячейки, стеллажи, полки)"""
    queryset = StorageLocation.objects.all()
    serializer_class = StorageLocationSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['zone', 'is_active']
    search_fields = ['code', 'barcode', 'zone', 'rack', 'shelf', 'notes']

    def get_serializer_class(self):
        if self.action == 'retrieve' or self.action == 'by_barcode':
            return StorageLocationDetailSerializer
        return StorageLocationSerializer

    @action(detail=False, methods=['get'])
    def by_barcode(self, request):
        """Поиск места хранения по штрихкоду (для сканера)"""
        barcode = request.query_params.get('code', '').strip()
        if not barcode:
            return Response({'error': 'Передайте ?code=ШТРИХКОД'}, status=400)
        try:
            loc = StorageLocation.objects.get(barcode=barcode)
            return Response(StorageLocationDetailSerializer(loc).data)
        except StorageLocation.DoesNotExist:
            return Response({'error': 'Место с таким штрихкодом не найдено', 'barcode': barcode}, status=404)

    @action(detail=True, methods=['post'])
    def move_items(self, request, pk=None):
        """Переместить товары из этого места в другое (тело: {target_location_id: N, item_ids: [...]})"""
        source = self.get_object()
        target_id = request.data.get('target_location_id')
        item_ids = request.data.get('item_ids', [])

        if not target_id:
            return Response({'error': 'Укажите target_location_id'}, status=400)
        try:
            target = StorageLocation.objects.get(id=target_id)
        except StorageLocation.DoesNotExist:
            return Response({'error': 'Целевое место не найдено'}, status=404)

        if target.is_full:
            return Response({'error': 'Целевое место заполнено'}, status=400)

        items = InventoryItem.objects.filter(id__in=item_ids, storage_location=source)
        moved_count = items.update(storage_location=target)

        return Response({
            'success': True,
            'moved_count': moved_count,
            'source': StorageLocationSerializer(source).data,
            'target': StorageLocationSerializer(target).data,
        })

    @action(detail=True, methods=['post'])
    def recount(self, request, pk=None):
        """Пересчёт: подтвердить список товаров в ячейке (тело: {item_ids: [...]})"""
        loc = self.get_object()
        confirmed_ids = request.data.get('item_ids', [])
        confirmed_set = set(confirmed_ids)

        # Товары, которые есть в ячейке, но не в списке — помечаем как missing
        current_items = InventoryItem.objects.filter(storage_location=loc)
        missing = []
        for item in current_items:
            if item.id not in confirmed_set:
                missing.append(item.id)

        return Response({
            'success': True,
            'location': StorageLocationDetailSerializer(loc).data,
            'confirmed_count': len(confirmed_set),
            'missing_item_ids': missing,
            'missing_count': len(missing),
        })


# ══════════════════════════════════════════════════════════════════
# Исходящие накладные (УПД)
# ══════════════════════════════════════════════════════════════════

class OutgoingInvoiceViewSet(viewsets.ModelViewSet):
    """Исходящие накладные (УПД) — выдача товаров со склада"""
    queryset = OutgoingInvoice.objects.prefetch_related('items__inventory_item', 'from_legal', 'to_client')
    serializer_class = OutgoingInvoiceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status', 'from_legal', 'to_client']
    search_fields = ['number', 'to_client__name', 'from_legal__name', 'basis']

    def get_serializer_class(self):
        if self.action == 'create':
            return OutgoingInvoiceCreateSerializer
        return OutgoingInvoiceSerializer

    def create(self, request, *args, **kwargs):
        """Создать УПД с позициями"""
        serializer = OutgoingInvoiceCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({'error': f'Ошибка валидации: {serializer.errors}'}, status=400)
        data = serializer.validated_data

        from_legal_id = data.get('from_legal_id')
        to_client_id = data.get('to_client_id')
        items_data = data.get('items', [])

        if not from_legal_id:
            return Response({'error': 'Не указано отправитель (from_legal_id)'}, status=400)
        if not to_client_id:
            return Response({'error': 'Не указан получатель (to_client_id)'}, status=400)

        from .models import LegalEntity, Client, InventoryItem

        try:
            from_legal = LegalEntity.objects.get(id=from_legal_id)
        except LegalEntity.DoesNotExist:
            return Response({'error': f'Юрлицо id={from_legal_id} не найдено'}, status=400)

        try:
            to_client = Client.objects.get(id=to_client_id)
        except Client.DoesNotExist:
            return Response({'error': f'Клиент id={to_client_id} не найден'}, status=400)

        invoice = OutgoingInvoice.objects.create(
            from_legal=from_legal,
            to_client=to_client,
            basis=data.get('basis', ''),
            received_by_name=data.get('received_by_name', ''),
            notes=data.get('notes', ''),
        )

        errors = []
        for item_data in items_data:
            item_id = item_data.get('inventory_item')
            if not item_id:
                errors.append('Пропущена позиция без inventory_item')
                continue
            try:
                inv_item = InventoryItem.objects.get(id=item_id)
                qty = item_data.get('quantity', 1)
                price = item_data.get('unit_price', inv_item.sale_price or 0)
                OutgoingInvoiceItem.objects.create(
                    invoice=invoice,
                    inventory_item=inv_item,
                    quantity=qty,
                    unit_price=price,
                    vat_rate=item_data.get('vat_rate', '20%'),
                    notes=item_data.get('notes', ''),
                )
            except InventoryItem.DoesNotExist:
                errors.append(f"Товар id={item_id} не найден")

        return Response({
            'success': True,
            'invoice': OutgoingInvoiceSerializer(invoice).data,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def issue(self, request, pk=None):
        """Провести выдачу: списать товары со склада"""
        invoice = self.get_object()
        if invoice.status != 'draft':
            return Response({'error': 'Накладная уже проведена или аннулирована'}, status=400)
        try:
            invoice.issue(request.user)
            return Response({'success': True, 'invoice': OutgoingInvoiceSerializer(invoice).data})
        except ValueError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Аннулировать накладную"""
        invoice = self.get_object()
        if invoice.status == 'cancelled':
            return Response({'error': 'Уже аннулирована'}, status=400)
        invoice.status = 'cancelled'
        invoice.save()
        return Response({'success': True, 'invoice': OutgoingInvoiceSerializer(invoice).data})

    @action(detail=False, methods=['get'])
    def print(self, request):
        """Данные для печати УПД"""
        invoice_id = request.query_params.get('id')
        if not invoice_id:
            return Response({'error': 'Укажите ?id=...'}, status=400)
        try:
            invoice = OutgoingInvoice.objects.prefetch_related(
                'items__inventory_item', 'from_legal', 'to_client'
            ).get(id=invoice_id)
        except OutgoingInvoice.DoesNotExist:
            return Response({'error': 'Накладная не найдена'}, status=404)

        return Response({
            'number': invoice.number,
            'date': invoice.date.isoformat(),
            'from_legal': {
                'name': invoice.from_legal.name,
                'short_name': invoice.from_legal.short_name,
                'inn': invoice.from_legal.inn,
                'kpp': invoice.from_legal.kpp,
                'address': invoice.from_legal.legal_address,
            },
            'to_client': {
                'name': invoice.to_client.name,
                'phone': invoice.to_client.phone,
                'address': invoice.to_client.address,
                'inn': invoice.to_client.inn,
                'is_legal': invoice.to_client.is_legal,
            },
            'basis': invoice.basis,
            'received_by_name': invoice.received_by_name,
            'total_amount': str(invoice.total_amount),
            'total_vat': str(invoice.total_vat),
            'items': [{
                'name': item.inventory_item.name,
                'barcode': item.inventory_item.barcode,
                'unit': item.inventory_item.unit,
                'quantity': item.quantity,
                'unit_price': str(item.unit_price),
                'amount': str(item.amount),
                'vat_rate': item.vat_rate,
            } for item in invoice.items.all()],
        })


# ══════════════════════════════════════════════════════════════════
# Asterisk PBX — управление телефонией
# ══════════════════════════════════════════════════════════════════

class AsteriskSipPeerViewSet(viewsets.ModelViewSet):
    """SIP-аккаунты (внутренние номера)"""
    queryset = AsteriskSipPeer.objects.all().order_by('name')
    serializer_class = AsteriskSipPeerSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'display_name', 'caller_id']

    def perform_create(self, serializer):
        # Генерируем пароль, если не задан
        if not serializer.validated_data.get('secret'):
            import secrets
            serializer.validated_data['secret'] = secrets.token_urlsafe(12)
        serializer.save()


class AsteriskTrunkViewSet(viewsets.ModelViewSet):
    """SIP-транки (подключение к провайдерам)"""
    queryset = AsteriskTrunk.objects.all().order_by('name')
    serializer_class = AsteriskTrunkSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'provider', 'host']


class AsteriskRouteViewSet(viewsets.ModelViewSet):
    """Маршруты звонков"""
    queryset = AsteriskRoute.objects.select_related('trunk').all().order_by('direction', 'priority')
    serializer_class = AsteriskRouteSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['direction', 'is_active']
    search_fields = ['name', 'match_pattern']


class AsteriskIvrViewSet(viewsets.ModelViewSet):
    """Голосовые меню (IVR)"""
    queryset = AsteriskIvr.objects.prefetch_related('options').all().order_by('name')
    serializer_class = AsteriskIvrSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description']

    @action(detail=True, methods=['post'])
    def add_option(self, request, pk=None):
        """Добавить опцию в IVR-меню"""
        ivr = self.get_object()
        serializer = AsteriskIvrOptionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(ivr=ivr)
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['delete'], url_path='remove-option/(?P<option_id>[^/.]+)')
    def remove_option(self, request, pk=None, option_id=None):
        """Удалить опцию из IVR-меню"""
        ivr = self.get_object()
        opt = get_object_or_404(AsteriskIvrOption, id=option_id, ivr=ivr)
        opt.delete()
        return Response({'ok': True})


class AsteriskIvrOptionViewSet(viewsets.ModelViewSet):
    """Опции IVR (без вложенности)"""
    queryset = AsteriskIvrOption.objects.all()
    serializer_class = AsteriskIvrOptionSerializer


class AsteriskVoicemailViewSet(viewsets.ModelViewSet):
    """Автоответчики / голосовая почта"""
    queryset = AsteriskVoicemail.objects.all().order_by('mailbox')
    serializer_class = AsteriskVoicemailSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['mailbox', 'display_name', 'email']


class AsteriskCallRecordingViewSet(viewsets.ReadOnlyModelViewSet):
    """Записи звонков"""
    queryset = AsteriskCallRecording.objects.select_related('client').all().order_by('-start_time')
    serializer_class = AsteriskCallRecordingSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['direction']
    search_fields = ['caller', 'callee', 'client__name']


# ══════════════════════════════════════════════════════════════════
# Ростелеком АТС — журнал звонков
# ══════════════════════════════════════════════════════════════════

class CallLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Журнал звонков (CDR) из Ростелеком АТС"""
    queryset = CallLog.objects.select_related('client').order_by('-start_time')
    serializer_class = CallLogSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['direction', 'status', 'call_type']
    search_fields = ['phone', 'client__name']


# ══════════════════════════════════════════════════════════════════
# ЕРЦ (Единый расчётный центр)
# ══════════════════════════════════════════════════════════════════

class ErcAccountViewSet(viewsets.ModelViewSet):
    """Лицевые счета ЕРЦ"""
    queryset = ErcAccount.objects.all()
    serializer_class = ErcAccountSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active']
    search_fields = ['account_number', 'full_name', 'address']


class ErcBillingRecordViewSet(viewsets.ModelViewSet):
    """Платёжные записи ЕРЦ"""
    queryset = ErcBillingRecord.objects.all()
    serializer_class = ErcBillingRecordSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['account', 'period']
    search_fields = ['account__account_number', 'account__full_name', 'account__address']


# ══════════════════════════════════════════════════════════════════
# Справочники: подъезды, УК, тарифы, внутренние платежи
# ══════════════════════════════════════════════════════════════════

class BuildingEntranceViewSet(viewsets.ModelViewSet):
    """Подъезды домов"""
    queryset = BuildingEntrance.objects.select_related('building', 'building__management_company_fk', 'building__region').all()
    serializer_class = BuildingEntranceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['building', 'building__management_company_fk', 'building__region']
    search_fields = ['building__street_name', 'building__city', 'ip_address', 'access_code', 'programming_code', 'notes']
    ordering_fields = ['building__street_name', 'building__city', 'building__apartments_count', 'building__management_company_fk__name', 'number']
    ordering = ['building__street_name']
    pagination_class = None

    @action(detail=True, methods=['get'])
    def orders(self, request, pk=None):
        """История заявок по подъезду (ищем по адресу дома)."""
        entrance = self.get_object()
        building = entrance.building
        orders = Order.objects.filter(building=building).select_related('master__user').order_by('-created_at')
        return Response([{'id': o.id, 'number': o.number, 'order_type': o.order_type,
                          'order_type_display': o.get_order_type_display(), 'status': o.status,
                          'status_display': o.get_status_display(),
                          'master_name': o.master.user.get_full_name() or o.master.user.username if o.master else '—',
                          'created_at': o.created_at.isoformat()} for o in orders])

    @action(detail=True, methods=['get'])
    def apartments(self, request, pk=None):
        """Список квартир в подъезде (клиенты)."""
        entrance = self.get_object()
        building = entrance.building
        clients = Client.objects.filter(building=building).order_by('apartment')
        if entrance.apartment_from and entrance.apartment_to:
            # Пытаемся отфильтровать по диапазону
            try:
                from_num = int(entrance.apartment_from)
                to_num = int(entrance.apartment_to)
                # apartment — CharField, фильтруем примерно
                result = []
                for c in clients:
                    if c.apartment:
                        try:
                            ap = int(c.apartment)
                            if from_num <= ap <= to_num:
                                result.append({'id': c.id, 'name': c.name, 'apartment': c.apartment,
                                               'phone': c.phone, 'personal_account_number': c.personal_account_number})
                        except ValueError:
                            result.append({'id': c.id, 'name': c.name, 'apartment': c.apartment,
                                           'phone': c.phone, 'personal_account_number': c.personal_account_number})
                return Response(result)
            except (ValueError, TypeError):
                pass
        result = [{'id': c.id, 'name': c.name, 'apartment': c.apartment,
                   'phone': c.phone, 'personal_account_number': c.personal_account_number}
                  for c in clients[:100]]
        return Response(result)


class ManagementCompanyViewSet(viewsets.ModelViewSet):
    """Управляющие компании / ТСЖ"""
    queryset = ManagementCompany.objects.all()
    serializer_class = ManagementCompanySerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'short_name', 'inn']
    pagination_class = None  # все 49 компаний сразу

    @action(detail=True, methods=['get'])
    def buildings(self, request, pk=None):
        """Список домов, которые обслуживает эта УК (из Building FK)"""
        company = self.get_object()
        buildings = company.buildings_list.select_related('region').order_by('street_name', 'house_number')

        return Response([{
            'id': b.id,
            'city': b.city,
            'street_name': b.street_name,
            'house_number': b.house_number,
            'building_number': b.building_number,
            'apartments_count': b.apartments_count,
            'entrances_count': b.entrances_count,
            'management_company_fk': b.management_company_fk_id,
            'clients_count': b.residents.count(),
        } for b in buildings])

    @action(detail=True, methods=['post'])
    def add_building(self, request, pk=None):
        """Привязать дом к этой УК"""
        company = self.get_object()
        building_id = request.data.get('building_id')
        if not building_id:
            return Response({'error': 'Укажите building_id'}, status=400)
        building = get_object_or_404(Building, id=building_id)
        building.management_company_fk = company
        building.management_company = company.name
        building.save()
        return Response({'ok': True})

    @action(detail=True, methods=['post'])
    def remove_building(self, request, pk=None):
        """Отвязать дом от этой УК"""
        company = self.get_object()
        building_id = request.data.get('building_id')
        if not building_id:
            return Response({'error': 'Укажите building_id'}, status=400)
        building = get_object_or_404(Building, id=building_id, management_company_fk=company)
        building.management_company_fk = None
        building.save()
        return Response({'ok': True})

    @action(detail=True, methods=['post'])
    def generate_clients(self, request, pk=None):
        """Создать клиентов-квартиры для дома этой УК по подъездам."""
        company = self.get_object()
        building_id = request.data.get('building_id')
        building = get_object_or_404(Building, id=building_id)
        created = 0
        for entrance in building.entrances.all():
            if entrance.apartment_from and entrance.apartment_to:
                for apt in range(entrance.apartment_from, entrance.apartment_to + 1):
                    _, c = Client.objects.get_or_create(
                        building=building, apartment=str(apt),
                        defaults={
                            'name': f'Квартира {apt}',
                            'address': f'г. {building.city}, {building.street_name}, д. {building.house_number}, кв. {apt}',
                            'management_company': company,
                            'entrance': entrance,
                            'source': 'manual', 'erc_enabled': True,
                        }
                    )
                    created += 1
        return Response({'ok': True, 'building_id': building.id, 'clients_created': created})

    @action(detail=True, methods=['post'])
    def apply_tariff(self, request, pk=None):
        """Применить тариф ко всем клиентам УК: сумма = тариф × квартир в доме."""
        company = self.get_object()
        tariff_id = request.data.get('tariff_id')
        building_id = request.data.get('building_id')  # опционально: только для одного дома

        if not tariff_id:
            return Response({'error': 'Укажите tariff_id'}, status=400)
        tariff = get_object_or_404(Tariff, id=tariff_id)

        clients_qs = Client.objects.filter(management_company=company)
        if building_id:
            clients_qs = clients_qs.filter(building_id=building_id)

        # Считаем сумму: тариф × количество квартир в доме
        if not building_id:
            # Для каждого дома УК: своя сумма
            for bld in company.buildings_list.all():
                clients_in_bld = clients_qs.filter(building=bld)
                clients_in_bld.update(tariff=tariff, monthly_payment=tariff.amount * bld.apartments_count)
        else:
            building = get_object_or_404(Building, id=building_id)
            clients_qs.update(tariff=tariff, monthly_payment=tariff.amount * building.apartments_count)

        updated = clients_qs.count()
        return Response({'ok': True, 'updated_clients': updated, 'tariff': TariffSerializer(tariff).data,
                         'monthly_total': float(tariff.amount) * (building.apartments_count if building_id else sum(b.apartments_count for b in company.buildings_list.all()))})

    @action(detail=True, methods=['get'])
    def orders(self, request, pk=None):
        """История заявок по всем домам этой УК"""
        company = self.get_object()
        building_ids = company.buildings_list.values_list('id', flat=True)
        orders = Order.objects.filter(building_id__in=building_ids).select_related('master__user').order_by('-created_at')
        return Response([{
            'id': o.id, 'number': o.number, 'order_type': o.order_type,
            'order_type_display': o.get_order_type_display(), 'status': o.status,
            'status_display': o.get_status_display(),
            'master_name': o.master.user.get_full_name() or o.master.user.username if o.master else '—',
            'created_at': o.created_at.isoformat(),
        } for o in orders])

    @action(detail=True, methods=['get', 'post'])
    def contacts(self, request, pk=None):
        """Контакты УК"""
        company = self.get_object()
        if request.method == 'GET':
            qs = MCContact.objects.filter(management_company=company)
            return Response(MCContactSerializer(qs, many=True).data)
        # POST
        serializer = MCContactSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(management_company=company)
        return Response(serializer.data, status=201)

    @action(detail=True, methods=['delete'], url_path='contacts/(?P<contact_id>[^/.]+)')
    def delete_contact(self, request, pk=None, contact_id=None):
        company = self.get_object()
        get_object_or_404(MCContact, id=contact_id, management_company=company).delete()
        return Response({'ok': True})

    @action(detail=True, methods=['get', 'post'])
    def comments(self, request, pk=None):
        """История обращений УК"""
        company = self.get_object()
        if request.method == 'GET':
            qs = MCComment.objects.filter(management_company=company).select_related('author')
            return Response(MCCommentSerializer(qs, many=True).data)
        # POST
        serializer = MCCommentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(management_company=company, author=request.user)
        return Response(serializer.data, status=201)

    @action(detail=True, methods=['get', 'post'])
    def payments(self, request, pk=None):
        """Бухгалтерия: начисления и оплаты УК"""
        company = self.get_object()
        if request.method == 'GET':
            qs = MCPayment.objects.filter(management_company=company)
            return Response(MCPaymentSerializer(qs, many=True).data)
        # POST
        serializer = MCPaymentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(management_company=company)
        return Response(serializer.data, status=201)

    @action(detail=True, methods=['post'])
    def confirm_payment(self, request, pk=None):
        """Бухгалтер подтверждает поступление оплаты"""
        payment_id = request.data.get('payment_id')
        payment = get_object_or_404(MCPayment, id=payment_id, management_company_id=pk)
        payment.is_confirmed = True
        payment.confirmed_by = request.user
        payment.confirmed_at = timezone.now()
        payment.save()
        return Response(MCPaymentSerializer(payment).data)

    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Включить/выключить обслуживание УК"""
        company = self.get_object()
        company.is_active = not company.is_active
        if not company.is_active:
            company.terminated_at = timezone.now().date()
            company.termination_reason = request.data.get('reason', '')
            # Обновляем статус клиентов
            Client.objects.filter(management_company=company).update(
                erc_enabled=False, notes=f'[АВТО] УК снята с обслуживания: {company.termination_reason}'
            )
        else:
            company.terminated_at = None
            company.termination_reason = ''
            Client.objects.filter(management_company=company).update(erc_enabled=True)
        company.save()
        return Response(ManagementCompanySerializer(company).data)

    @action(detail=False, methods=['post'])
    def create_with_buildings(self, request):
        """
        Создать УК/ТСЖ сразу с домами, подъездами и квартирами.
        """
        company_data = request.data.get('company', {})
        buildings_data = request.data.get('buildings', [])

        serializer = ManagementCompanySerializer(data=company_data)
        serializer.is_valid(raise_exception=True)
        company = serializer.save()

        stats = {'buildings': 0, 'entrances': 0, 'clients': 0}
        for bld_data in buildings_data:
            entrances_list = bld_data.pop('entrances', [])
            bld_data['management_company_fk'] = company
            bld_data['management_company'] = company.name

            building, created = Building.objects.get_or_create(
                city=bld_data.get('city', 'Санкт-Петербург'),
                street_name=bld_data.get('street_name', ''),
                house_number=bld_data.get('house_number', ''),
                building_number=bld_data.get('building_number', ''),
                defaults={k: v for k, v in bld_data.items()
                          if k not in ['city', 'street_name', 'house_number', 'building_number']}
            )
            if not building.management_company_fk:
                building.management_company_fk = company
                building.management_company = company.name
                building.save()
            stats['buildings'] += 1

            for ent_data in entrances_list:
                create_clients = ent_data.pop('create_clients', False)
                ent_num = ent_data.get('number', 1)
                apt_from = ent_data.get('apartments_from', 0)
                apt_to = ent_data.get('apartments_to', 0)

                entrance, _ = BuildingEntrance.objects.get_or_create(
                    building=building, number=ent_num,
                    defaults={'apartment_from': apt_from, 'apartment_to': apt_to,
                              'apartments_count': max(0, apt_to - apt_from + 1)}
                )
                if apt_from and apt_to and (entrance.apartment_from != apt_from or entrance.apartment_to != apt_to):
                    entrance.apartment_from = apt_from
                    entrance.apartment_to = apt_to
                    entrance.apartments_count = apt_to - apt_from + 1
                    entrance.save()
                stats['entrances'] += 1

                if create_clients and apt_from and apt_to:
                    for apt_num in range(apt_from, apt_to + 1):
                        Client.objects.get_or_create(
                            building=building, apartment=str(apt_num),
                            defaults={
                                'name': f'Квартира {apt_num}',
                                'address': f'г. {building.city}, {building.street_name}, д. {building.house_number}, кв. {apt_num}',
                                'management_company': company, 'entrance': entrance,
                                'source': 'manual', 'erc_enabled': True,
                            }
                        )
                        stats['clients'] += 1

        return Response({'ok': True, 'company': ManagementCompanySerializer(company).data, 'stats': stats}, status=201)


class TariffViewSet(viewsets.ModelViewSet):
    """Тарифы на обслуживание"""
    queryset = Tariff.objects.all()
    serializer_class = TariffSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']


class PaymentRecordViewSet(viewsets.ModelViewSet):
    """Внутренние платежи (не ЕРЦ)"""
    queryset = PaymentRecord.objects.select_related('client').all()
    serializer_class = PaymentRecordSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['client', 'payment_type', 'period']


class BewardDeviceViewSet(viewsets.ModelViewSet):
    """Справочник IP-адресов панелей Beward (умные домофоны)"""
    queryset = BewardDevice.objects.select_related('building', 'entrance').all()
    serializer_class = BewardDeviceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['building', 'entrance']
    search_fields = ['ip_address', 'address', 'region', 'notes']
    pagination_class = None  # все записи сразу (9555 шт.), фильтрация на фронте


# ══════════════════════════════════════════════════════════════════
# Эндпоинты импорта из Excel
# ══════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_clients_excel_view(request):
    """
    Импорт базы клиентов из Excel-файла.
    Принимает multipart/form-data с полем 'file'.
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=status.HTTP_400_BAD_REQUEST)

        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'error': 'Поддерживаются только файлы .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        from .import_service import import_clients_from_excel
        # Принимаем опциональный column_map для гибкого маппинга колонок
        column_map = None
        if 'column_map' in request.data:
            try:
                import json
                raw_map = request.data['column_map']
                if isinstance(raw_map, str):
                    column_map = json.loads(raw_map)
                else:
                    column_map = dict(raw_map)
                # Конвертируем ключи в int
                column_map = {k: int(v) for k, v in column_map.items() if v}
            except (ValueError, TypeError):
                column_map = None

        result = import_clients_from_excel(uploaded.read(), request.user, column_map)
        return Response(result)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Ошибка импорта: {str(e)}',
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_erc_excel_view(request):
    """
    Импорт данных ЕРЦ из Excel-файла (оборотная ведомость).
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=status.HTTP_400_BAD_REQUEST)

        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'error': 'Поддерживаются только файлы .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        period_date = None
        period_str = request.data.get('period', '')
        if period_str:
            from datetime import datetime
            try:
                period_date = datetime.strptime(period_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'success': False, 'error': 'Формат периода: YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

        from .converters import auto_convert
        from .unified_importer import import_unified_csv
        import csv, io

        # Конвертируем Excel в унифицированные CSV-строки через автоопределение формата
        rows, format_name = auto_convert(uploaded.read(), period_date)

        if not rows:
            return Response({
                'success': False,
                'error': 'Не удалось определить формат файла или файл пуст',
            }, status=status.HTTP_400_BAD_REQUEST)

        # Создаём CSV в памяти из унифицированных строк
        from .converters import UNIFIED_FIELDS
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=UNIFIED_FIELDS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)

        # Импортируем полученный CSV в БД
        csv_bytes = output.getvalue().encode('utf-8-sig')
        stats = import_unified_csv(csv_bytes, request.user)

        # Приводим ответ к формату, ожидаемому фронтендом
        result = {
            'success': True,
            'total': stats['total_rows'],
            'created': stats['erc_records_created'] + stats['clients_created'],
            'updated': stats['erc_records_updated'] + stats['clients_updated'],
            'errors': stats['errors'],
            'format': format_name,
            'details': {
                'payment_rows': stats['payment_rows'],
                'client_rows': stats['client_rows'],
                'buildings_created': stats['buildings_created'],
                'clients_created': stats['clients_created'],
                'clients_updated': stats['clients_updated'],
                'erc_accounts_created': stats['erc_accounts_created'],
                'erc_records_created': stats['erc_records_created'],
                'erc_records_updated': stats['erc_records_updated'],
                'skipped': stats['skipped'],
            },
        }

        return Response(result)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Ошибка импорта ЕРЦ: {str(e)}',
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_preview_view(request):
    """
    Предпросмотр Excel-файла (первые 10 строк).
    """
    if 'file' not in request.FILES:
        return Response({'success': False, 'error': 'Файл не прикреплён'}, status=status.HTTP_400_BAD_REQUEST)

    uploaded = request.FILES['file']
    if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
        return Response({'success': False, 'error': 'Поддерживаются только файлы .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

    from .import_service import preview_excel
    result = preview_excel(uploaded.read())
    return Response(result)


# ══════════════════════════════════════════════════════════════════
# Новые эндпоинты: Конвертация Excel → CSV + Унифицированный импорт
# ══════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def convert_excel_view(request):
    """
    Конвертирует загруженный Excel (любого формата: ТСЖ, ЕРЦ СПб, ЕРЦ ЛО, ...)
    в унифицированный CSV.
    
    Принимает:
      - file: .xlsx файл
      - period (опционально): дата периода YYYY-MM-DD для ЕРЦ
    
    Возвращает:
      - converted_rows: массив строк унифицированного CSV (первые 200 для предпросмотра)
      - total_rows: всего строк
      - format: определённый формат (tszh, erc_spb, erc_lo, ...)
      - csv_content: полный CSV как текст (для скачивания)
      - csv_filename: предлагаемое имя файла
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=status.HTTP_400_BAD_REQUEST)

        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'error': 'Поддерживаются только файлы .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        # Период
        from datetime import datetime
        period_date = None
        period_str = request.data.get('period', '')
        if period_str:
            try:
                period_date = datetime.strptime(period_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'success': False, 'error': 'Формат периода: YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

        from .converters import auto_convert

        file_bytes = uploaded.read()
        rows, fmt = auto_convert(file_bytes, period_date)

        # Генерируем CSV
        import csv, io
        from .converters import UNIFIED_FIELDS

        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=UNIFIED_FIELDS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)
        csv_content = buf.getvalue()

        basename = uploaded.name.rsplit('.', 1)[0]

        return Response({
            'success': True,
            'format': fmt,
            'total_rows': len(rows),
            'converted_preview': rows[:200],  # первые 200 для отображения
            'csv_content': csv_content,
            'csv_filename': f'{basename}.csv',
        })

    except Exception as e:
        import traceback
        return Response({
            'success': False,
            'error': f'Ошибка конвертации: {str(e)}',
            'traceback': traceback.format_exc(),
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_unified_view(request):
    """
    Импортирует унифицированный CSV в базу данных.
    
    Принимает:
      - file: .csv файл (унифицированный формат)
      - или csv_text: CSV-текст напрямую (если конвертировали только что)
    
    Возвращает статистику импорта.
    """
    try:
        csv_data = None
        csv_filename = 'data.csv'

        if 'file' in request.FILES:
            uploaded = request.FILES['file']
            csv_data = uploaded.read()
            csv_filename = uploaded.name
        elif 'csv_text' in request.data:
            csv_data = request.data['csv_text'].encode('utf-8')
        else:
            return Response({'success': False, 'error': 'Нет файла или csv_text'}, status=status.HTTP_400_BAD_REQUEST)

        from .unified_importer import import_unified_csv

        stats = import_unified_csv(csv_data, request.user, dry_run=False)

        return Response({
            'success': True,
            'file': csv_filename,
            'total_rows': stats['total_rows'],
            'client_rows': stats['client_rows'],
            'payment_rows': stats['payment_rows'],
            'buildings_created': stats['buildings_created'],
            'clients_created': stats['clients_created'],
            'clients_updated': stats['clients_updated'],
            'erc_accounts_created': stats['erc_accounts_created'],
            'erc_records_created': stats['erc_records_created'],
            'erc_records_updated': stats['erc_records_updated'],
            'skipped': stats['skipped'],
            'errors': stats['errors'],
        })

    except Exception as e:
        import traceback
        return Response({
            'success': False,
            'error': f'Ошибка импорта: {str(e)}',
            'traceback': traceback.format_exc(),
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ══════════════════════════════════════════════════════════════════
# Универсальный импорт клиентов (CSV) — единый формат
# 14 полей клиента + 5 полей ЕРЦ (опционально)
# ══════════════════════════════════════════════════════════════════

UNIVERSAL_CLIENT_FIELDS = [
    'city', 'region', 'district', 'street_name', 'house_number',
    'building_number', 'apartment', 'entrance_number',
    'full_name', 'personal_account', 'phone', 'source', 'source_file',
    'period', 'balance_start', 'charged', 'paid', 'balance_end',
]

SAMPLE_CSV_CONTENT = """city,region,district,street_name,house_number,building_number,apartment,entrance_number,full_name,personal_account,phone,source,source_file,period,balance_start,charged,paid,balance_end
Коммунар,Ленинградская обл,Гатчинский р-н,Бумажников,2,,3,1,Небензя Эдуард Александрович,070000081442,,АО ЕИРЦ коммунар,05_ЕИРЦ Коммунар_май 2026.xlsx,2026-05-01,0,0,0,0
Коммунар,Ленинградская обл,Гатчинский р-н,Бумажников,2,,5,1,Смородинов Валентин Николаевич,070000081444,,АО ЕИРЦ коммунар,05_ЕИРЦ Коммунар_май 2026.xlsx,2026-05-01,0,0,0,0
Агалатово,Ленинградская обл,Всеволожский р-н,,144,1,34,1,Иванов Петр Сергеевич,050000123456,,"ТСЖ Агалатово",агалатово_май.xlsx,2026-05-01,1200.50,300,300,1200.50
Санкт-Петербург,Ленинградская обл,,Ленина,5,,1,1,Иванова Мария,120000012345,79111234567,ТСЖ Пример,example.xlsx,2026-06-01,500,250,250,500
Санкт-Петербург,Ленинградская обл,,Мира,10,корп.1,42,1,Сидоров Алексей,120000012346,79330001122,Жилкомсервис,example.xlsx,,0,0,0,0"""


@api_view(['GET'])
@permission_classes([AllowAny])
def download_sample_csv_view(request):
    """Скачать образец CSV для импорта клиентов (без авторизации)."""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="clients_sample.csv"'
    response.write('\uFEFF')  # BOM для Excel
    response.write(SAMPLE_CSV_CONTENT)
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_clients_csv_view(request):
    """
    Единый импорт: клиенты + начисления ЕРЦ из одного CSV.

    Поля (19):
      city, region, district, street_name, house_number, building_number,
      apartment, entrance_number, full_name, personal_account, phone,
      source, source_file, period, balance_start, charged, paid, balance_end

    Уникальность клиента — по personal_account.
    Здание — по city + street_name + house_number.
    Если period заполнен — создаётся ErcBillingRecord.
    """
    try:
        csv_data = None
        if 'file' in request.FILES:
            csv_data = request.FILES['file'].read().decode('utf-8-sig')
        elif 'csv_text' in request.data:
            csv_data = request.data['csv_text']
        else:
            return Response({'success': False, 'error': 'Нет файла или csv_text'}, status=400)

        import csv as csv_mod, io
        reader = csv_mod.DictReader(io.StringIO(csv_data))
        rows = list(reader)

        if not rows:
            return Response({'success': False, 'error': 'CSV пустой'}, status=400)

        from .models import (
            Client as ClientModel, Building, ManagementCompany, BuildingEntrance,
            ErcAccount, ErcBillingRecord,
        )
        from datetime import datetime

        created_clients = 0
        updated_clients = 0
        created_buildings = 0
        created_entrances = 0
        created_erc = 0
        updated_erc = 0
        errors = []
        skipped_no_account = 0

        for i, row in enumerate(rows, 2):
            try:
                city = (row.get('city') or row.get('город') or '').strip()
                region = (row.get('region') or row.get('область') or '').strip()
                district = (row.get('district') or row.get('район') or '').strip()
                street_name = (row.get('street_name') or row.get('street') or row.get('улица') or '').strip()
                house_number = (row.get('house_number') or row.get('house') or row.get('дом') or '').strip()
                building_number = (row.get('building_number') or row.get('building') or row.get('корпус') or row.get('строение') or '').strip()
                apartment = (row.get('apartment') or row.get('квартира') or row.get('кв') or '').strip()
                entrance = (row.get('entrance_number') or row.get('entrance') or row.get('подъезд') or row.get('под') or '').strip()
                full_name = (row.get('full_name') or row.get('фио') or row.get('name') or '').strip()
                personal_account = (row.get('personal_account') or row.get('лицевой_счет') or row.get('лс') or row.get('account') or '').strip()
                phone = (row.get('phone') or row.get('телефон') or row.get('тел') or '').strip()
                source = (row.get('source') or row.get('источник') or row.get('ук') or row.get('management_company') or '').strip()
                source_file = (row.get('source_file') or row.get('файл') or '').strip()

                # --- ЕРЦ поля ---
                period_str = (row.get('period') or '').strip()
                balance_start = (row.get('balance_start') or '0').strip()
                charged = (row.get('charged') or '0').strip()
                paid = (row.get('paid') or '0').strip()
                balance_end = (row.get('balance_end') or '0').strip()

                # --- Пропуск мусора ---
                if not personal_account or len(personal_account) < 5:
                    skipped_no_account += 1
                    continue
                if not personal_account.isdigit():
                    skipped_no_account += 1
                    continue
                skip_words = ['лицевого', 'счета', 'лицевой', 'номер', 'счёт', 'полное', 'фио', 'account']
                if full_name.lower() in skip_words:
                    continue
                if not full_name:
                    full_name = 'Не определено'

                # --- Нормализация ---
                if not city:
                    city = 'Санкт-Петербург'
                import re as _re
                clean_street = _re.sub(r'^(ул\.?|улица)\s+', '', street_name, flags=_re.IGNORECASE).strip()

                # --- Адресная строка ---
                address_parts = [f'г. {city}']
                if region:
                    address_parts.append(region)
                if district:
                    address_parts.append(district)
                if clean_street:
                    address_parts.append(f'ул. {clean_street}')
                if house_number:
                    addr_house = f'д. {house_number}'
                    if building_number:
                        addr_house += f' корп. {building_number}'
                    address_parts.append(addr_house)
                if apartment:
                    address_parts.append(f'кв. {apartment}')
                full_address = ', '.join(address_parts)

                # --- Здание ---
                building = None
                if clean_street and house_number:
                    building = Building.objects.filter(
                        city__icontains=city,
                        street_name__icontains=clean_street,
                        house_number=house_number,
                    ).first()
                    if not building:
                        building = Building.objects.create(
                            city=city,
                            street_name=clean_street,
                            house_number=house_number,
                            building_number=building_number,
                        )
                        created_buildings += 1
                    elif not building.city:
                        building.city = city
                        building.save(update_fields=['city'])
                elif house_number:
                    # Деревня/посёлок без улицы
                    building = Building.objects.filter(
                        city__icontains=city,
                        street_name='',
                        house_number=house_number,
                    ).first()
                    if not building:
                        building = Building.objects.create(
                            city=city,
                            street_name='',
                            house_number=house_number,
                            building_number=building_number,
                        )
                        created_buildings += 1

                # --- Подъезд ---
                entrance_obj = None
                if building and entrance:
                    try:
                        ent_num = int(float(entrance))
                        entrance_obj, ent_created = BuildingEntrance.objects.get_or_create(
                            building=building, number=ent_num,
                            defaults={'apartments_count': 0}
                        )
                        if ent_created:
                            created_entrances += 1
                    except (ValueError, TypeError):
                        pass

                # --- УК ---
                mc_obj = None
                if source:
                    mc_obj, _ = ManagementCompany.objects.get_or_create(name=source)
                    if building and mc_obj and not building.management_company_fk:
                        building.management_company_fk = mc_obj
                        building.save(update_fields=['management_company_fk'])

                # --- Клиент (КЛЮЧ = personal_account) ---
                existing = ClientModel.objects.filter(personal_account_number=personal_account).first()

                if existing:
                    if building and not existing.building:
                        existing.building = building
                    if entrance_obj and not existing.entrance:
                        existing.entrance = entrance_obj
                    if not existing.address or existing.address == 'г. , д. ':
                        existing.address = full_address
                    if phone and not existing.phone:
                        existing.phone = phone
                    if full_name and existing.name == 'Не определено':
                        existing.name = full_name
                    existing.save()
                    updated_clients += 1
                else:
                    ClientModel.objects.create(
                        name=full_name,
                        phone=phone,
                        address=full_address,
                        building=building,
                        entrance=entrance_obj,
                        apartment=apartment,
                        management_company=mc_obj,
                        personal_account_number=personal_account,
                        source='erc' if 'ерц' in (source + source_file).lower() else 'manual',
                    )
                    created_clients += 1

                # --- ЕРЦ начисление (если есть period) ---
                if period_str:
                    try:
                        period_date = datetime.strptime(period_str, '%Y-%m-%d').date()
                        if period_date.day != 1:
                            period_date = period_date.replace(day=1)

                        erc_account, _ = ErcAccount.objects.get_or_create(
                            account_number=personal_account,
                            defaults={'full_name': full_name, 'address': full_address}
                        )

                        billing, billing_created = ErcBillingRecord.objects.update_or_create(
                            account=erc_account,
                            period=period_date,
                            defaults={
                                'balance_start': float(balance_start) if balance_start else 0,
                                'charged': float(charged) if charged else 0,
                                'paid': float(paid) if paid else 0,
                                'balance_end': float(balance_end) if balance_end else 0,
                            }
                        )
                        if billing_created:
                            created_erc += 1
                        else:
                            updated_erc += 1
                    except Exception:
                        pass  # не фатально

            except Exception as e:
                errors.append(f'Строка {i}: {str(e)}')

        import logging
        logging.warning(
            f'[IMPORT] rows={len(rows)} clients: +{created_clients} ~{updated_clients} '
            f'skip={skipped_no_account} bld={created_buildings} entr={created_entrances} '
            f'erc: +{created_erc} ~{updated_erc}'
        )

        return Response({
            'success': True,
            'total_rows': len(rows),
            'clients_created': created_clients,
            'clients_updated': updated_clients,
            'skipped_no_account': skipped_no_account,
            'buildings_created': created_buildings,
            'entrances_created': created_entrances,
            'erc_billing_created': created_erc,
            'erc_billing_updated': updated_erc,
            'errors': errors[:20],
        })

    except Exception as e:
        import traceback
        return Response({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc(),
        }, status=500)


# ══════════════════════════════════════════════════════════════════
# Прямой импорт XLSX с Dadata (все форматы ЕРЦ)
# ══════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_xlsx_preview_view(request):
    """
    Предпросмотр XLSX: возвращает первые 20 строк в унифицированном формате БЕЗ импорта в базу.
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=400)
        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith('.xlsx'):
            return Response({'success': False, 'error': 'Только .xlsx'}, status=400)

        import tempfile, os
        from openpyxl import load_workbook
        from .xlsx_importer import _detect_format, _extract_row, _find_data_start
        from .dadata_service import normalize_address

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        fmt = _detect_format(ws)
        data_start = _find_data_start(ws, fmt)
        preview = []

        for r in range(data_start, min(data_start + 20, ws.max_row + 1)):
            row_data = _extract_row(ws, r, fmt)
            if not row_data:
                continue

            personal_account = row_data.get('personal_account', '')
            name = row_data.get('name', '')
            raw_address = row_data.get('raw_address', '')

            if not personal_account or len(personal_account) < 5 or not personal_account.isdigit():
                continue
            skip_words = ['лицевого', 'счета', 'лицевой', 'номер', 'фамилия', 'фио']
            if any(w in name.lower() for w in skip_words):
                continue

            # Парсим адрес как при импорте
            pre_city = row_data.get('_city', '')
            pre_street = row_data.get('_street_name', '')
            pre_house = row_data.get('_house_number', '')
            pre_apt = row_data.get('_apartment', '')
            pre_district = row_data.get('_district', '')
            pre_region = row_data.get('_region', '')

            if pre_city or pre_house:
                city = pre_city or 'Санкт-Петербург'
                region_str = pre_region
                district = pre_district
                street_name = pre_street
                house_number = pre_house
                apartment = pre_apt
            else:
                addr = normalize_address(raw_address) if raw_address else {'success': False}
                city = addr.get('city', '') or 'Санкт-Петербург'
                region_str = addr.get('region', '')
                district = addr.get('district', '')
                street_name = addr.get('street_name', '')
                house_number = addr.get('house_number', '')
                apartment = row_data.get('apartment', '') or addr.get('apartment', '')

            preview.append({
                'л/с': personal_account,
                'ФИО': name,
                'регион': region_str,
                'район': district,
                'город': city,
                'улица': street_name,
                'дом': house_number,
                'квартира': apartment,
                'подъезд': row_data.get('entrance', '1') or '1',
                'сальдо': row_data.get('balance_start', 0),
                'начислено': row_data.get('charged', 0),
                'оплачено': row_data.get('paid', 0),
                '_raw': raw_address[:100],
            })

        os.unlink(tmp_path)
        return Response({'success': True, 'format': fmt, 'preview': preview, 'preview_count': len(preview)})
    except Exception as e:
        import traceback
        return Response({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_xlsx_direct_view(request):
    """
    Прямой импорт XLSX-файлов ЕРЦ с нормализацией адресов через Dadata.
    Поддерживает: Коммунар, СПб, Агалатово, ЛО, Красное Село, Стр.6-3, ТСЖ Битрикс.
    
    Принимает multipart: file (.xlsx)
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=400)
        
        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith('.xlsx'):
            return Response({'success': False, 'error': 'Только .xlsx'}, status=400)
        
        import tempfile, os
        from .xlsx_importer import import_xlsx_file
        
        # Сохраняем во временный файл
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        
        try:
            period_str = request.data.get('period', '')
            from datetime import datetime
            period_date = None
            if period_str:
                period_date = datetime.strptime(period_str, '%Y-%m-%d').date()
            
            stats = import_xlsx_file(tmp_path, source_filename=uploaded.name, period_date=period_date)
        finally:
            os.unlink(tmp_path)
        
        return Response({
            'success': True,
            **stats,
        })
    
    except Exception as e:
        import traceback
        return Response({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc(),
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_erc_update_view(request):
    """
    Обновление ЕРЦ: обновляет ТОЛЬКО начисления по номеру лицевого счёта.
    Не создаёт клиентов, не парсит адреса. 
    Принимает: file (.xlsx) + period (YYYY-MM-DD)
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=400)
        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith('.xlsx'):
            return Response({'success': False, 'error': 'Только .xlsx'}, status=400)

        import tempfile, os
        from .xlsx_importer import import_erc_update_only

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            period_str = request.data.get('period', '')
            from datetime import datetime
            period_date = None
            if period_str:
                period_date = datetime.strptime(period_str, '%Y-%m-%d').date()

            stats = import_erc_update_only(tmp_path, period_date=period_date)
        finally:
            os.unlink(tmp_path)

        return Response({
            'success': True,
            **stats,
        })
    except Exception as e:
        import traceback
        return Response({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc(),
        }, status=500)


# ══════════════════════════════════════════════════════════════════
# Импорт справочника Beward (IP-адреса и коды)
# ══════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_beward_ip_view(request):
    """
    Импорт справочника IP-адресов Beward из Excel.
    Формат: Район | Адрес | Подъезд | IP панели
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=400)

        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'error': 'Поддерживаются .xlsx/.xls'}, status=400)

        import openpyxl, re, tempfile, os

        # Сохраняем во временный файл (openpyxl требует путь к файлу)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
        finally:
            os.unlink(tmp_path)

        created, skipped, with_building = 0, 0, 0

        for row in rows:
            region = str(row[0]).strip() if row[0] else ''
            address = str(row[1]).strip() if row[1] else ''
            entrance_raw = str(row[2]).strip() if row[2] else ''
            ip_addr = str(row[3]).strip() if row[3] else ''

            if not ip_addr or not address:
                skipped += 1
                continue

            ent_num = ''
            try:
                ent_num = str(int(float(entrance_raw)))
            except (ValueError, TypeError):
                ent_num = entrance_raw

            # Пробуем найти Building
            building = None
            if address:
                addr = re.sub(r',(?!\s)', ', ', address)
                # Грубый парсинг
                m = re.search(r'(?:дом|дсм|д\.)\s*(\d+[а-яА-Я]?)', addr)
                house = m.group(1) if m else ''
                parts = [p.strip() for p in addr.split(',')]
                street_candidates = [p for p in parts if 'улица' in p.lower() or 'проспект' in p.lower() or 'шоссе' in p.lower() or 'пр-кт' in p.lower() or 'пер' in p.lower()]
                street = street_candidates[-1] if street_candidates else (parts[1] if len(parts) > 1 else parts[0] if parts else '')
                for prefix in ['улица ', 'проспект ', 'переулок ', 'шоссе ']:
                    if street.lower().startswith(prefix):
                        street = street[len(prefix):]
                if street and house:
                    qs = Building.objects.filter(house_number=house, street_name__icontains=street.strip())
                    building = qs.first()

            BewardDevice.objects.update_or_create(
                ip_address=ip_addr,
                defaults={
                    'region': region,
                    'address': address,
                    'entrance_number': ent_num,
                    'building': building,
                }
            )
            if building:
                with_building += 1
            created += 1

        return Response({
            'success': True,
            'total_rows': len(rows),
            'created': created,
            'updated': 0,
            'skipped': skipped,
            'with_building': with_building,
            'message': f'Импортировано {created} IP-адресов, {with_building} привязано к домам, пропущено {skipped}',
        })

    except Exception as e:
        import traceback
        return Response({
            'success': False,
            'error': f'Ошибка: {str(e)}',
            'traceback': traceback.format_exc(),
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_beward_codes_view(request):
    """
    Импорт кодов доступа Beward из Excel (второй файл).
    Колонки: № п/п | Дата выдачи | Район | АДРЕС | № под. | Нумерация квартир |
             Код доступа | IP | код для программирования ключей | примечание
    
    Данные записываются в BuildingEntrance (коды) и BewardDevice (IP + коды).
    """
    try:
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'Файл не прикреплён'}, status=400)

        uploaded = request.FILES['file']
        if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'error': 'Поддерживаются только .xlsx'}, status=400)

        import openpyxl, re, tempfile, os
        from .models import Building, BuildingEntrance

        # Сохраняем во временный файл (openpyxl требует путь к файлу)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
        finally:
            os.unlink(tmp_path)
        stats = {
            'total_rows': len(rows),
            'entrances_created': 0,
            'entrances_updated': 0,
            'devices_updated': 0,
            'no_building': 0,
            'skipped': 0,
        }

        for row in rows:
            region = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            address = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            entrance_raw = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            apartment_range = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            access_code = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            ip_addr = str(row[7]).strip() if len(row) > 7 and row[7] else ''
            programming_code = str(row[8]).strip() if len(row) > 8 and row[8] else ''
            notes = str(row[9]).strip() if len(row) > 9 and row[9] else ''
            date_issued = row[1] if len(row) > 1 else None

            if not address:
                stats['skipped'] += 1
                continue

            # Парсим адрес — улучшенный поиск улицы и дома
            addr = address

            # Ищем дом: «дом 21», «дом 8», «д. 15»
            m = re.search(r'(?:дом|д\.)\s*(\d+[а-яА-Я]?)', addr.lower())
            house = m.group(1) if m else ''

            # Ищем корпус
            m = re.search(r'корпус\s*(\d+[а-яА-Я]?)', addr.lower())
            bldg = m.group(1) if m else ''

            # Улица — ищем «улица XXX», «шоссе XXX», «проспект XXX»
            street = ''
            for pattern in [r'(?:улица|шоссе|проспект|переулок|бульвар)\s+([^,]+)', r'(?:ул\.|ш\.)\s+([^,]+)']:
                m = re.search(pattern, addr, re.IGNORECASE)
                if m:
                    street = m.group(1).strip()
                    break

            # Находим Building
            building = None
            if house:
                qs = Building.objects.filter(house_number=house)
                if street:
                    qs_street = qs.filter(street_name__icontains=street.strip())
                    if qs_street.exists():
                        qs = qs_street
                if bldg and qs.count() > 1:
                    qs_b = qs.filter(building_number=bldg)
                    if qs_b.exists():
                        qs = qs_b
                building = qs.first()

            # Парсим номер подъезда
            ent_num = None
            try:
                ent_num = int(float(entrance_raw))
            except (ValueError, TypeError):
                m = re.search(r'\d+', str(entrance_raw))
                if m:
                    ent_num = int(m.group(0))

            # Создаём/обновляем BuildingEntrance
            entrance = None
            if building and ent_num:
                entrance, created = BuildingEntrance.objects.get_or_create(
                    building=building,
                    number=ent_num,
                    defaults={'apartments_count': 0}
                )
                entrance.ip_address = ip_addr or entrance.ip_address
                entrance.access_code = access_code or entrance.access_code
                entrance.programming_code = programming_code or entrance.programming_code
                if apartment_range and '-' in apartment_range:
                    try:
                        p = apartment_range.split('-')
                        entrance.apartment_from = int(p[0].strip())
                        entrance.apartment_to = int(p[-1].strip())
                        entrance.apartments_count = entrance.apartment_to - entrance.apartment_from + 1
                    except (ValueError, IndexError):
                        pass
                if notes and notes not in (entrance.notes or ''):
                    entrance.notes = ((entrance.notes or '') + '\n' + notes).strip()
                entrance.save()
                if created:
                    stats['entrances_created'] += 1
                else:
                    stats['entrances_updated'] += 1
            elif ent_num:
                stats['no_building'] += 1

            # Обновляем BewardDevice
            if ip_addr:
                dt = None
                if date_issued and hasattr(date_issued, 'strftime'):
                    dt = timezone.make_aware(
                        date_issued.replace(tzinfo=None),
                        timezone.get_current_timezone()
                    ) if date_issued and date_issued.tzinfo is None else date_issued

                BewardDevice.objects.update_or_create(
                    ip_address=ip_addr,
                    defaults={
                        'region': region,
                        'address': address,
                        'entrance_number': str(ent_num) if ent_num else entrance_raw,
                        'access_code': access_code,
                        'programming_code': programming_code,
                        'apartment_range': apartment_range,
                        'date_issued': dt,
                        'notes': (notes or '')[:1000],
                        'building': building,
                        'entrance': entrance,
                    }
                )
                stats['devices_updated'] += 1

        return Response({
            'success': True,
            **stats,
            'message': f'Подъездов: +{stats["entrances_created"]} обновлено {stats["entrances_updated"]}, '
                       f'устройств дополнено: {stats["devices_updated"]}, '
                       f'не найдено домов: {stats["no_building"]}, пропущено: {stats["skipped"]}',
        })

    except Exception as e:
        import traceback
        return Response({
            'success': False,
            'error': f'Ошибка: {str(e)}',
            'traceback': traceback.format_exc(),
        }, status=500)


# ══════════════════════════════════════════════════════════════════
# Системная статистика, экспорт, очистка
# ══════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def erc_report_summary_view(request):
    """
    Отчёт ЕРЦ: агрегация начислений и оплат по месяцам.
    Возвращает:
      - months: список { period, charged_total, paid_total, records_count, paid_percent }
      - totals: { charged_total, paid_total, balance_start, balance_end }
    """
    from django.db.models import Sum, Count
    from datetime import date

    records = ErcBillingRecord.objects.values('period').annotate(
        charged_total=Sum('charged'),
        paid_total=Sum('paid'),
        balance_start_total=Sum('balance_start'),
        balance_end_total=Sum('balance_end'),
        records_count=Count('id'),
    ).order_by('-period')

    months = []
    for r in records[:24]:  # последние 24 месяца
        ch = float(r['charged_total'] or 0)
        pd = float(r['paid_total'] or 0)
        months.append({
            'period': r['period'].strftime('%Y-%m-%d') if r['period'] else '',
            'period_label': r['period'].strftime('%B %Y') if r['period'] else '',
            'charged_total': round(ch, 2),
            'paid_total': round(pd, 2),
            'balance_start_total': round(float(r['balance_start_total'] or 0), 2),
            'balance_end_total': round(float(r['balance_end_total'] or 0), 2),
            'records_count': r['records_count'],
            'paid_percent': round(pd / ch * 100, 1) if ch > 0 else 0,
        })

    # Итоги за всё время
    agg = ErcBillingRecord.objects.aggregate(
        total_charged=Sum('charged'),
        total_paid=Sum('paid'),
        total_balance_start=Sum('balance_start'),
        total_balance_end=Sum('balance_end'),
        total_records=Count('id'),
    )

    ch_all = float(agg['total_charged'] or 0)
    pd_all = float(agg['total_paid'] or 0)

    return Response({
        'months': months,
        'totals': {
            'charged_total': round(ch_all, 2),
            'paid_total': round(pd_all, 2),
            'balance_start_total': round(float(agg['total_balance_start'] or 0), 2),
            'balance_end_total': round(float(agg['total_balance_end'] or 0), 2),
            'records_count': agg['total_records'],
            'paid_percent': round(pd_all / ch_all * 100, 1) if ch_all > 0 else 0,
        },
        'accounts_count': ErcAccount.objects.count(),
        'clients_with_erc': Client.objects.filter(personal_account_number__isnull=False).exclude(personal_account_number='').count(),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def system_stats_view(request):
    """Возвращает статистику системы: место на диске, размер медиафайлов."""
    import os, shutil
    from django.conf import settings as django_settings

    result = {
        'disk_total_gb': 0,
        'disk_used_gb': 0,
        'disk_free_gb': 0,
        'media_count': 0,
        'media_size_mb': 0,
        'media_path': '',
        'db_size_mb': 0,
    }

    # Место на диске
    try:
        media_root = getattr(django_settings, 'MEDIA_ROOT', '/app/media')
        stat = shutil.disk_usage(os.path.dirname(str(media_root)) if media_root else '/')
        result['disk_total_gb'] = round(stat.total / (1024**3), 1)
        result['disk_used_gb'] = round(stat.used / (1024**3), 1)
        result['disk_free_gb'] = round(stat.free / (1024**3), 1)
        result['media_path'] = str(media_root)
    except Exception:
        pass

    # Размер медиафайлов
    from .models import OrderMedia
    result['media_count'] = OrderMedia.objects.count()
    try:
        total_size = 0
        for m in OrderMedia.objects.all():
            try:
                total_size += m.file.size
            except Exception:
                pass
        result['media_size_mb'] = round(total_size / (1024**2), 1)
    except Exception:
        pass

    # Размер БД (приблизительно)
    try:
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT pg_database_size(current_database())")
            db_size = cursor.fetchone()[0]
            result['db_size_mb'] = round(db_size / (1024**2), 1)
    except Exception:
        pass

    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_clients_excel_view(request):
    """Экспорт всех клиентов в Excel-файл."""
    import openpyxl, io
    from .models import Client

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    # Заголовки
    headers = ['ID', 'ФИО', 'Телефон', 'Email', 'Адрес', 'Л/счет', '№ парадной', 'УК/ТСЖ',
               'Район (мун.)', 'Источник', 'Дата добавления', 'Примечания']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Маппинг source на читаемые названия
    source_labels = {'manual': 'Ручной ввод', 'excel_import': 'Импорт (ТСЖ/УК)', 'erc': 'ЕРЦ'}

    for row_idx, c in enumerate(Client.objects.all().order_by('id'), 2):
        ws.cell(row=row_idx, column=1, value=c.id)
        ws.cell(row=row_idx, column=2, value=c.name)
        ws.cell(row=row_idx, column=3, value=c.phone)
        ws.cell(row=row_idx, column=4, value=c.email)
        ws.cell(row=row_idx, column=5, value=c.address)
        ws.cell(row=row_idx, column=6, value=c.personal_account_number or '')
        ws.cell(row=row_idx, column=7, value=c.entrance_number)
        ws.cell(row=row_idx, column=8, value=c.management_company)
        ws.cell(row=row_idx, column=9, value=c.district)
        ws.cell(row=row_idx, column=10, value=source_labels.get(c.source, c.source))
        ws.cell(row=row_idx, column=11, value=c.created_at.strftime('%d.%m.%Y %H:%M') if c.created_at else '')
        ws.cell(row=row_idx, column=12, value=c.notes)

    buf = io.BytesIO()
    wb.save(buf)

    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clients_export.xlsx'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cleanup_media_view(request):
    """Очистка медиафайлов (фото/видео отчётов) старше N дней или все."""
    from django.utils import timezone
    from .models import OrderMedia
    import os

    days = request.data.get('days')
    delete_all = request.data.get('delete_all', False)

    deleted_count = 0
    deleted_size = 0

    qs = OrderMedia.objects.all()

    if delete_all:
        pass  # удаляем все
    elif days:
        cutoff = timezone.now() - timezone.timedelta(days=int(days))
        qs = qs.filter(uploaded_at__lt=cutoff)
    else:
        return Response({'success': False, 'error': 'Укажите days или delete_all'}, status=400)

    for media in qs:
        try:
            if media.file:
                deleted_size += media.file.size
                media.file.delete(save=False)
            media.delete()
            deleted_count += 1
        except Exception as e:
            pass

    return Response({
        'success': True,
        'deleted_count': deleted_count,
        'deleted_size_mb': round(deleted_size / (1024**2), 1),
    })


# ══════════════════════════════════════════════════════════════════════════════
# Прямой перенос данных между серверами (миграция «online»)
# ══════════════════════════════════════════════════════════════════════════════
import threading

_migration_state = {
    'running': False,
    'progress': 0,
    'total': 0,
    'current_model': '',
    'current_step': '',
    'created': 0,
    'updated': 0,
    'errors': [],
    'log': [],
    'finished': False,
}
_migration_lock = threading.Lock()

# Порядок переноса моделей (зависимости: сначала справочники, потом основные)
MIGRATION_MODELS = [
    # Справочники
    {'model': 'Region', 'label': 'Регионы'},
    {'model': 'ManagementCompany', 'label': 'УК/ТСЖ'},
    {'model': 'Tariff', 'label': 'Тарифы'},
    {'model': 'Building', 'label': 'Дома'},
    {'model': 'BuildingEntrance', 'label': 'Подъезды'},
    # Клиенты и ЕРЦ
    {'model': 'Client', 'label': 'Клиенты'},
    {'model': 'ErcAccount', 'label': 'Лицевые счета'},
    {'model': 'ErcBillingRecord', 'label': 'ЕРЦ-записи'},
    {'model': 'PaymentRecord', 'label': 'Внутренние платежи'},
    # Пользователи
    {'model': 'UserProfile', 'label': 'Профили'},
    {'model': 'Master', 'label': 'Мастера'},
    # Оборудование
    {'model': 'InventoryItem', 'label': 'Оборудование'},
    {'model': 'InventoryMovement', 'label': 'Движения'},
    {'model': 'StorageLocation', 'label': 'Места хранения'},
    {'model': 'Supplier', 'label': 'Поставщики'},
    # Заявки
    {'model': 'Order', 'label': 'Заявки'},
    {'model': 'OrderHistory', 'label': 'История'},
    {'model': 'OrderMedia', 'label': 'Медиа'},
    {'model': 'Payment', 'label': 'Оплаты'},
    # Настройки
    {'model': 'SystemSettings', 'label': 'Настройки'},
    {'model': 'BewardDevice', 'label': 'Beward'},
    {'model': 'LegalEntity', 'label': 'Юрлица'},
    {'model': 'CallLog', 'label': 'Звонки'},
]

MIGRATION_MODEL_NAMES = [m['model'] for m in MIGRATION_MODELS]


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def migration_export_model_view(request):
    """
    Постраничный экспорт модели для прямого переноса.
    GET /api/system/migrate/export/?model=Region&page=1&page_size=500

    Возвращает: { model, page, total_pages, total_count, data: [...] }
    """
    from django.apps import apps
    from django.core import serializers

    model_name = request.query_params.get('model', '')
    if model_name not in MIGRATION_MODEL_NAMES:
        return Response({'error': f'Unknown model: {model_name}'}, status=400)

    try:
        Model = apps.get_model('main', model_name)
        if not Model:
            # Пробуем auth.User
            from django.contrib.auth.models import User as AuthUser
            Model = AuthUser if model_name == 'User' else None
            if not Model:
                return Response({'error': f'Model not found: {model_name}'}, status=400)
    except LookupError:
        return Response({'error': f'Model not found: {model_name}'}, status=400)

    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 500))
    total = Model.objects.count()
    total_pages = max(1, (total + page_size - 1) // page_size)

    qs = Model.objects.all().order_by('pk')
    start = (page - 1) * page_size
    end = start + page_size
    items = list(qs[start:end].values())

    # Конвертируем поля, которые не сериализуются (ForeignKey → _id)
    for item in items:
        for k, v in list(item.items()):
            if hasattr(v, 'isoformat'):
                if v is not None:
                    item[k] = v.isoformat()

    return Response({
        'model': model_name,
        'page': page,
        'total_pages': total_pages,
        'total_count': total,
        'page_size': page_size,
        'data': items,
    })


def _run_migration(source_host, source_port, username, password, sections):
    """Фоновый процесс переноса данных. Сначала авторизуется на источнике."""
    import requests as req

    global _migration_state
    with _migration_lock:
        _migration_state = {
            'running': True, 'progress': 0, 'total': 0,
            'current_model': '', 'current_step': 'Подготовка...',
            'created': 0, 'updated': 0, 'errors': [], 'log': [], 'finished': False,
        }

    source_base = f'http://{source_host}:{source_port}'
    source_url = f'{source_base}/api/system/migrate/export/'

    # ── Авторизация на сервере-источнике ──
    token = None
    try:
        with _migration_lock:
            _migration_state['current_step'] = 'Авторизация на источнике...'
        login_resp = req.post(
            f'{source_base}/api/auth/login/',
            json={'username': username, 'password': password},
            timeout=10,
        )
        if login_resp.status_code == 200:
            token = login_resp.json().get('token')
        if not token:
            with _migration_lock:
                _migration_state['running'] = False
                _migration_state['finished'] = True
                _migration_state['errors'].append('Не удалось авторизоваться на сервере-источнике. Проверьте логин/пароль.')
                _migration_state['current_step'] = 'Ошибка авторизации'
            return
    except Exception as e:
        with _migration_lock:
            _migration_state['running'] = False
            _migration_state['finished'] = True
            _migration_state['errors'].append(f'Ошибка подключения к источнику: {str(e)[:100]}')
            _migration_state['current_step'] = 'Ошибка подключения'
        return

    headers = {'Authorization': f'Token {token}'}

    # Подсчёт общего количества записей
    all_models = [m for m in MIGRATION_MODELS if m['model'] in sections or 'all' in sections]
    total_items = 0
    for m_cfg in all_models:
        try:
            r = req.get(source_url, params={'model': m_cfg['model'], 'page': 1, 'page_size': 1}, headers=headers, timeout=10)
            if r.status_code == 200:
                total_items += r.json().get('total_count', 0)
        except Exception:
            pass

    with _migration_lock:
        _migration_state['total'] = total_items

    created_total = 0

    for m_cfg in all_models:
        model_name = m_cfg['model']
        label = m_cfg['label']

        with _migration_lock:
            _migration_state['current_model'] = label
            _migration_state['current_step'] = f'Загрузка {label}...'

        page = 1
        while True:
            try:
                r = req.get(source_url, params={'model': model_name, 'page': page, 'page_size': 500}, headers=headers, timeout=30)
                if r.status_code != 200:
                    with _migration_lock:
                        _migration_state['errors'].append(f'{label}: HTTP {r.status_code}')
                        _migration_state['log'].append(f'✗ {label}: HTTP {r.status_code}')
                    break
                data = r.json()
                items = data.get('data', [])
                if not items:
                    break

                page += 1
                from django.apps import apps
                try:
                    Model = apps.get_model('main', model_name)
                except LookupError:
                    with _migration_lock:
                        _migration_state['errors'].append(f'{label}: модель не найдена')
                    break

                for item in items:
                    pk = item.pop('id', None)
                    try:
                        if pk and Model.objects.filter(pk=pk).exists():
                            Model.objects.filter(pk=pk).update(**item)
                            created_total += 1
                        else:
                            obj = Model(**item)
                            if pk:
                                obj.id = pk
                            obj.save(force_insert=True)
                            created_total += 1
                    except Exception as e:
                        with _migration_lock:
                            err = str(e)[:100]
                            _migration_state['errors'].append(f'{label} pk={pk}: {err}')

                with _migration_lock:
                    _migration_state['progress'] += len(items)
                    _migration_state['created'] = created_total
                    _migration_state['current_step'] = f'{label}: {_migration_state["progress"]}/{total_items}'
                    _migration_state['log'].append(f'✓ {label}: +{len(items)} (стр. {page - 1})')

                if data.get('page', 0) >= data.get('total_pages', 1):
                    break
            except Exception as e:
                with _migration_lock:
                    _migration_state['errors'].append(f'{label}: {str(e)[:100]}')
                break

    with _migration_lock:
        _migration_state['running'] = False
        _migration_state['finished'] = True
        _migration_state['current_step'] = 'Перенос завершён'


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def migration_start_view(request):
    """
    Запуск прямого переноса данных с другого сервера.
    POST /api/system/migrate/start/
    Body: { host, port, username, password, sections }
    """
    host = request.data.get('host', '')
    port = request.data.get('port', '8000')
    username = request.data.get('username', 'admin')
    password = request.data.get('password', 'admin123')
    sections = request.data.get('sections', ['all'])

    if not host:
        return Response({'success': False, 'error': 'Укажите IP-адрес сервера-источника'}, status=400)

    global _migration_state
    with _migration_lock:
        if _migration_state.get('running'):
            return Response({'success': False, 'error': 'Перенос уже выполняется'}, status=400)

    thread = threading.Thread(target=_run_migration, args=(host, port, username, password, sections), daemon=True)
    thread.start()

    return Response({'success': True, 'message': 'Перенос запущен'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def migration_status_view(request):
    """Статус текущего переноса."""
    global _migration_state
    with _migration_lock:
        return Response({**_migration_state})
